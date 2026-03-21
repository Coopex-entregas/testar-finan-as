from __future__ import annotations

# ============ Stdlib ============
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
import os, io, csv, re, json, time, difflib, unicodedata
from datetime import datetime, date, timedelta, time as dtime
from collections import defaultdict, namedtuple
import uuid
from functools import wraps
from types import SimpleNamespace

# MIME types (fixes p/ Office)
import mimetypes
mimetypes.add_type("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx")
mimetypes.add_type("application/vnd.ms-excel", ".xls")
mimetypes.add_type("application/vnd.openxmlformats-officedocument.wordprocessingml.document", ".docx")
mimetypes.add_type("application/msword", ".doc")
mimetypes.add_type("application/vnd.openxmlformats-officedocument.presentationml.presentation", ".pptx")
mimetypes.add_type("application/vnd.ms-powerpoint", ".ppt")

# ============ Terceiros ============
from flask import (
    Flask, render_template, request, redirect, url_for, session,
    flash, send_file, abort, jsonify, current_app
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from dateutil.relativedelta import relativedelta

# ✅ FALTAVA ISSO (resolve o erro do UserMixin e já deixa pronto p/ login)
from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    login_required,
    logout_user,
    current_user,
)

from sqlalchemy import text
from sqlalchemy import func, text as sa_text, or_, and_, case
from sqlalchemy import func, text as sa_text, or_, and_, case, literal
from sqlalchemy.inspection import inspect as sa_inspect
from sqlalchemy.pool import QueuePool
from sqlalchemy import event
from sqlalchemy.engine import Engine
from sqlalchemy.exc import OperationalError, SQLAlchemyError, IntegrityError, DisconnectionError
from sqlalchemy import delete as sa_delete

# 👉 Novo: para gerar XLSX em memória
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse, parse_qsl, urlencode, urlunparse

# ============ App / Diretórios ============
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

UPLOAD_DIR = os.path.join(BASE_DIR, "static", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Legado / compatibilidade
DOCS_DIR = os.path.join(UPLOAD_DIR, "docs")
os.makedirs(DOCS_DIR, exist_ok=True)

STATIC_TABLES = os.path.join(UPLOAD_DIR, "tabelas")
os.makedirs(STATIC_TABLES, exist_ok=True)

# Persistência real (Render Disk)
PERSIST_ROOT = os.environ.get("PERSIST_ROOT", "/var/data")
if not os.path.isdir(PERSIST_ROOT):
    PERSIST_ROOT = os.path.join(BASE_DIR, "data")
os.makedirs(PERSIST_ROOT, exist_ok=True)

# Tabelas em disco persistente
TABELAS_DIR = os.path.join(PERSIST_ROOT, "tabelas")
os.makedirs(TABELAS_DIR, exist_ok=True)

# Documentos em disco persistente
DOCS_PERSIST_DIR = os.path.join(PERSIST_ROOT, "docs")
os.makedirs(DOCS_PERSIST_DIR, exist_ok=True)

def _merge_qs(url: str, extra: dict[str, str]) -> str:
    """Insere parâmetros de query no URI sem duplicar os já existentes."""
    p = urlparse(url)
    q = dict(parse_qsl(p.query, keep_blank_values=True))
    for k, v in (extra or {}).items():
        q.setdefault(k, v)
    return urlunparse(p._replace(query=urlencode(q, doseq=True)))


def _build_db_uri() -> str:
    raw = os.environ.get("DATABASE_URL")
    if not raw:
        return "sqlite:///" + os.path.join(BASE_DIR, "app.db")

    # força driver psycopg3 (SQLAlchemy)
    if raw.startswith("postgres://"):
        raw = raw.replace("postgres://", "postgresql+psycopg://", 1)
    elif raw.startswith("postgresql://") and "+psycopg" not in raw:
        raw = raw.replace("postgresql://", "postgresql+psycopg://", 1)

    # SSL + keepalive + app name via libpq (idempotente)
    extras = {
        "sslmode": "require",
        "keepalives": "1",
        "keepalives_idle": "30",   # segundos ocioso antes de mandar keepalive
        "keepalives_interval": "10",
        "keepalives_count": "3",
        "application_name": os.environ.get("APP_NAME", "financas-dxsu"),
    }
    return _merge_qs(raw, extras)


app = Flask(__name__, static_folder="static", template_folder="templates")

app.secret_key = os.environ.get("SECRET_KEY", "coopex-secret")

URI = _build_db_uri()

# 🚫 Guard: impede cair em SQLite em produção se DATABASE_URL não existir
if "sqlite" in URI and os.environ.get("FLASK_ENV") == "production":
    raise RuntimeError("DATABASE_URL ausente em produção")

# ===== Pool dimensionado por worker =====
workers = int(os.environ.get("WEB_CONCURRENCY", "1") or "1")
threads = int(os.environ.get("GTHREADS", "1") or "1")  # se não usar threads, fica 1
req_concurrency = max(1, workers * threads)

# alvo global de conexões por instância (ajustável via env)
# 40 é bem suficiente para ~70 pessoas usando o sistema sem estourar limite do Postgres
target_total = int(os.environ.get("DB_TARGET_CONNS", "40") or "40")

# por worker, limitado para não exagerar:
per_worker_target = max(5, min(target_total // max(1, workers), 15))

default_pool_size = min(per_worker_target, req_concurrency + 2)
default_max_overflow = max(5, default_pool_size)

pool_size = int(os.environ.get("DB_POOL_SIZE", str(default_pool_size)))
max_overflow = int(os.environ.get("DB_MAX_OVERFLOW", str(default_max_overflow)))
pool_recycle = int(os.environ.get("SQL_POOL_RECYCLE", "240"))  # 4 min < timeout de idle do provider
pool_timeout = int(os.environ.get("SQL_POOL_TIMEOUT", "20"))   # espera máx. por conexão do pool (s)

app.config.update(
    SQLALCHEMY_DATABASE_URI=URI,
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    JSON_SORT_KEYS=False,
    MAX_CONTENT_LENGTH=32 * 1024 * 1024,  # 32MB
    SESSION_COOKIE_HTTPONLY=True,
    REMEMBER_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("FLASK_SECURE_COOKIES", "1") == "1",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    SQLALCHEMY_ENGINE_OPTIONS={
        "poolclass": QueuePool,
        "pool_size": pool_size,
        "max_overflow": max_overflow,
        "pool_timeout": pool_timeout,
        "pool_pre_ping": True,       # testa conexão antes de usar (evita usar conn morta)
        "pool_use_lifo": True,       # reduz churn de conexões sob carga
        "pool_recycle": pool_recycle,  # recicla sockets ociosos (evita timeout de idle / TLS)
        "connect_args": {
            # tempo máx. para abrir conexão
            "connect_timeout": int(os.getenv("PGCONNECT_TIMEOUT", "5")),
            # tempo máx. por statement no servidor (defensivo)
            "options": "-c statement_timeout=15000",
        },
    },
)

db = SQLAlchemy(app)

def _sso_serializer():
    secret = os.environ.get("SSO_SHARED_SECRET") or app.secret_key
    # "salt" separa o token SSO de outros usos do secret
    return URLSafeTimedSerializer(secret_key=secret, salt="coopex-sso-v1")

def sso_load(token: str, max_age_seconds: int = 45) -> dict:
    s = _sso_serializer()
    return s.loads(token, max_age=max_age_seconds)

def _get_or_create_sso_user(tipo: str = "admin") -> Usuario:
    """
    Garante um usuário técnico para sessão SSO, evitando quebrar rotas que consultam Usuario.
    """
    username = f"sso_{tipo}"
    u = Usuario.query.filter_by(usuario=username).first()
    if u:
        return u

    # cria user técnico sem senha (não loga pelo /login)
    u = Usuario(usuario=username, tipo=tipo, senha_hash="!")
    db.session.add(u)
    db.session.commit()
    return u
    

def ajustar_banco():
    try:
        # se você tiver essa função _is_sqlite, use; senão pode checar pela URI
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(despesas_cooperado);")).fetchall()
            colnames = {row[1] for row in cols}
            if "eh_adiantamento" not in colnames:
                db.session.execute(sa_text(
                    "ALTER TABLE despesas_cooperado ADD COLUMN eh_adiantamento BOOLEAN DEFAULT 0"
                ))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                ALTER TABLE IF EXISTS public.despesas_cooperado
                ADD COLUMN IF NOT EXISTS eh_adiantamento BOOLEAN DEFAULT FALSE
            """))
            db.session.commit()
    except Exception:
        db.session.rollback()


# ========= Retry de conexão p/ rotas críticas =========
def with_db_retry(fn):
    """
    Decorator simples para repetir a operação 1x em caso de queda de conexão
    (OperationalError / DisconnectionError). Usa junto de rotas que batem no banco.
    """
    @wraps(fn)
    def wrapper(*args, **kwargs):
        last_exc = None
        for attempt in range(2):  # tentativa 0 e 1
            try:
                return fn(*args, **kwargs)
            except (OperationalError, DisconnectionError) as e:
                last_exc = e
                # rollback defensivo e pequeno backoff
                db.session.rollback()
                try:
                    app.logger.warning(f"[DB_RETRY] tentativa {attempt+1} falhou: {e}")
                except Exception:
                    pass
                time.sleep(0.2)
        # se chegou aqui, estourou as tentativas
        raise last_exc
    return wrapper


# Health checks
@app.get("/healthz")
def healthz():
    return "ok", 200


@app.get("/readyz")
def readyz():
    try:
        db.session.execute(sa_text("SELECT 1"))
        return "ready", 200
    except Exception:
        return "not-ready", 503


# Liga foreign_keys no SQLite
@event.listens_for(Engine, "connect")
def _set_sqlite_pragma(dbapi_con, con_record):
    try:
        if app.config["SQLALCHEMY_DATABASE_URI"].startswith("sqlite"):
            cur = dbapi_con.cursor()
            cur.execute("PRAGMA foreign_keys=ON")
            cur.close()
    except Exception:
        pass


def _is_sqlite() -> bool:
    try:
        return db.session.get_bind().dialect.name == "sqlite"
    except Exception:
        return "sqlite" in (app.config.get("SQLALCHEMY_DATABASE_URI") or "")



# ==========================================
# MODELOS
# ==========================================

class Usuario(db.Model, UserMixin):
    __tablename__ = "usuarios"

    id = db.Column(db.Integer, primary_key=True)
    usuario = db.Column(db.String(80), unique=True, nullable=False)
    nome = db.Column(db.String(120))
    senha_hash = db.Column(db.String(200), nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # admin | cooperado | restaurante

    is_master = db.Column(
        db.Boolean,
        nullable=False,
        default=False,
        server_default=text("false")
    )

    ativo = db.Column(
        db.Boolean,
        nullable=False,
        default=True,
        server_default=text("true")
    )

    @property
    def is_active(self) -> bool:
        return bool(self.ativo is not False)

    def set_password(self, raw: str):
        self.senha_hash = generate_password_hash(raw)

    def check_password(self, raw: str) -> bool:
        return check_password_hash(self.senha_hash, raw)
        
class AdminPermissao(db.Model):
    __tablename__ = "admin_permissoes"

    id = db.Column(db.Integer, primary_key=True)

    usuario_id = db.Column(
        db.Integer,
        db.ForeignKey("usuarios.id", ondelete="CASCADE"),
        nullable=False,
        index=True
    )

    aba = db.Column(db.String(50), nullable=False)

    pode_ver = db.Column(db.Boolean, default=False)
    pode_criar = db.Column(db.Boolean, default=False)
    pode_editar = db.Column(db.Boolean, default=False)
    pode_excluir = db.Column(db.Boolean, default=False)

    usuario = db.relationship(
        "Usuario",
        backref=db.backref("permissoes_admin", cascade="all, delete-orphan")
    )

    __table_args__ = (
        db.UniqueConstraint("usuario_id", "aba", name="uq_admin_permissao_usuario_aba"),
    )

class Cooperado(db.Model):
    __tablename__ = "cooperados"

    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(120), nullable=False)

    usuario_id = db.Column(db.Integer, db.ForeignKey("usuarios.id"), nullable=False, unique=True)

    # 1 usuário -> 1 cooperado
    usuario_ref = db.relationship("Usuario", backref=db.backref("coop_account", uselist=False))

    # NOVO
    telefone = db.Column(db.String(30))

    # Foto no banco
    foto_bytes = db.Column(db.LargeBinary)
    foto_mime = db.Column(db.String(100))
    foto_filename = db.Column(db.String(255))
    foto_url = db.Column(db.String(255))

    cnh_numero = db.Column(db.String(50))
    cnh_validade = db.Column(db.Date)

    placa = db.Column(db.String(20))
    placa_validade = db.Column(db.Date)

    ultima_atualizacao = db.Column(db.DateTime)


class Restaurante(db.Model):
    __tablename__ = "restaurantes"

    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(120), nullable=False)
    periodo = db.Column(db.String(20), nullable=False)  # seg-dom | sab-sex | sex-qui

    usuario_id = db.Column(db.Integer, db.ForeignKey("usuarios.id"), nullable=False, unique=True)

    # 1 usuário -> 1 restaurante
    usuario_ref = db.relationship("Usuario", backref=db.backref("rest_account", uselist=False))

    # Taxa administrativa
    taxa_admin_valor = db.Column(db.Float, default=0.0)
    taxa_admin_data_base = db.Column(db.Date)
    taxa_admin_multa_percentual = db.Column(db.Float, default=2.0)
    taxa_admin_juros_dia_percentual = db.Column(db.Float, default=0.03)
    ativo = db.Column(db.Boolean, default=True)

    # Foto no banco (bytea)
    foto_bytes = db.Column(db.LargeBinary)
    foto_mime = db.Column(db.String(100))
    foto_filename = db.Column(db.String(255))

    # compatibilidade
    foto_url = db.Column(db.String(255))


class Lancamento(db.Model):
    __tablename__ = "lancamentos"

    id = db.Column(db.Integer, primary_key=True)

    restaurante_id = db.Column(db.Integer, db.ForeignKey("restaurantes.id"), nullable=False)
    cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"), nullable=False)

    restaurante = db.relationship("Restaurante")
    cooperado = db.relationship("Cooperado")

    descricao = db.Column(db.String(200))
    valor = db.Column(db.Float, default=0.0)

    # Se sempre precisa de data, deixe nullable=False
    data = db.Column(db.Date, nullable=False)

    hora_inicio = db.Column(db.String(10))
    hora_fim = db.Column(db.String(10))

    # opcional: quantidade de entregas
    qtd_entregas = db.Column(db.Integer)


# === AVALIAÇÕES DE COOPERADO (NOVO) =========================================
class AvaliacaoCooperado(db.Model):
    __tablename__ = "avaliacoes"
    id = db.Column(db.Integer, primary_key=True)

    restaurante_id = db.Column(db.Integer, db.ForeignKey("restaurantes.id"), nullable=False)
    cooperado_id   = db.Column(db.Integer, db.ForeignKey("cooperados.id"),  nullable=False)

    # 🔴 IMPORTANTE: CASCADE na FK para o lançamento
    lancamento_id  = db.Column(
        db.Integer,
        db.ForeignKey("lancamentos.id", ondelete="CASCADE"),
        nullable=True  # deixe True para permitir trocar para SET NULL futuramente, se quiser
    )

    # notas 1..5
    estrelas_geral         = db.Column(db.Integer)
    estrelas_pontualidade  = db.Column(db.Integer)
    estrelas_educacao      = db.Column(db.Integer)
    estrelas_eficiencia    = db.Column(db.Integer)
    estrelas_apresentacao  = db.Column(db.Integer)  # "Bem apresentado"

    comentario       = db.Column(db.Text)

    # IA/heurísticas
    media_ponderada  = db.Column(db.Float)
    sentimento       = db.Column(db.String(12))     # positivo | neutro | negativo
    temas            = db.Column(db.String(255))    # palavras-chave resumidas
    alerta_crise     = db.Column(db.Boolean, default=False)
    feedback_motoboy = db.Column(db.Text)

    criado_em = db.Column(db.DateTime, default=datetime.utcnow)


class ReceitaCooperativa(db.Model):
    __tablename__ = "receitas_coop"
    id = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(200), nullable=False)
    valor_total = db.Column(db.Float, default=0.0)
    data = db.Column(db.Date, nullable=True)

    restaurante_id = db.Column(db.Integer, db.ForeignKey("restaurantes.id"), nullable=True, index=True)
    restaurante = db.relationship("Restaurante")

    auto_taxa_adm = db.Column(db.Boolean, default=False)
    competencia = db.Column(db.String(7))
    valor_previsto = db.Column(db.Float, default=0.0)
    valor_principal = db.Column(db.Float, default=0.0)
    valor_pago = db.Column(db.Float, default=0.0)
    valor_multa = db.Column(db.Float, default=0.0)
    valor_juros = db.Column(db.Float, default=0.0)
    data_vencimento = db.Column(db.Date)
    data_pagamento = db.Column(db.Date)
    status_pagamento = db.Column(db.String(20), default="nao_pago")
    multa_percentual = db.Column(db.Float, default=2.0)
    juros_dia_percentual = db.Column(db.Float, default=0.03)


class DespesaCooperativa(db.Model):
    __tablename__ = "despesas_coop"
    id = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, default=0.0)
    data = db.Column(db.Date)


class ReceitaCooperado(db.Model):
    __tablename__ = "receitas_cooperado"
    id = db.Column(db.Integer, primary_key=True)
    cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"), nullable=False)
    cooperado = db.relationship("Cooperado")
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, default=0.0)
    data = db.Column(db.Date)


# =========================
# Models
# =========================
class DespesaCooperado(db.Model):
    __tablename__ = "despesas_cooperado"
    id = db.Column(db.Integer, primary_key=True)
    cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"), nullable=True)  # None = Todos
    cooperado = db.relationship("Cooperado")
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, default=0.0)
    # legado (pontual)
    data = db.Column(db.Date)

    # novo (período)
    data_inicio = db.Column(db.Date)
    data_fim    = db.Column(db.Date)

    # NOVO: vínculo forte para sabermos quem gerou a despesa
    beneficio_id = db.Column(
        db.Integer,
        db.ForeignKey("beneficios_registro.id", ondelete="CASCADE"),
        index=True,
        nullable=True  # deixa True para migrar suave
    )

    # 🔴 NOVO: marca se é adiantamento
    eh_adiantamento = db.Column(db.Boolean, default=False)
    competencia_desconto = db.Column(db.String(20), default='atual')
    


class DespesaCooperadoAbatimento(db.Model):
    __tablename__ = "despesas_cooperado_abatimentos"
    id = db.Column(db.Integer, primary_key=True)
    despesa_id = db.Column(db.Integer, db.ForeignKey("despesas_cooperado.id", ondelete="CASCADE"), nullable=False, index=True)
    data = db.Column(db.Date, nullable=False, default=date.today)
    valor = db.Column(db.Float, default=0.0)
    origem = db.Column(db.String(30), default="manual")
    observacao = db.Column(db.String(255))
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)

    despesa = db.relationship(
        "DespesaCooperado",
        backref=db.backref("abatimentos", cascade="all, delete-orphan", order_by="DespesaCooperadoAbatimento.data.desc(), DespesaCooperadoAbatimento.id.desc()")
    )


class BeneficioRegistro(db.Model):
    __tablename__ = "beneficios_registro"
    id = db.Column(db.Integer, primary_key=True)
    data_inicial = db.Column(db.Date, nullable=False)
    data_final = db.Column(db.Date, nullable=False)
    data_lancamento = db.Column(db.Date)
    tipo = db.Column(db.String(40), nullable=False)  # hospitalar | farmaceutico | alimentar
    valor_total = db.Column(db.Float, default=0.0)
    recebedores_nomes = db.Column(db.Text)  # nomes separados por ';'
    recebedores_ids = db.Column(db.Text)    # ids separados por ';'

    # NOVO: relacionamento reverso (não remove nada seu)
    despesas = db.relationship(
        "DespesaCooperado",
        primaryjoin="BeneficioRegistro.id==DespesaCooperado.beneficio_id",
        cascade="all, delete-orphan",
        passive_deletes=True
    )

# =========================
# Semana seg→dom + Normalização automática
# =========================
from sqlalchemy import event

def semana_bounds(d: date):
    """
    Para uma data 'd', retorna (segunda, domingo) da mesma semana.
    0=segunda .. 6=domingo
    """
    dow = d.weekday()
    ini = d - timedelta(days=dow)     # segunda
    fim = ini + timedelta(days=6)     # domingo
    return ini, fim

def normaliza_periodo(data: date|None, data_inicio: date|None, data_fim: date|None):
    """
    Regra global para DespesaCooperado:
      - data_inicio = segunda
      - data_fim    = domingo
      - data        = domingo (== data_fim)
    Se vier só 'data', deriva o período pela semana dessa data.
    Se vier início/fim, ajusta para seg/dom da(s) semana(s) informada(s).
    """
    if data_inicio and data_fim:
        # garante ordem e aplica bordas semanais
        if data_fim < data_inicio:
            data_inicio, data_fim = data_fim, data_inicio
        ini, _ = semana_bounds(data_inicio)
        _, fim = semana_bounds(data_fim)
        return fim, ini, fim

    base = data or date.today()
    ini, fim = semana_bounds(base)
    return fim, ini, fim

# === Normalização automática de período em DespesaCooperado ===
def _ajusta_semana(target: DespesaCooperado):
    d, di, df = normaliza_periodo(target.data, target.data_inicio, target.data_fim)
    target.data = d
    target.data_inicio = di
    target.data_fim = df

@event.listens_for(DespesaCooperado, "before_insert")
def _desp_before_insert(mapper, connection, target):
    _ajusta_semana(target)

@event.listens_for(DespesaCooperado, "before_update")
def _desp_before_update(mapper, connection, target):
    _ajusta_semana(target)



class Escala(db.Model):
    __tablename__ = "escalas"
    id = db.Column(db.Integer, primary_key=True)
    cooperado_id = db.Column(
        db.Integer,
        db.ForeignKey("cooperados.id", ondelete="CASCADE"),
        nullable=True  # pode não ter cadastro
    )
    restaurante_id = db.Column(
        db.Integer,
        db.ForeignKey("restaurantes.id", ondelete="CASCADE"),
        nullable=True
    )

    data = db.Column(db.String(40))
    turno = db.Column(db.String(50))
    horario = db.Column(db.String(50))
    contrato = db.Column(db.String(80))
    cor = db.Column(db.String(200))
    cooperado_nome = db.Column(db.String(120))  # nome bruto da planilha quando não há cadastro


class TrocaSolicitacao(db.Model):
    __tablename__ = "trocas"
    id = db.Column(db.Integer, primary_key=True)
    solicitante_id = db.Column(
        db.Integer,
        db.ForeignKey("cooperados.id", ondelete="CASCADE"),
        nullable=False
    )
    destino_id = db.Column(
        db.Integer,
        db.ForeignKey("cooperados.id", ondelete="CASCADE"),
        nullable=False
    )
    origem_escala_id = db.Column(
        db.Integer,
        db.ForeignKey("escalas.id", ondelete="CASCADE"),
        nullable=False
    )
    mensagem = db.Column(db.Text)
    status = db.Column(db.String(20), default="pendente")
    criada_em = db.Column(db.DateTime, default=datetime.utcnow)
    aplicada_em = db.Column(db.DateTime)


class EscalaHistorico(db.Model):
    __tablename__ = "escalas_historico"
    id = db.Column(db.Integer, primary_key=True)
    grupo_ref = db.Column(db.String(40), index=True)
    origem = db.Column(db.String(30), index=True)  # upload, edicao_manual, troca_aprovada, passagem_aprovada
    acao = db.Column(db.String(30), index=True)    # snapshot, substituicao, troca, passagem
    escala_ref_id = db.Column(db.Integer, index=True)
    troca_ref_id = db.Column(db.Integer, index=True)
    admin_usuario_id = db.Column(db.Integer, nullable=True)
    data = db.Column(db.String(40), index=True)
    turno = db.Column(db.String(50), index=True)
    horario = db.Column(db.String(50))
    contrato = db.Column(db.String(80), index=True)
    cooperado_id = db.Column(db.Integer, nullable=True, index=True)
    cooperado_nome = db.Column(db.String(120))
    saiu_nome = db.Column(db.String(120))
    entrou_nome = db.Column(db.String(120))
    snapshot_em = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class TrocaHistorico(db.Model):
    __tablename__ = "trocas_historico"
    id = db.Column(db.Integer, primary_key=True)
    troca_ref_id = db.Column(db.Integer, index=True)
    tipo = db.Column(db.String(20), index=True)  # troca ou passagem
    solicitante_id = db.Column(db.Integer, nullable=True, index=True)
    solicitante_nome = db.Column(db.String(120))
    destino_id = db.Column(db.Integer, nullable=True, index=True)
    destino_nome = db.Column(db.String(120))
    data = db.Column(db.String(40), index=True)
    turno = db.Column(db.String(50), index=True)
    horario = db.Column(db.String(50))
    contrato = db.Column(db.String(80), index=True)
    saiu_nome = db.Column(db.String(120))
    entrou_nome = db.Column(db.String(120))
    aplicada_em = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class Config(db.Model):
    __tablename__ = "config"
    id = db.Column(db.Integer, primary_key=True)
    salario_minimo = db.Column(db.Float, default=0.0)


class Documento(db.Model):
    __tablename__ = "documentos"
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    categoria = db.Column(db.String(40))
    descricao = db.Column(db.String(255))
    arquivo_url = db.Column(db.String(255), nullable=False)
    arquivo_nome = db.Column(db.String(255))
    enviado_em = db.Column(db.DateTime, default=datetime.utcnow)


class Tabela(db.Model):
    __tablename__ = "tabelas"
    
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    categoria = db.Column(db.String(40))
    descricao = db.Column(db.String(255))
    arquivo_url = db.Column(db.String(255), nullable=False)
    arquivo_nome = db.Column(db.String(255))
    enviado_em = db.Column(db.DateTime, default=datetime.utcnow)


# ---------- AVISOS (NOVO) ----------
aviso_restaurantes = db.Table(
    "aviso_restaurantes",
    db.Column("aviso_id", db.Integer, db.ForeignKey("avisos.id"), primary_key=True),
    db.Column("restaurante_id", db.Integer, db.ForeignKey("restaurantes.id"), primary_key=True),
)

aviso_cooperados = db.Table(
    "aviso_cooperados",
    db.Column("aviso_id", db.Integer, db.ForeignKey("avisos.id"), primary_key=True),
    db.Column("cooperado_id", db.Integer, db.ForeignKey("cooperados.id"), primary_key=True),
)

class Aviso(db.Model):
    __tablename__ = "avisos"
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(140), nullable=False)
    corpo = db.Column(db.Text, nullable=False)
    # escopo: global | restaurante | cooperado
    tipo = db.Column(db.String(20), nullable=False, default="global")

    # destino individual (opcional / legado)
    destino_cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"))
    destino_cooperado = db.relationship("Cooperado", foreign_keys=[destino_cooperado_id])

    # destinos por lista (N:N)
    cooperados = db.relationship("Cooperado", secondary=aviso_cooperados, backref="avisos_recebidos")
    restaurantes = db.relationship("Restaurante", secondary=aviso_restaurantes, backref="avisos")

    prioridade = db.Column(db.String(10), default="normal")  # normal | alta
    fixado = db.Column(db.Boolean, default=False)
    ativo = db.Column(db.Boolean, default=True)
    inicio_em = db.Column(db.DateTime)  # janela de exibição opcional
    fim_em = db.Column(db.DateTime)

    criado_por_id = db.Column(db.Integer, db.ForeignKey("usuarios.id"), nullable=False)
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)
    atualizado_em = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class AvisoLeitura(db.Model):
    __tablename__ = "aviso_leituras"
    id = db.Column(db.Integer, primary_key=True)
    aviso_id = db.Column(db.Integer, db.ForeignKey("avisos.id"), nullable=False, index=True)
    cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"), nullable=True, index=True)
    restaurante_id = db.Column(db.Integer, db.ForeignKey("restaurantes.id"), nullable=True, index=True)
    lido_em = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint("aviso_id", "cooperado_id", "restaurante_id", name="uq_aviso_dest"), )

# =========================
# Helpers
# =========================
def _is_sqlite() -> bool:
    try:
        return db.session.get_bind().dialect.name == "sqlite"
    except Exception:
        return "sqlite" in (app.config.get("SQLALCHEMY_DATABASE_URI") or "")


from sqlalchemy.exc import InvalidRequestError

# Se não existir, define _is_sqlite() de forma segura
try:
    _is_sqlite  # type: ignore
except NameError:
    def _is_sqlite() -> bool:
        try:
            return db.engine.url.get_backend_name() == "sqlite"
        except Exception:
            return "sqlite" in str(db.engine.url).lower()

# ===== Alíquotas (parametrizáveis por env) =====
# Por padrão:
# INSS = 4% (0.04)
# SEST/SENAT = 0,5% (0.005)
INSS_ALIQ = float(os.environ.get("ALIQUOTA_INSS", "0.04"))
SEST_ALIQ = float(os.environ.get("ALIQUOTA_SEST", "0.005"))

def calc_descontos(valor: float) -> dict:
    """
    Calcula descontos padronizados (2 casas) e retorna um dicionário:
      - inss: 4%
      - sest: 0,5%
      - encargos: total (INSS + SEST)
      - liquido: valor - encargos
    """
    v = float(valor or 0.0)
    inss = round(v * INSS_ALIQ, 2)
    sest = round(v * SEST_ALIQ, 2)
    encargos = round(inss + sest, 2)
    liquido = round(v - encargos, 2)
    return {
        "inss": inss,
        "sest": sest,
        "encargos": encargos,
        "liquido": liquido,
    }


# =========================
# Init DB / Migração leve
# =========================
def init_db():
    """
    Versão unificada e idempotente:
      1) Ajustes de performance para SQLite (WAL/synchronous)
      2) Criação de todas as tabelas (create_all)
      3) Índices úteis (cooperado/restaurante/criado_em)
      4) Migrações leves (qtd_entregas, ativo em usuarios, colunas de escalas, fotos, tabela avaliacoes_restaurante)
      5) Bootstrap mínimo (admin e config) — só se os modelos existirem
    """

    # 1) Perf no SQLite
    try:
        if _is_sqlite():
            db.session.execute(sa_text("PRAGMA journal_mode=WAL;"))
            db.session.execute(sa_text("PRAGMA synchronous=NORMAL;"))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 2) Tabelas/mapeamentos
    try:
        db.create_all()
    except Exception:
        db.session.rollback()

    # 3) Índices de performance (idempotentes)
    try:
        if _is_sqlite():
            stmts = [
                "CREATE INDEX IF NOT EXISTS ix_avaliacoes_criado_em   ON avaliacoes (criado_em)",
                "CREATE INDEX IF NOT EXISTS ix_avaliacoes_rest_criado ON avaliacoes (restaurante_id, criado_em)",
                "CREATE INDEX IF NOT EXISTS ix_avaliacoes_coop_criado ON avaliacoes (cooperado_id,  criado_em)",

                "CREATE INDEX IF NOT EXISTS ix_av_rest_criado_em      ON avaliacoes_restaurante (criado_em)",
                "CREATE INDEX IF NOT EXISTS ix_av_rest_rest_criado    ON avaliacoes_restaurante (restaurante_id, criado_em)",
                "CREATE INDEX IF NOT EXISTS ix_av_rest_coop_criado    ON avaliacoes_restaurante (cooperado_id,  criado_em)",
            ]
        else:
            stmts = [
                "CREATE INDEX IF NOT EXISTS ix_avaliacoes_criado_em   ON public.avaliacoes (criado_em)",
                "CREATE INDEX IF NOT EXISTS ix_avaliacoes_rest_criado ON public.avaliacoes (restaurante_id, criado_em)",
                "CREATE INDEX IF NOT EXISTS ix_avaliacoes_coop_criado ON public.avaliacoes (cooperado_id,  criado_em)",

                "CREATE INDEX IF NOT EXISTS ix_av_rest_criado_em      ON public.avaliacoes_restaurante (criado_em)",
                "CREATE INDEX IF NOT EXISTS ix_av_rest_rest_criado    ON public.avaliacoes_restaurante (restaurante_id, criado_em)",
                "CREATE INDEX IF NOT EXISTS ix_av_rest_coop_criado    ON public.avaliacoes_restaurante (cooperado_id,  criado_em)",
            ]

        for sql in stmts:
            db.session.execute(sa_text(sql))
        db.session.commit()
    except Exception:
        db.session.rollback()

    # 3.9) is_master em usuarios
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(usuarios);")).fetchall()
            colnames = {row[1] for row in cols}
            if "is_master" not in colnames:
                db.session.execute(sa_text("ALTER TABLE usuarios ADD COLUMN is_master BOOLEAN DEFAULT 0"))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                ALTER TABLE IF EXISTS public.usuarios
                ADD COLUMN IF NOT EXISTS is_master BOOLEAN DEFAULT FALSE
            """))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 3.95) nome em usuarios
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(usuarios);")).fetchall()
            colnames = {row[1] for row in cols}
            if "nome" not in colnames:
                db.session.execute(sa_text("ALTER TABLE usuarios ADD COLUMN nome VARCHAR(120)"))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                ALTER TABLE IF EXISTS public.usuarios
                ADD COLUMN IF NOT EXISTS nome VARCHAR(120)
            """))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4) Migração leve: garantir coluna qtd_entregas em lancamentos
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(lancamentos);")).fetchall()
            colnames = {row[1] for row in cols}
            if "qtd_entregas" not in colnames:
                db.session.execute(sa_text("ALTER TABLE lancamentos ADD COLUMN qtd_entregas INTEGER"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF EXISTS public.lancamentos "
                "ADD COLUMN IF NOT EXISTS qtd_entregas INTEGER"
            ))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 3.99) corrige registros antigos com ativo NULL
    try:
        if _is_sqlite():
            db.session.execute(sa_text("""
                UPDATE usuarios
                   SET ativo = 1
                 WHERE ativo IS NULL
            """))
        else:
            db.session.execute(sa_text("""
                UPDATE public.usuarios
                   SET ativo = TRUE
                 WHERE ativo IS NULL
            """))
        db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.x) período em despesas_cooperado (data_inicio / data_fim) + backfill
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(despesas_cooperado);")).fetchall()
            colnames = {row[1] for row in cols}

            if "data_inicio" not in colnames:
                db.session.execute(sa_text("ALTER TABLE despesas_cooperado ADD COLUMN data_inicio DATE"))
            if "data_fim" not in colnames:
                db.session.execute(sa_text("ALTER TABLE despesas_cooperado ADD COLUMN data_fim DATE"))
            db.session.commit()

            # Retropreenche linhas antigas
            db.session.execute(sa_text("""
                UPDATE despesas_cooperado
                   SET data_inicio = COALESCE(data_inicio, data),
                       data_fim    = COALESCE(data_fim,    data)
                 WHERE data IS NOT NULL
                   AND (data_inicio IS NULL OR data_fim IS NULL)
            """))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                ALTER TABLE IF NOT EXISTS public.despesas_cooperado
                ADD COLUMN IF NOT EXISTS data_inicio DATE
            """))
            db.session.execute(sa_text("""
                ALTER TABLE IF NOT EXISTS public.despesas_cooperado
                ADD COLUMN IF NOT EXISTS data_fim DATE
            """))
            db.session.commit()

            db.session.execute(sa_text("""
                UPDATE public.despesas_cooperado
                   SET data_inicio = COALESCE(data_inicio, data),
                       data_fim    = COALESCE(data_fim,    data)
                 WHERE data IS NOT NULL
                   AND (data_inicio IS NULL OR data_fim IS NULL)
            """))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.y) beneficio_id em despesas_cooperado (FK p/ beneficios_registro)
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(despesas_cooperado);")).fetchall()
            colnames = {row[1] for row in cols}
            if "beneficio_id" not in colnames:
                db.session.execute(sa_text("ALTER TABLE despesas_cooperado ADD COLUMN beneficio_id INTEGER"))
            db.session.commit()
            # OBS: SQLite não permite adicionar uma FK com ON DELETE CASCADE via ALTER TABLE;
            # para ter a FK de fato, teria que recriar a tabela. Em dev, costuma bastar só a coluna.
        else:
            # Index para consultas
            db.session.execute(sa_text("""
                CREATE INDEX IF NOT EXISTS ix_despesas_beneficio_id
                ON public.despesas_cooperado (beneficio_id)
            """))
            # Normaliza sequência: DROP antes do ADD (idempotente entre deploys)
            db.session.execute(sa_text("""
                ALTER TABLE public.despesas_cooperado
                DROP CONSTRAINT IF EXISTS despesas_cooperado_beneficio_id_fkey
            """))
            db.session.execute(sa_text("""
                ALTER TABLE public.despesas_cooperado
                ADD CONSTRAINT despesas_cooperado_beneficio_id_fkey
                FOREIGN KEY (beneficio_id) REFERENCES public.beneficios_registro (id)
                ON DELETE CASCADE
            """))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.z) eh_adiantamento em despesas_cooperado (marca adiantamento separado)
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(despesas_cooperado);")).fetchall()
            colnames = {row[1] for row in cols}
            if "eh_adiantamento" not in colnames:
                db.session.execute(sa_text("ALTER TABLE despesas_cooperado ADD COLUMN eh_adiantamento BOOLEAN DEFAULT 0"))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                ALTER TABLE IF EXISTS public.despesas_cooperado
                ADD COLUMN IF NOT EXISTS eh_adiantamento BOOLEAN DEFAULT FALSE
            """))
            db.session.commit()
    except Exception:
        db.session.rollback()



    # 4.za) competencia_desconto em despesas_cooperado
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(despesas_cooperado);")).fetchall()
            colnames = {row[1] for row in cols}
            if "competencia_desconto" not in colnames:
                db.session.execute(sa_text("ALTER TABLE despesas_cooperado ADD COLUMN competencia_desconto VARCHAR(20) DEFAULT 'atual'"))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                ALTER TABLE IF EXISTS public.despesas_cooperado
                ADD COLUMN IF NOT EXISTS competencia_desconto VARCHAR(20) DEFAULT 'atual'
            """))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.zb) tabela de abatimentos de despesas do cooperado
    try:
        if _is_sqlite():
            db.session.execute(sa_text("""
                CREATE TABLE IF NOT EXISTS despesas_cooperado_abatimentos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    despesa_id INTEGER NOT NULL,
                    data DATE NOT NULL,
                    valor FLOAT DEFAULT 0.0,
                    origem VARCHAR(30) DEFAULT 'manual',
                    observacao VARCHAR(255),
                    criado_em DATETIME,
                    FOREIGN KEY(despesa_id) REFERENCES despesas_cooperado(id) ON DELETE CASCADE
                )
            """))
            db.session.execute(sa_text("CREATE INDEX IF NOT EXISTS ix_despesas_cooperado_abatimentos_despesa_id ON despesas_cooperado_abatimentos (despesa_id)"))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                CREATE TABLE IF NOT EXISTS public.despesas_cooperado_abatimentos (
                    id SERIAL PRIMARY KEY,
                    despesa_id INTEGER NOT NULL REFERENCES public.despesas_cooperado(id) ON DELETE CASCADE,
                    data DATE NOT NULL,
                    valor DOUBLE PRECISION DEFAULT 0.0,
                    origem VARCHAR(30) DEFAULT 'manual',
                    observacao VARCHAR(255),
                    criado_em TIMESTAMP
                )
            """))
            db.session.execute(sa_text("""
                CREATE INDEX IF NOT EXISTS ix_despesas_cooperado_abatimentos_despesa_id
                ON public.despesas_cooperado_abatimentos (despesa_id)
            """))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.1) cooperado_nome em escalas
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(escalas);")).fetchall()
            colnames = {row[1] for row in cols}
            if "cooperado_nome" not in colnames:
                db.session.execute(sa_text("ALTER TABLE escalas ADD COLUMN cooperado_nome VARCHAR(120)"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS escalas "
                "ADD COLUMN IF NOT EXISTS cooperado_nome VARCHAR(120)"
            ))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.2) restaurante_id em escalas
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(escalas);")).fetchall()
            colnames = {row[1] for row in cols}
            if "restaurante_id" not in colnames:
                db.session.execute(sa_text("ALTER TABLE escalas ADD COLUMN restaurante_id INTEGER"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS escalas "
                "ADD COLUMN IF NOT EXISTS restaurante_id INTEGER"
            ))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.3) fotos no banco (cooperados)
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(cooperados);")).fetchall()
            colnames = {row[1] for row in cols}
            if "foto_bytes" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN foto_bytes BLOB"))
            if "foto_mime" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN foto_mime VARCHAR(100)"))
            if "foto_filename" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN foto_filename VARCHAR(255)"))
            if "foto_url" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN foto_url VARCHAR(255)"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados ADD COLUMN IF NOT EXISTS foto_bytes BYTEA"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados ADD COLUMN IF NOT EXISTS foto_mime VARCHAR(100)"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados ADD COLUMN IF NOT EXISTS foto_filename VARCHAR(255)"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados ADD COLUMN IF NOT EXISTS foto_url VARCHAR(255)"))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.3.x) telefone em cooperados
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(cooperados);")).fetchall()
            colnames = {row[1] for row in cols}
            if "telefone" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN telefone VARCHAR(30)"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados "
                "ADD COLUMN IF NOT EXISTS telefone VARCHAR(30)"
            ))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.4) tabela avaliacoes_restaurante (se não existir)
    try:
        if _is_sqlite():
            db.session.execute(sa_text("""
                CREATE TABLE IF NOT EXISTS avaliacoes_restaurante (
                id SERIAL PRIMARY KEY,
                restaurante_id INTEGER NOT NULL,
                cooperado_id INTEGER NOT NULL,
                lancamento_id INTEGER UNIQUE,
                estrelas_geral DOUBLE PRECISION,
                estrelas_ambiente INTEGER,
                estrelas_tratamento INTEGER,
                estrelas_suporte INTEGER,
                comentario TEXT,
                media_ponderada DOUBLE PRECISION,
                sentimento VARCHAR(12),
                temas VARCHAR(255),
                alerta_crise BOOLEAN DEFAULT FALSE,
                criado_em TIMESTAMP
              );
           """))
            db.session.execute(sa_text(
                "CREATE INDEX IF NOT EXISTS ix_av_rest_rest ON avaliacoes_restaurante(restaurante_id, criado_em)"))
            db.session.execute(sa_text(
                "CREATE INDEX IF NOT EXISTS ix_av_rest_coop ON avaliacoes_restaurante(cooperado_id)"))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                CREATE TABLE IF NOT EXISTS avaliacoes_restaurante (
                  id SERIAL PRIMARY KEY,
                  restaurante_id INTEGER NOT NULL,
                  cooperado_id   INTEGER NOT NULL,
                  lancamento_id  INTEGER UNIQUE,
                  estrelas_geral INTEGER,
                  estrelas_ambiente   = db.Column(db.Integer)
                  estrelas_tratamento = db.Column(db.Integer)
                  estrelas_suporte    = db.Column(db.Integer)
                  comentario TEXT,
                  media_ponderada DOUBLE PRECISION,
                  sentimento VARCHAR(12),
                  temas VARCHAR(255),
                  alerta_crise BOOLEAN DEFAULT FALSE,
                  criado_em TIMESTAMP
                );
            """))
            db.session.execute(sa_text(
                "CREATE INDEX IF NOT EXISTS ix_av_rest_rest ON avaliacoes_restaurante(restaurante_id, criado_em)"))
            db.session.execute(sa_text(
                "CREATE INDEX IF NOT EXISTS ix_av_rest_coop ON avaliacoes_restaurante(cooperado_id)"))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.5) fotos no banco (restaurantes)
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(restaurantes);")).fetchall()
            colnames = {row[1] for row in cols}
            if "foto_bytes" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN foto_bytes BLOB"))
            if "foto_mime" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN foto_mime VARCHAR(100)"))
            if "foto_filename" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN foto_filename VARCHAR(255)"))
            if "foto_url" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN foto_url VARCHAR(255)"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS restaurantes ADD COLUMN IF NOT EXISTS foto_bytes BYTEA"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS restaurantes ADD COLUMN IF NOT EXISTS foto_mime VARCHAR(100)"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS restaurantes ADD COLUMN IF NOT EXISTS foto_filename VARCHAR(255)"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS restaurantes ADD COLUMN IF NOT EXISTS foto_url VARCHAR(255)"))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.5.x) taxa administrativa em restaurantes + receitas_coop
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(restaurantes);")).fetchall()
            colnames = {row[1] for row in cols}
            if "taxa_admin_valor" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN taxa_admin_valor FLOAT DEFAULT 0"))
            if "taxa_admin_data_base" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN taxa_admin_data_base DATE"))
            if "taxa_admin_multa_percentual" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN taxa_admin_multa_percentual FLOAT DEFAULT 2.0"))
            if "taxa_admin_juros_dia_percentual" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN taxa_admin_juros_dia_percentual FLOAT DEFAULT 0.03"))
            if "ativo" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN ativo BOOLEAN DEFAULT 1"))
            db.session.commit()

            cols = db.session.execute(sa_text("PRAGMA table_info(receitas_coop);")).fetchall()
            colnames = {row[1] for row in cols}
            adds = {
                "restaurante_id": "INTEGER",
                "auto_taxa_adm": "BOOLEAN DEFAULT 0",
                "competencia": "VARCHAR(7)",
                "valor_previsto": "FLOAT DEFAULT 0",
                "valor_principal": "FLOAT DEFAULT 0",
                "valor_pago": "FLOAT DEFAULT 0",
                "valor_multa": "FLOAT DEFAULT 0",
                "valor_juros": "FLOAT DEFAULT 0",
                "data_vencimento": "DATE",
                "data_pagamento": "DATE",
                "status_pagamento": "VARCHAR(20) DEFAULT 'nao_pago'",
                "multa_percentual": "FLOAT DEFAULT 2.0",
                "juros_dia_percentual": "FLOAT DEFAULT 0.03",
            }
            for col, ddl in adds.items():
                if col not in colnames:
                    db.session.execute(sa_text(f"ALTER TABLE receitas_coop ADD COLUMN {col} {ddl}"))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                ALTER TABLE IF EXISTS public.restaurantes ADD COLUMN IF NOT EXISTS taxa_admin_valor DOUBLE PRECISION DEFAULT 0;
                ALTER TABLE IF EXISTS public.restaurantes ADD COLUMN IF NOT EXISTS taxa_admin_data_base DATE;
                ALTER TABLE IF EXISTS public.restaurantes ADD COLUMN IF NOT EXISTS taxa_admin_multa_percentual DOUBLE PRECISION DEFAULT 2.0;
                ALTER TABLE IF EXISTS public.restaurantes ADD COLUMN IF NOT EXISTS taxa_admin_juros_dia_percentual DOUBLE PRECISION DEFAULT 0.03;
                ALTER TABLE IF EXISTS public.restaurantes ADD COLUMN IF NOT EXISTS ativo BOOLEAN DEFAULT TRUE;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS restaurante_id INTEGER;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS auto_taxa_adm BOOLEAN DEFAULT FALSE;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS competencia VARCHAR(7);
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS valor_previsto DOUBLE PRECISION DEFAULT 0;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS valor_principal DOUBLE PRECISION DEFAULT 0;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS valor_pago DOUBLE PRECISION DEFAULT 0;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS valor_multa DOUBLE PRECISION DEFAULT 0;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS valor_juros DOUBLE PRECISION DEFAULT 0;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS data_vencimento DATE;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS data_pagamento DATE;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS status_pagamento VARCHAR(20) DEFAULT 'nao_pago';
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS multa_percentual DOUBLE PRECISION DEFAULT 2.0;
                ALTER TABLE IF EXISTS public.receitas_coop ADD COLUMN IF NOT EXISTS juros_dia_percentual DOUBLE PRECISION DEFAULT 0.03;
            """))
            db.session.commit()
        try:
            db.session.execute(sa_text("UPDATE receitas_coop SET status_pagamento = 'nao_pago' WHERE status_pagamento IS NULL"))
            db.session.execute(sa_text("UPDATE receitas_coop SET auto_taxa_adm = FALSE WHERE auto_taxa_adm IS NULL"))
            db.session.commit()
        except Exception:
            db.session.rollback()
    except Exception:
        db.session.rollback()

    # 5) Bootstrap mínimo (admin e config) — só se os modelos existirem
    try:
        # Garante que o model Usuario está acessível
        _ = Usuario  # type: ignore[name-defined]

        # Admin
        try:
            tem_admin = Usuario.query.filter_by(tipo="admin").first()  # type: ignore[name-defined]
        except Exception:
            tem_admin = None

        if not tem_admin:
            admin_user = os.environ.get("ADMIN_USER", "admin")
            admin_pass = os.environ.get("ADMIN_PASS", os.urandom(8).hex())
            admin = Usuario(
                usuario=admin_user,
                tipo="admin",
                senha_hash="",
                is_master=True
            )  # type: ignore[name-defined]
            try:
                admin.set_password(admin_pass)  # type: ignore[attr-defined]
            except Exception:
                try:
                    from werkzeug.security import generate_password_hash
                    admin.senha_hash = generate_password_hash(admin_pass)  # type: ignore[attr-defined]
                except Exception:
                    pass
            db.session.add(admin)
            db.session.commit()
    except Exception:
        db.session.rollback()

    try:
        admin_master = Usuario.query.filter_by(tipo="admin", is_master=True).first()
        if not admin_master:
            primeiro_admin = Usuario.query.filter_by(tipo="admin").order_by(Usuario.id.asc()).first()
            if primeiro_admin:
                primeiro_admin.is_master = True
                db.session.commit()
    except Exception:
        db.session.rollback()

    try:
        # Se o model Config existir, cria default
        try:
            Config  # type: ignore[name-defined]
            has_config_model = True
        except NameError:
            has_config_model = False

        if has_config_model:
            if not Config.query.get(1):  # type: ignore[name-defined]
                db.session.add(Config(id=1, salario_minimo=0.0))  # type: ignore[name-defined]
                db.session.commit()
    except Exception:
        db.session.rollback()

# === Bootstrap do banco no start (Render/Gunicorn) ===
try:
    if os.environ.get("INIT_DB_ON_START", "1") == "1":
        _t0 = datetime.utcnow()
        with app.app_context():
       
            init_db()
        try:
            app.logger.info(f"init_db concluído em {(datetime.utcnow() - _t0).total_seconds():.2f}s")
        except Exception:
            pass
    else:
        try:
            app.logger.info("INIT_DB_ON_START=0: pulando init_db no boot.")
        except Exception:
            pass
except Exception as e:
    try:
        app.logger.warning(f"init_db falhou/pulado: {e}")
    except Exception:
        pass


def get_config() -> Config:
    cfg = Config.query.get(1)
    if not cfg:
        cfg = Config(id=1, salario_minimo=0.0)
        db.session.add(cfg)
        db.session.commit()
    return cfg


def role_required(role: str):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            uid = session.get("user_id")
            tipo = (session.get("user_tipo") or "").strip().lower()
            role_norm = (role or "").strip().lower()

            if not uid or not tipo:
                session.clear()
                return redirect(url_for("login"))

            u = Usuario.query.get(uid)
            if not u:
                session.clear()
                return redirect(url_for("login"))

            tipo_db = (u.tipo or "").strip().lower()

            if getattr(u, "ativo", None) is None:
                u.ativo = True
                db.session.commit()

            if u.ativo is False:
                session.clear()
                flash("Conta desativada. Fale com o administrador.", "danger")
                return redirect(url_for("login"))

            if tipo_db != role_norm:
                session.clear()
                return redirect(url_for("login"))

            if tipo != tipo_db:
                session["user_tipo"] = tipo_db

            return fn(*args, **kwargs)
        return wrapper
    return deco


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        uid = session.get("user_id")
        tipo = (session.get("user_tipo") or "").strip().lower()

        if not uid:
            session.clear()
            return redirect(url_for("login"))

        u = Usuario.query.get(uid)
        if not u:
            session.clear()
            flash("Sessão inválida. Faça login novamente.", "danger")
            return redirect(url_for("login"))

        tipo_db = (u.tipo or "").strip().lower()

        if tipo_db != "admin":
            session.clear()
            flash("Acesso restrito ao administrador.", "danger")
            return redirect(url_for("login"))

        if getattr(u, "ativo", None) is None:
            u.ativo = True
            db.session.commit()

        if u.ativo is False:
            session.clear()
            flash("Conta desativada. Fale com o administrador master.", "danger")
            return redirect(url_for("login"))

        if tipo != "admin":
            session["user_tipo"] = "admin"

        session["user_id"] = u.id

        return fn(*args, **kwargs)
    return wrapper


ADMIN_ABAS = {
    "lancamentos": "Lançamentos",
    "receitas": "Receitas Coop",
    "despesas": "Despesas Coop",
    "coop_receitas": "Receitas Cooperado",
    "coop_despesas": "Despesas Cooperado",
    "beneficios": "Benefícios",
    "cooperados": "Cooperados",
    "restaurantes": "Restaurantes",
    "escalas": "Escalas",
    "avisos": "Avisos",
    "documentos": "Documentos",
    "tabelas": "Tabelas",
    "avaliacoes": "Avaliações",
    "config": "Configurações",
    "folha": "Folha",
}


def _usuario_logado() -> Usuario | None:
    uid = session.get("user_id")
    if not uid:
        return None

    u = Usuario.query.get(uid)
    if not u:
        return None

    if getattr(u, "ativo", None) is None:
        u.ativo = True
        db.session.commit()

    if u.ativo is False:
        return None

    return u


def is_admin_master() -> bool:
    u = _usuario_logado()
    return bool(u and (u.tipo or "").strip().lower() == "admin" and getattr(u, "is_master", False))


def get_admin_permissions_map(usuario_id: int) -> dict:
    perms = AdminPermissao.query.filter_by(usuario_id=usuario_id).all()
    out = {}

    for p in perms:
        out[p.aba] = {
            "ver": bool(p.pode_ver),
            "criar": bool(p.pode_criar),
            "editar": bool(p.pode_editar),
            "excluir": bool(p.pode_excluir),
        }

    return out


def admin_has_perm(aba: str, acao: str = "ver") -> bool:
    u = _usuario_logado()
    if not u:
        return False

    if (u.tipo or "").strip().lower() != "admin":
        return False

    if getattr(u, "is_master", False):
        return True

    if aba not in ADMIN_ABAS:
        return False

    mapa = get_admin_permissions_map(u.id)
    perm = mapa.get(aba, {})
    return bool(perm.get(acao, False))


def admin_perm_required(aba: str, acao: str = "ver"):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            wants_json = (
                request.headers.get("X-Requested-With") == "XMLHttpRequest"
                or request.headers.get("Accept", "").lower().find("application/json") >= 0
                or (request.form.get("ajax") == "1")
                or (request.args.get("ajax") == "1")
            )

            def _deny_json(message: str, status: int):
                return jsonify({"ok": False, "message": message}), status

            uid = session.get("user_id")
            tipo = (session.get("user_tipo") or "").strip().lower()

            if not uid:
                session.clear()
                if wants_json:
                    return _deny_json("Sessão expirada. Faça login novamente.", 401)
                return redirect(url_for("login"))

            u = Usuario.query.get(uid)
            if not u:
                session.clear()
                if wants_json:
                    return _deny_json("Sessão inválida. Faça login novamente.", 401)
                flash("Sessão inválida. Faça login novamente.", "danger")
                return redirect(url_for("login"))

            tipo_db = (u.tipo or "").strip().lower()

            if tipo_db != "admin":
                session.clear()
                if wants_json:
                    return _deny_json("Acesso restrito ao administrador.", 403)
                flash("Acesso restrito ao administrador.", "danger")
                return redirect(url_for("login"))

            if getattr(u, "ativo", None) is None:
                u.ativo = True
                db.session.commit()

            if u.ativo is False:
                session.clear()
                if wants_json:
                    return _deny_json("Conta desativada. Fale com o administrador master.", 403)
                flash("Conta desativada. Fale com o administrador master.", "danger")
                return redirect(url_for("login"))

            if tipo != "admin":
                session["user_tipo"] = "admin"

            if not admin_has_perm(aba, acao):
                if wants_json:
                    return _deny_json("Você não tem permissão para editar Escalas.", 403)

                flash("Você não tem permissão para essa ação.", "danger")

                if getattr(u, "is_master", False):
                    return redirect(url_for("admin_dashboard", tab="lancamentos"))

                abas_liberadas = [
                    nome_aba
                    for nome_aba in ADMIN_ABAS.keys()
                    if admin_has_perm(nome_aba, "ver")
                ]

                if abas_liberadas:
                    return redirect(url_for("admin_dashboard", tab=abas_liberadas[0]))

                session.clear()
                flash("Seu usuário admin está sem permissões liberadas.", "warning")
                return redirect(url_for("login"))

            return fn(*args, **kwargs)
        return wrapper
    return deco


def _normalize_name(s: str) -> list[str]:
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^a-zA-Z0-9\s]", " ", s)
    parts = [p.lower() for p in s.split() if p.strip()]
    return parts

def _norm_login(s: str) -> str:
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.lower().strip()
    s = re.sub(r"\s+", "", s)
    return s


def _match_cooperado_by_login(login_planilha: str, cooperados: list[Cooperado]) -> Cooperado | None:
    """Casa EXATAMENTE com Usuario.usuario após normalização."""
    key = _norm_login(login_planilha)
    if not key:
        return None

    for c in cooperados:
        login = getattr(c.usuario_ref, "usuario", "") or ""
        if _norm_login(login) == key:
            return c

    return None


def _match_restaurante_id(contrato_txt: str) -> int | None:
    alvo = " ".join(_normalize_name(contrato_txt or ""))
    if not alvo:
        return None

    restaurantes = Restaurante.query.order_by(Restaurante.nome.asc()).all()

    for r in restaurantes:
        rn = " ".join(_normalize_name(r.nome))
        if alvo == rn or alvo in rn or rn in alvo:
            return r.id

    try:
        nomes_norm = [" ".join(_normalize_name(r.nome)) for r in restaurantes]
        close = difflib.get_close_matches(alvo, nomes_norm, n=1, cutoff=0.85)
        if close:
            alvo_norm = close[0]
            for r in restaurantes:
                if " ".join(_normalize_name(r.nome)) == alvo_norm:
                    return r.id
    except Exception:
        pass

    return None



def _escala_weekday_num(data_txt: str | None) -> int:
    """Retorna 1=segunda ... 7=domingo a partir do texto da data da escala."""
    try:
        wd = _weekday_from_data_str(data_txt)
        if wd in (1, 2, 3, 4, 5, 6, 7):
            return int(wd)
    except Exception:
        pass

    try:
        dt = _parse_data_escala_str(data_txt)
        if dt:
            return int(dt.weekday()) + 1
    except Exception:
        pass

    return 0


def _escala_weekday_label(data_txt: str | None) -> str:
    mapa = {1: "Segunda", 2: "Terça", 3: "Quarta", 4: "Quinta", 5: "Sexta", 6: "Sábado", 7: "Domingo"}
    return mapa.get(_escala_weekday_num(data_txt), "Sem dia")


def _escala_sort_key(e: Escala):
    dt = _parse_data_escala_str(getattr(e, "data", None))
    data_ord = dt.toordinal() if dt else 0

    horario_txt = str(getattr(e, "horario", "") or "")
    m = re.search(r"(\d{1,2}):(\d{2})", horario_txt)
    mins = (int(m.group(1)) * 60 + int(m.group(2))) if m else 9999

    return (
        _escala_weekday_num(getattr(e, "data", None)),
        (getattr(e, "contrato", "") or "").strip().lower(),
        data_ord,
        mins,
        (getattr(e, "turno", "") or "").strip().lower(),
        int(getattr(e, "id", 0) or 0),
    )


def _brasil_now() -> datetime:
    """Agora em horário de Brasília sem depender do timezone do servidor."""
    return datetime.utcnow() - timedelta(hours=3)


def _escala_inicio_datetime(e: Escala) -> datetime | None:
    dt = _parse_data_escala_str(getattr(e, "data", None))
    if not dt:
        return None

    horario_txt = str(getattr(e, "horario", "") or "")
    m = re.search(r"(\d{1,2}):(\d{2})", horario_txt)
    if not m:
        return None

    hh = int(m.group(1))
    mm = int(m.group(2))
    if hh > 23 or mm > 59:
        return None

    return datetime.combine(dt, dtime(hour=hh, minute=mm))


def _build_escala_alertas_1h(escalas_all: list[Escala], cooperados_map: dict[int, Cooperado] | None = None) -> list[dict]:
    agora = _brasil_now()
    limite = agora + timedelta(hours=1)
    out = []
    cooperados_map = cooperados_map or {}

    for e in escalas_all:
        sem_cooperado = not getattr(e, "cooperado_id", None) and not (getattr(e, "cooperado_nome", None) or "").strip()
        if not sem_cooperado:
            continue

        inicio = _escala_inicio_datetime(e)
        if not inicio:
            continue

        if not (agora <= inicio <= limite):
            continue

        minutos = int((inicio - agora).total_seconds() // 60)
        contrato = (getattr(e, "contrato", None) or "Sem contrato").strip() or "Sem contrato"
        coop = cooperados_map.get(getattr(e, "cooperado_id", 0) or 0)
        out.append({
            "id": int(getattr(e, "id", 0) or 0),
            "data": getattr(e, "data", "") or "",
            "turno": getattr(e, "turno", "") or "",
            "horario": getattr(e, "horario", "") or "",
            "contrato": contrato,
            "cooperado_nome": (coop.nome if coop else (getattr(e, "cooperado_nome", None) or "").strip()),
            "minutos_restantes": max(0, minutos),
            "inicio_iso": inicio.strftime("%Y-%m-%dT%H:%M:%S"),
            "weekday_label": _escala_weekday_label(getattr(e, "data", None)),
            "mensagem": f"Cobrir contrato {contrato} às {(getattr(e, 'horario', '') or '').strip() or '—'}",
        })

    out.sort(key=lambda item: (item.get("inicio_iso", ""), (item.get("contrato", "") or "").lower(), item.get("id", 0)))
    return out

def _match_cooperado_by_name(nome_planilha: str, cooperados: list[Cooperado]) -> Cooperado | None:
    def norm_join(s: str) -> str:
        return " ".join(_normalize_name(s))

    sheet_tokens = _normalize_name(nome_planilha)
    sheet_norm = " ".join(sheet_tokens)
    if not sheet_norm:
        return None

    for c in cooperados:
        c_norm = norm_join(c.nome)
        if sheet_norm == c_norm or sheet_norm in c_norm or c_norm in sheet_norm:
            return c

    parts_sheet = set(sheet_tokens)
    best = None
    best_count = 0

    for c in cooperados:
        parts_c = set(_normalize_name(c.nome))
        inter = parts_sheet & parts_c
        if len(inter) > best_count:
            best = c
            best_count = len(inter)

    if best and best_count >= 2:
        return best

    if len(sheet_tokens) == 1 and len(sheet_tokens[0]) >= 3:
        token = sheet_tokens[0]
        hits = [c for c in cooperados if token in set(_normalize_name(c.nome))]
        if hits:
            return hits[0]

    names_norm = [norm_join(c.nome) for c in cooperados]
    close = difflib.get_close_matches(sheet_norm, names_norm, n=1, cutoff=0.85)
    if close:
        target = close[0]
        for c in cooperados:
            if norm_join(c.nome) == target:
                return c

    return None


def _build_docinfo(c: Cooperado) -> dict:
    today = date.today()
    cnh_ok = (c.cnh_validade is not None and c.cnh_validade >= today)
    placa_ok = (c.placa_validade is not None and c.placa_validade >= today)
    return {
        "cnh": {"ok": cnh_ok},
        "placa": {"ok": placa_ok},
    }


def _save_upload(file_storage) -> str | None:
    # Mantido para compatibilidade com outras partes do app (ex.: uploads de xlsx)
    if not file_storage:
        return None

    fname = secure_filename(file_storage.filename or "")
    if not fname:
        return None

    path = os.path.join(UPLOAD_DIR, fname)
    file_storage.save(path)
    return f"/static/uploads/{fname}"

# =========================
# Helpers de arquivos persistentes
# =========================

# Diretórios persistentes de fotos
FOTOS_PERSIST_DIR = os.path.join(PERSIST_ROOT, "fotos")
FOTOS_COOPS_DIR = os.path.join(FOTOS_PERSIST_DIR, "cooperados")
FOTOS_RESTS_DIR = os.path.join(FOTOS_PERSIST_DIR, "restaurantes")

os.makedirs(FOTOS_PERSIST_DIR, exist_ok=True)
os.makedirs(FOTOS_COOPS_DIR, exist_ok=True)
os.makedirs(FOTOS_RESTS_DIR, exist_ok=True)

# Compatibilidade/legado em static
STATIC_FOTOS_DIR = os.path.join(UPLOAD_DIR, "fotos")
STATIC_FOTOS_COOPS_DIR = os.path.join(STATIC_FOTOS_DIR, "cooperados")
STATIC_FOTOS_RESTS_DIR = os.path.join(STATIC_FOTOS_DIR, "restaurantes")

os.makedirs(STATIC_FOTOS_DIR, exist_ok=True)
os.makedirs(STATIC_FOTOS_COOPS_DIR, exist_ok=True)
os.makedirs(STATIC_FOTOS_RESTS_DIR, exist_ok=True)


def salvar_tabela_upload(file_storage) -> str | None:
    """
    Salva a TABELA em disco persistente e retorna o nome único salvo.
    Esse nome deve ser gravado no banco em:
      - Tabela.arquivo_url
      - Tabela.arquivo_nome
    """
    if not file_storage or not file_storage.filename:
        return None

    fname = secure_filename(file_storage.filename)
    if not fname:
        return None

    base, ext = os.path.splitext(fname)
    unique = f"{base}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}{ext.lower()}"
    destino = os.path.join(TABELAS_DIR, unique)

    os.makedirs(TABELAS_DIR, exist_ok=True)
    file_storage.save(destino)

    return unique


def resolve_tabela_path(nome_arquivo: str) -> str | None:
    """
    Resolve o caminho real da tabela nesta ordem:
      1) disco persistente
      2) legado em static/uploads/tabelas
      3) caminho absoluto derivado de /static/...
    """
    if not nome_arquivo:
        return None

    nome_limpo = str(nome_arquivo).split("?", 1)[0].split("#", 1)[0].strip()
    if not nome_limpo:
        return None

    somente_nome = os.path.basename(nome_limpo)

    candidatos = [
        os.path.join(TABELAS_DIR, somente_nome),
        os.path.join(STATIC_TABLES, somente_nome),
        _abs_path_from_url(nome_limpo) if nome_limpo.startswith("/") else None,
        nome_limpo if os.path.isabs(nome_limpo) else None,
    ]

    for p in candidatos:
        if p and os.path.isfile(p):
            return p

    app.logger.warning(
        "Arquivo de Tabela não encontrado. nome='%s' tents=%s",
        nome_arquivo,
        [c for c in candidatos if c]
    )
    return None


def _save_foto_to_db(entidade, file_storage, *, is_cooperado: bool) -> str | None:
    """
    Mantido com o mesmo nome para não quebrar o restante do sistema,
    mas agora salva a FOTO EM DISCO PERSISTENTE em vez de salvar no banco.

    Retorna a URL pública da foto salva em /static/uploads/fotos/...,
    enquanto o arquivo real fica protegido no disco persistente.

    Também grava:
      - foto_filename
      - foto_mime
      - foto_url

    E limpa os campos binários antigos do banco, se existirem.
    """
    if not file_storage or not file_storage.filename:
        return getattr(entidade, "foto_url", None)

    raw_name = secure_filename(file_storage.filename or "")
    if not raw_name:
        return getattr(entidade, "foto_url", None)

    db.session.flush()

    ext = os.path.splitext(raw_name)[1].lower() or ".jpg"
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if is_cooperado:
        persist_dir = FOTOS_COOPS_DIR
        static_dir = STATIC_FOTOS_COOPS_DIR
        prefix = f"coop_{entidade.id}_{stamp}"
        public_url = f"/static/uploads/fotos/cooperados/{prefix}{ext}"
    else:
        persist_dir = FOTOS_RESTS_DIR
        static_dir = STATIC_FOTOS_RESTS_DIR
        prefix = f"rest_{entidade.id}_{stamp}"
        public_url = f"/static/uploads/fotos/restaurantes/{prefix}{ext}"

    os.makedirs(persist_dir, exist_ok=True)
    os.makedirs(static_dir, exist_ok=True)

    final_name = f"{prefix}{ext}"
    persist_path = os.path.join(persist_dir, final_name)
    static_path = os.path.join(static_dir, final_name)

    file_storage.save(persist_path)

    # espelho em static para manter compatibilidade com templates antigos
    try:
        import shutil
        shutil.copy2(persist_path, static_path)
    except Exception:
        pass

    # remove foto antiga do mesmo tipo, se existir
    foto_antiga = getattr(entidade, "foto_url", None)
    if foto_antiga:
        try:
            antiga_nome = os.path.basename(str(foto_antiga).split("?", 1)[0])
            if antiga_nome and antiga_nome != final_name:
                if is_cooperado:
                    antigos = [
                        os.path.join(FOTOS_COOPS_DIR, antiga_nome),
                        os.path.join(STATIC_FOTOS_COOPS_DIR, antiga_nome),
                    ]
                else:
                    antigos = [
                        os.path.join(FOTOS_RESTS_DIR, antiga_nome),
                        os.path.join(STATIC_FOTOS_RESTS_DIR, antiga_nome),
                    ]
                for oldp in antigos:
                    if os.path.isfile(oldp):
                        os.remove(oldp)
        except Exception:
            pass

    entidade.foto_filename = raw_name
    entidade.foto_mime = (file_storage.mimetype or "application/octet-stream")
    entidade.foto_url = f"{public_url}?v={int(datetime.utcnow().timestamp())}"

    # limpa dados binários antigos do banco, se os campos existirem
    if hasattr(entidade, "foto_bytes"):
        entidade.foto_bytes = None

    return entidade.foto_url


def _abs_path_from_url(rel_url: str) -> str:
    """
    Converte '/static/uploads/arquivo.pdf' para o caminho absoluto no disco.
    """
    if not rel_url:
        return ""

    rel_url = str(rel_url).split("?", 1)[0].split("#", 1)[0].strip()
    if not rel_url:
        return ""

    if rel_url.startswith("/"):
        rel_url = rel_url.lstrip("/")

    return os.path.join(BASE_DIR, rel_url.replace("/", os.sep))


def _serve_uploaded(rel_url: str, *, download_name: str | None = None, force_download: bool = False):
    """
    Entrega arquivo salvo em /static/uploads ou em fallback do disco persistente.
    - PDFs abrem inline por padrão.
    - Outros tipos baixam, salvo se force_download=False e o navegador suportar.
    """
    if not rel_url:
        abort(404)

    rel_limpa = str(rel_url).split("?", 1)[0].split("#", 1)[0].strip()
    if not rel_limpa:
        abort(404)

    abs_path = _abs_path_from_url(rel_limpa)

    # fallback especial para fotos persistentes
    if not os.path.exists(abs_path):
        nome = os.path.basename(rel_limpa)

        candidatos = [
            os.path.join(FOTOS_COOPS_DIR, nome),
            os.path.join(FOTOS_RESTS_DIR, nome),
            os.path.join(DOCS_PERSIST_DIR, nome),
            os.path.join(TABELAS_DIR, nome),
        ]

        achado = next((p for p in candidatos if os.path.isfile(p)), None)
        if not achado:
            abort(404)
        abs_path = achado

    mime, _ = mimetypes.guess_type(abs_path)
    is_pdf = (mime == "application/pdf") or abs_path.lower().endswith(".pdf")

    return send_file(
        abs_path,
        mimetype=mime or "application/octet-stream",
        as_attachment=(force_download or not is_pdf),
        download_name=(download_name or os.path.basename(abs_path)),
        conditional=True,
    )


# ========= Helpers de DOCUMENTOS =========
def salvar_documento_upload(file_storage) -> str | None:
    """
    Salva o documento em disco persistente e retorna o nome único salvo.
    Esse nome deve ser gravado no banco em:
      - Documento.arquivo_url (ou URL /docs/<nome>)
      - Documento.arquivo_nome
    """
    if not file_storage or not file_storage.filename:
        return None

    fname = secure_filename(file_storage.filename)
    if not fname:
        return None

    base, ext = os.path.splitext(fname)
    unique = f"{base}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}{ext.lower()}"
    destino = os.path.join(DOCS_PERSIST_DIR, unique)

    os.makedirs(DOCS_PERSIST_DIR, exist_ok=True)
    file_storage.save(destino)

    return unique


def resolve_documento_path(nome_arquivo: str) -> str | None:
    """
    Resolve o caminho real do documento nesta ordem:
      1) persistente
      2) legado em static/uploads/docs
      3) caminho absoluto derivado de /static/...
    """
    if not nome_arquivo:
        return None

    nome_limpo = str(nome_arquivo).split("?", 1)[0].split("#", 1)[0].strip()
    if not nome_limpo:
        return None

    somente_nome = os.path.basename(nome_limpo)

    candidatos = [
        os.path.join(DOCS_PERSIST_DIR, somente_nome),
        os.path.join(DOCS_DIR, somente_nome),
        _abs_path_from_url(nome_limpo) if nome_limpo.startswith("/") else None,
        nome_limpo if os.path.isabs(nome_limpo) else None,
    ]

    for p in candidatos:
        if p and os.path.isfile(p):
            return p

    app.logger.warning(
        "Documento não encontrado. nome='%s' tents=%s",
        nome_arquivo,
        [c for c in candidatos if c]
    )
    return None


def _assert_cooperado_ativo(cooperado_id: int):
    c = (
        Cooperado.query
        .join(Usuario, Cooperado.usuario_id == Usuario.id)
        .filter(Cooperado.id == cooperado_id, Usuario.ativo.is_(True))
        .first()
    )
    if not c:
        abort(400, description="Cooperado inativo ou inexistente.")
    return c
# ========= ROTA: /docs/<nome> (abre inline PDF; baixa outros tipos) =========
@app.get("/docs/<path:nome>")
def serve_documento(nome: str):
    """
    Abre PDFs inline (no navegador) e força download para outros tipos.
    Busca primeiro no disco persistente e faz fallback pro legado.
    """
    path = resolve_documento_path(nome)
    if not path:
        abort(404)

    mime, _ = mimetypes.guess_type(path)
    is_pdf = (mime == "application/pdf") or path.lower().endswith(".pdf")

    return send_file(
        path,
        mimetype=mime or "application/octet-stream",
        as_attachment=not is_pdf,                  # PDF inline, outros baixam
        download_name=os.path.basename(path),
        conditional=True,
    )

def _prox_ocorrencia_anual(dt: date | None) -> date | None:
    if not dt:
        return None
    hoje = date.today()
    alvo = date(hoje.year, dt.month, dt.day)
    if alvo < hoje:
        alvo = date(hoje.year + 1, dt.month, dt.day)
    return alvo

def _parse_ymd(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _bounds_mes(yyyy_mm: str):
    # "2025-06" -> [2025-06-01, 2025-07-01)
    y, m = map(int, yyyy_mm.split("-"))
    ini = date(y, m, 1)
    fim = (ini + relativedelta(months=1))
    return ini, fim

def _bounds_semana_atual(ref: date | None = None) -> tuple[date, date]:
    """
    Retorna o intervalo da semana atual (segunda a domingo).
    """
    ref = ref or date.today()
    ini = ref - timedelta(days=ref.weekday())   # segunda
    fim = ini + timedelta(days=6)               # domingo
    return ini, fim

def _parse_data_ymd(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _fmt_br(d: date | None) -> str:
    return d.strftime("%d/%m/%Y") if d else ""

# === Helpers de Avaliação (NLP leve + métricas) =============================
def _clamp_star(v):
    try:
        v = int(v)
    except Exception:
        return None
    return min(5, max(1, v))

def _media_ponderada(geral, pont, educ, efic, apres):
    """
    Ponderação (soma=1.0):
      Geral 0.40 + Pontualidade 0.15 + Educação 0.15 + Eficiência 0.15 + Bem Apresentado 0.15
    Calcula só com os campos presentes (ignora None) e renormaliza pesos.
    """
    pares = [
        (geral, 0.40),
        (pont,  0.15),
        (educ,  0.15),
        (efic,  0.15),
        (apres, 0.15),
    ]
    num = 0.0
    den = 0.0
    for nota, w in pares:
        if nota is not None:
            num += float(nota) * w
            den += w
    return round(num / den, 2) if den > 0 else None

_POS = set("""
bom ótima otimo excelente parabéns educado gentil atencioso cordial limpo cheiroso organizado rápido rapida rapido pontual
""".split())
_NEG = set("""
ruim péssimo pessimo horrível horrivel sujo atrasado grosseiro mal educado agressivo impaciente amassado quebrado frio derramou
""".split())

def _analise_sentimento(txt: str | None) -> str:
    if not txt:
        return "neutro"
    t = (txt or "").lower()
    # contagem bem simples
    pos = sum(1 for w in _POS if w in t)
    neg = sum(1 for w in _NEG if w in t)
    if neg > pos + 0: return "negativo"
    if pos > neg + 0: return "positivo"
    return "neutro"

# mapeia temas por palavras-chave simples
_TEMAS = {
    "Pontualidade":  ["pontual", "atras", "horario", "horário", "demor", "rápido", "rapido", "lent"],
    "Educação":      ["educad", "grosseir", "simpat", "antipatic", "mal trat", "sem paciencia", "sem paciência", "atencios"],
    "Eficiência":    ["amass", "vazou", "quebrad", "frio", "bagunça", "bagunca", "cuidado", "eficien", "desorgan"],
    "Bem apresentado": ["uniform", "higien", "apresenta", "limpo", "cheiroso", "aparencia", "aparência"],
}

def _identifica_temas(txt: str | None) -> list[str]:
    if not txt:
        return []
    t = (txt or "").lower()
    hits = []
    for tema, keys in _TEMAS.items():
        if any(k in t for k in keys):
            hits.append(tema)
    return hits[:4]

_RISCO = ["ameaça","ameaca","acidente","quebrado","agress","roubo","violên","violenc","lesão","lesao","sangue","caiu","bateu","droga","alcool","álcool"]

def _sinaliza_crise(nota_geral: int | None, txt: str | None) -> bool:
    if nota_geral == 1 and txt:
        low = txt.lower()
        return any(k in low for k in _RISCO)
    return False

def _gerar_feedback(pont, educ, efic, apres, comentario, sentimento):
    partes = []
    def badge(nome, nota):
        return f"{nome}: {nota} ★" if nota is not None else None

    for nome, nota in (("Pontualidade", pont), ("Educação", educ), ("Eficiência", efic), ("Apresentação", apres)):
        b = badge(nome, nota)
        if b: partes.append(b)

    dicas = []
    if educ is not None and educ <= 2: dicas.append("melhore a abordagem/educação ao falar com o cliente")
    if pont is not None and pont <= 2: dicas.append("tente chegar no horário combinado")
    if efic is not None and efic <= 2: dicas.append("redobre o cuidado com o pedido durante o transporte")
    if apres is not None and apres <= 2: dicas.append("capriche na apresentação pessoal (higiene/uniforme)")

    txt = f"Notas — " + " | ".join(partes) if partes else "Obrigado pelo trabalho!"
    if comentario:
        txt += f". Cliente comentou: \"{comentario.strip()}\""
    if dicas:
        txt += ". Dica: " + "; ".join(dicas) + "."
    if sentimento == "positivo":
        txt += " 👏"
    return txt[:1000]

from datetime import datetime, date
from sqlalchemy import or_, and_

# routes_avisos.py
from datetime import datetime
from flask import Blueprint, render_template, redirect, request, url_for, abort
from flask_login import current_user


def _cooperado_atual() -> Cooperado | None:
    """
    Retorna o Cooperado do usuário logado, usando a sessão da aplicação.
    """
    uid = session.get("user_id")
    if not uid:
        return None
    return Cooperado.query.filter_by(usuario_id=uid).first()


# --- Blueprint Portal ---
portal_bp = Blueprint("portal", __name__, url_prefix="/portal")

@portal_bp.get("/avisos", endpoint="portal_cooperado_avisos")
@role_required("cooperado")
def avisos_list():
    coop = _cooperado_atual()
    if not coop:
        abort(403)

    # pega todos os avisos que se aplicam ao cooperado (seu helper)
    avisos = get_avisos_for_cooperado(coop)

    # busca leituras de uma vez (evita N+1)
    lidos_ids = {
        r.aviso_id
        for r in AvisoLeitura.query.filter_by(cooperado_id=coop.id).all()
    }

    # injeta flag lido para o template (sem tocar no banco)
    for a in avisos:
        a.lido = (a.id in lidos_ids)

    avisos_nao_lidos_count = sum(1 for a in avisos if not getattr(a, "lido", False))
    current_year = datetime.now().year

    return render_template(
        "portal_cooperado_avisos.html",
        avisos=avisos,
        avisos_nao_lidos_count=avisos_nao_lidos_count,
        current_year=current_year
    )

# === AVALIAÇÕES: Cooperado -> Restaurante (AJUSTADO) =========================
class AvaliacaoRestaurante(db.Model):
    __tablename__ = "avaliacoes_restaurante"

    id = db.Column(db.Integer, primary_key=True)

    restaurante_id = db.Column(db.Integer, db.ForeignKey("restaurantes.id"), nullable=False, index=True)
    cooperado_id   = db.Column(db.Integer, db.ForeignKey("cooperados.id"),   nullable=False, index=True)

    # 1 avaliação por lançamento (o cooperado só avalia uma vez aquele turno/expediente)
    lancamento_id  = db.Column(
        db.Integer,
        db.ForeignKey("lancamentos.id", ondelete="CASCADE"),
        unique=True,
        index=True,
        nullable=True,
    )

    # ===== SOMENTE 3 DIMENSÕES =====
    estrelas_ambiente   = db.Column(db.Integer)   # 1..5
    estrelas_tratamento = db.Column(db.Integer)   # 1..5
    estrelas_suporte    = db.Column(db.Integer)   # 1..5

    # Derivados
    estrelas_geral      = db.Column(db.Float)     # média arredondada 1 casa
    media_ponderada     = db.Column(db.Float)     # pode ser igual à geral (sem pesos)

    comentario          = db.Column(db.Text)
    sentimento          = db.Column(db.String(12))
    temas               = db.Column(db.String(255))
    alerta_crise        = db.Column(db.Boolean, default=False)

    criado_em = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    def recompute(self, pesos=(1, 1, 1)):
        vals = [self.estrelas_ambiente, self.estrelas_tratamento, self.estrelas_suporte]
        nums = [float(v) for v in vals if isinstance(v, (int, float)) and v is not None]
        if not nums:
            self.estrelas_geral = None
            self.media_ponderada = None
            return
        if pesos and len(pesos) == 3:
            total = 0.0
            wsum = 0.0
            for v, p in zip(vals, pesos):
                if v is not None:
                    total += float(v) * p
                    wsum += p
            media = (total / wsum) if wsum > 0 else (sum(nums) / len(nums))
        else:
            media = sum(nums) / len(nums)
        self.media_ponderada = round(media, 2)
        self.estrelas_geral = round(media, 1)

@portal_bp.post("/avisos/<int:aviso_id>/lido", endpoint="marcar_aviso_lido")
@role_required("cooperado")
def avisos_marcar_lido(aviso_id: int):
    coop = _cooperado_atual()
    if not coop:
        abort(403)

    aviso = Aviso.query.get_or_404(aviso_id)

    # idempotente: só cria se ainda não houver leitura
    ja_leu = AvisoLeitura.query.filter_by(
        cooperado_id=coop.id,
        aviso_id=aviso.id
    ).first()

    if not ja_leu:
        db.session.add(AvisoLeitura(
            cooperado_id=coop.id,
            aviso_id=aviso.id,
            lido_em=datetime.utcnow()
        ))
        db.session.commit()

    next_url = request.form.get("next") or (url_for("portal.portal_cooperado_avisos") + f"#aviso-{aviso.id}")
    return redirect(next_url)

@portal_bp.post("/avisos/marcar-todos", endpoint="marcar_todos_avisos_lidos")
@role_required("cooperado")
def avisos_marcar_todos():
    coop = _cooperado_atual()
    if not coop:
        abort(403)

    avisos = get_avisos_for_cooperado(coop)
    if not avisos:
        return redirect(url_for("portal.portal_cooperado_avisos"))
    ids_todos = {a.id for a in avisos}
    ids_ja_lidos = {
        r.aviso_id
        for r in AvisoLeitura.query.filter_by(cooperado_id=coop.id).all()
    }
    ids_pendentes = list(ids_todos - ids_ja_lidos)

    if ids_pendentes:
        db.session.bulk_save_objects([
            AvisoLeitura(cooperado_id=coop.id, aviso_id=aid, lido_em=datetime.utcnow())
            for aid in ids_pendentes
        ])
        db.session.commit()

    return redirect(url_for("portal.portal_cooperado_avisos"))

  # --- Registro do blueprint 'portal' (uma única vez, após definir TODAS as rotas dele)
# --- Registro do blueprint 'portal' (depois de definir TODAS as rotas do blueprint)
def register_blueprints_once(app):
    if "portal" not in app.blueprints:
        app.register_blueprint(portal_bp)

register_blueprints_once(app)

# --- Alias para compatibilidade com o template (endpoint esperado: 'portal_cooperado_avisos')
from flask import redirect, url_for

def _portal_cooperado_avisos_alias():
    # redireciona para a rota real dentro do blueprint 'portal'
    return redirect(url_for("portal.portal_cooperado_avisos"))

# publica a URL "antiga" (ajuste o path se o seu antigo era outro)
app.add_url_rule(
    "/portal/cooperado/avisos",         # caminho acessado
    endpoint="portal_cooperado_avisos", # nome que o template usa no url_for(...)
    view_func=_portal_cooperado_avisos_alias,
    methods=["GET"],
)

# ======== Helpers p/ troca: data/weekday/turno ========
def _parse_data_escala_str(s: str) -> date | None:
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', str(s or ''))
    if not m:
        return None
    d_, mth, y = map(int, m.groups())
    if y < 100:
        y += 2000
    try:
        return date(y, mth, d_)
    except Exception:
        return None

def _weekday_from_data_str(s: str) -> int | None:
    dt = _parse_data_escala_str(s)
    if dt:
        return (dt.weekday() % 7) + 1
    txt = unicodedata.normalize("NFD", str(s or "").lower())
    txt = "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")
    M = re.search(
        r"\b(seg|segunda|ter|terca|terça|qua|quarta|qui|quinta|sex|sexta|sab|sabado|sábado|dom|domingo)\b",
        txt
    )
    if not M:
        M = re.search(r"\b(seg|ter|qua|qui|sex|sab|dom)\b", txt)
        if not M:
            return None
    token = M.group(1)
    mapa = {
        "seg":1,"segunda":1,
        "ter":2,"terca":2,"terça":2,
        "qua":3,"quarta":3,
        "qui":4,"quinta":4,
        "sex":5,"sexta":5,
        "sab":6,"sabado":6,"sábado":6,
        "dom":7,"domingo":7,
    }
    return mapa.get(token)

def _weekday_abbr(num: int | None) -> str:
    return {1:"SEG",2:"TER",3:"QUA",4:"QUI",5:"SEX",6:"SÁB",7:"DOM"}.get(num or 0, "")

def _turno_bucket(turno: str | None, horario: str | None) -> str:
    t = (turno or "").lower()
    t = unicodedata.normalize("NFD", t)
    t = "".join(ch for ch in t if unicodedata.category(ch) != "Mn")
    if "noite" in t or "noturn" in t:
        return "noite"
    if any(x in t for x in ["dia", "diurn", "manha", "manhã", "tarde"]):
        return "dia"
    m = re.search(r'(\d{1,2}):(\d{2})', str(horario or ""))
    if m:
        h = int(m.group(1))
        return "noite" if (h >= 17 or h <= 6) else "dia"
    return "dia"

def _escala_label(e: Escala | None) -> str:
    if not e:
        return "—"
    wd = _weekday_from_data_str(e.data)
    wd_abbr = _weekday_abbr(wd)
    dt = _parse_data_escala_str(e.data)
    if dt:
        data_txt = dt.strftime("%d/%m/%y") + (f"-{wd_abbr}" if wd_abbr else "")
    else:
        data_txt = (str(e.data or "").strip() or wd_abbr)
    turno_txt = (e.turno or "").strip() or _turno_bucket(e.turno, e.horario).upper()
    horario_txt = (e.horario or "").strip()
    contrato_txt = (e.contrato or "").strip()
    parts = [x for x in [data_txt, turno_txt, horario_txt, contrato_txt] if x]
    return " • ".join(parts)

HIST_RETENTION_DAYS = 31


def _history_cutoff_dt() -> datetime:
    return datetime.utcnow() - timedelta(days=HIST_RETENTION_DAYS)


def _safe_coop_nome_by_id(coop_id: int | None) -> str:
    if not coop_id:
        return ""
    try:
        c = Cooperado.query.get(int(coop_id))
        return (c.nome or "").strip() if c else ""
    except Exception:
        return ""


def _prune_histories() -> None:
    cutoff = _history_cutoff_dt()
    try:
        EscalaHistorico.query.filter(EscalaHistorico.snapshot_em < cutoff).delete(synchronize_session=False)
    except Exception:
        pass
    try:
        TrocaHistorico.query.filter(TrocaHistorico.aplicada_em < cutoff).delete(synchronize_session=False)
    except Exception:
        pass


def _log_escala_historico(*, origem: str, acao: str, escala_ref_id: int | None = None, troca_ref_id: int | None = None,
                          grupo_ref: str | None = None, data: str = "", turno: str = "", horario: str = "",
                          contrato: str = "", cooperado_id: int | None = None, cooperado_nome: str | None = None,
                          saiu_nome: str | None = None, entrou_nome: str | None = None, admin_usuario_id: int | None = None,
                          snapshot_em: datetime | None = None) -> None:
    row = EscalaHistorico(
        grupo_ref=grupo_ref,
        origem=origem,
        acao=acao,
        escala_ref_id=escala_ref_id,
        troca_ref_id=troca_ref_id,
        admin_usuario_id=admin_usuario_id,
        data=data or "",
        turno=turno or "",
        horario=horario or "",
        contrato=contrato or "",
        cooperado_id=cooperado_id,
        cooperado_nome=(cooperado_nome or "").strip() or None,
        saiu_nome=(saiu_nome or "").strip() or None,
        entrou_nome=(entrou_nome or "").strip() or None,
        snapshot_em=snapshot_em or datetime.utcnow(),
    )
    db.session.add(row)


def _snapshot_escalas_atual(*, grupo_ref: str, origem: str = "upload", acao: str = "snapshot", admin_usuario_id: int | None = None,
                            snapshot_em: datetime | None = None) -> None:
    when = snapshot_em or datetime.utcnow()
    escalas = Escala.query.order_by(Escala.id.asc()).all()
    for e in escalas:
        nome = _safe_coop_nome_by_id(getattr(e, 'cooperado_id', None)) or (getattr(e, 'cooperado_nome', None) or "")
        _log_escala_historico(
            origem=origem,
            acao=acao,
            escala_ref_id=e.id,
            grupo_ref=grupo_ref,
            admin_usuario_id=admin_usuario_id,
            data=e.data or "",
            turno=e.turno or "",
            horario=e.horario or "",
            contrato=e.contrato or "",
            cooperado_id=e.cooperado_id,
            cooperado_nome=nome,
            snapshot_em=when,
        )


def _log_troca_historico_rows(troca_ref_id: int, linhas: list[dict], *, solicitante, destinatario, tipo: str, when: datetime | None = None) -> None:
    ts = when or datetime.utcnow()
    for linha in (linhas or []):
        turno_txt, horario_txt = "", ""
        raw = (linha.get('turno_horario') or '').strip()
        if '•' in raw:
            parts = [p.strip() for p in raw.split('•')]
            if parts:
                turno_txt = parts[0]
            if len(parts) > 1:
                horario_txt = parts[1]
        else:
            turno_txt = raw
        db.session.add(TrocaHistorico(
            troca_ref_id=troca_ref_id,
            tipo=tipo,
            solicitante_id=(getattr(solicitante, 'id', None) if solicitante else None),
            solicitante_nome=(getattr(solicitante, 'nome', None) if solicitante else None),
            destino_id=(getattr(destinatario, 'id', None) if destinatario else None),
            destino_nome=(getattr(destinatario, 'nome', None) if destinatario else None),
            data=(linha.get('dia') or ''),
            turno=turno_txt,
            horario=horario_txt,
            contrato=(linha.get('contrato') or ''),
            saiu_nome=(linha.get('saiu') or ''),
            entrou_nome=(linha.get('entrou') or ''),
            aplicada_em=ts,
        ))


def _parse_ymd_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value.strip(), '%Y-%m-%d').date()
    except Exception:
        return None


def _history_rows_between(q, col, ini: date | None, fim: date | None):
    if ini:
        q = q.filter(col >= datetime.combine(ini, dtime.min))
    if fim:
        q = q.filter(col <= datetime.combine(fim, dtime.max))
    return q


def _resolve_change_columns(rows: list[dict], hist_rows: list[EscalaHistorico]) -> list[dict]:
    latest = {}
    for h in hist_rows:
        key = ((h.data or '').strip(), (h.turno or '').strip(), (h.horario or '').strip(), (h.contrato or '').strip(), (h.cooperado_nome or '').strip())
        prev = latest.get(key)
        if prev is None or (h.snapshot_em or datetime.min) >= (prev.snapshot_em or datetime.min):
            latest[key] = h
    out = []
    for r in rows:
        key = ((r.get('data') or '').strip(), (r.get('turno') or '').strip(), (r.get('horario') or '').strip(), (r.get('contrato') or '').strip(), (r.get('cooperado_nome') or '').strip())
        h = latest.get(key)
        nr = dict(r)
        nr['saiu_nome'] = (getattr(h, 'saiu_nome', '') or '') if h else ''
        nr['entrou_nome'] = (getattr(h, 'entrou_nome', '') or '') if h else ''
        out.append(nr)
    return out


def _carry_forward_contrato(escalas: list[Escala]) -> dict[int, str]:
    eff = {}
    atual = ""
    for e in escalas:
        raw = (e.contrato or "").strip()
        if raw:
            atual = raw
        eff[e.id] = atual
    return eff

def _parse_linhas_from_msg(msg: str | None) -> list[dict]:
    if not msg:
        return []
    blobs = re.findall(r"__AFETACAO_JSON__\s*[:=]\s*(\{.*?\})\s*$", str(msg), flags=re.DOTALL)
    if not blobs:
        return []
    raw = blobs[-1]
    try:
        payload = json.loads(raw)
    except Exception:
        try:
            payload = json.loads(raw.replace("'", '\"'))
        except Exception:
            return []
    linhas = payload.get("linhas") or payload.get("rows") or []
    out = []
    for r in (linhas if isinstance(linhas, list) else []):
        turno = str(r.get("turno") or "").strip()
        horario = str(r.get("horario") or "").strip()
        turno_horario = (r.get("turno_horario") or " • ".join(x for x in [turno, horario] if x)).strip()
        out.append({
            "dia": str(r.get("dia") or ""),
            "turno_horario": turno_horario,
            "contrato": str(r.get("contrato") or ""),
            "saiu": str(r.get("saiu") or ""),
            "entrou": str(r.get("entrou") or ""),
        })
    return out

def _strip_afetacao_blob(msg: str | None) -> str:
    if not msg:
        return ""
    return re.sub(r"__AFETACAO_JSON__\s*[:=]\s*\{.*\}\s*$", "", str(msg), flags=re.DOTALL).strip()

def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or "").strip().lower())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s

def to_css_color(v: str) -> str:
    t = str(v or "").strip()
    if not t:
        return ""
    t_low = t.lower().strip()
    if re.fullmatch(r"[0-9a-fA-F]{8}", t):
        return f"#{t[2:8]}"
    if re.fullmatch(r"[0-9a-fA-F]{6}", t):
        return f"#{t}"
    if re.fullmatch(r"#?[0-9a-fA-F]{6,8}", t):
        if not t.startswith("#"):
            t = f"#{t}"
        if len(t) == 9:
            a = int(t[1:3], 16) / 255.0
            r = int(t[3:5], 16)
            g = int(t[5:7], 16)
            b = int(t[7:9], 16)
            return f"rgba({r},{g},{b},{a:.3f})"
        return t
    m = re.fullmatch(r"\s*(\d{1,3})\s*[,;]\s*(\d{1,3})\s*[,;]\s*(\d{1,3})\s*", t)
    if m:
        r, g, b = [max(0, min(255, int(x))) for x in m.groups()]
        return f"rgb({r},{g},{b})"
    mapa = {
        "azul": "blue", "vermelho": "red", "verde": "green",
        "amarelo": "yellow", "cinza": "gray", "preto": "black",
        "branco": "white", "laranja": "orange", "roxo": "purple",
    }
    return mapa.get(t_low, t)

# ---------- AVISOS: helpers ----------
from sqlalchemy import case, or_, and_, func
from sqlalchemy.orm import selectinload

def _avisos_base_query():
    # usa o relógio do banco; evita divergência de TZ/UTC da app
    now = func.now()
    return (
        Aviso.query
        .options(selectinload(Aviso.restaurantes), selectinload(Aviso.cooperados))  # evita N+1 no template
        .filter(Aviso.ativo.is_(True))
        .filter(or_(Aviso.inicio_em.is_(None), Aviso.inicio_em <= now))
        .filter(or_(Aviso.fim_em.is_(None),    Aviso.fim_em    >= now))
    )

# PRIORIDADE: "alta" (0), "media"/"média" (1), outras/NULL (2)
_PRIORD = case(
    (func.lower(Aviso.prioridade) == "alta", 0),
    (func.lower(Aviso.prioridade).in_(("media", "média")), 1),
    else_=2,
)


def _parse_datetime_local(value: str | None):
    raw = (value or "").strip()
    if not raw:
        return None
    raw = raw.replace("Z", "")
    for candidate in (raw, raw.replace("T", " ")):
        try:
            return datetime.fromisoformat(candidate)
        except Exception:
            pass
    try:
        d = _parse_date(raw)
        if d:
            return datetime.combine(d, dtime.min)
    except Exception:
        pass
    return None


def _aviso_destinatarios(aviso: Aviso, cooperados_all=None, restaurantes_all=None):
    cooperados_all = cooperados_all if cooperados_all is not None else Cooperado.query.order_by(Cooperado.nome.asc()).all()
    restaurantes_all = restaurantes_all if restaurantes_all is not None else Restaurante.query.order_by(Restaurante.nome.asc()).all()

    if aviso.tipo == "cooperado":
        alvo_ids = {c.id for c in (list(getattr(aviso, "cooperados", []) or []))}
        if alvo_ids:
            coops = [c for c in cooperados_all if c.id in alvo_ids]
        elif aviso.destino_cooperado_id:
            coops = [c for c in cooperados_all if c.id == aviso.destino_cooperado_id]
        else:
            coops = list(cooperados_all)
        rests = []
    elif aviso.tipo == "restaurante":
        alvo_ids = {r.id for r in (list(getattr(aviso, "restaurantes", []) or []))}
        rests = [r for r in restaurantes_all if (not alvo_ids or r.id in alvo_ids)]
        coops = []
    elif aviso.tipo == "global":
        coops = list(cooperados_all)
        rests = list(restaurantes_all)
    else:
        coops = []
        rests = []
    return coops, rests


def get_avisos_for_cooperado(coop: Cooperado):
    q = (
        _avisos_base_query()
        .filter(
            or_(
                (Aviso.tipo == "global"),
                and_(
                    Aviso.tipo == "cooperado",
                    or_(
                        Aviso.cooperados.any(Cooperado.id == coop.id),
                        Aviso.destino_cooperado_id == coop.id,
                        and_(
                            Aviso.destino_cooperado_id.is_(None),
                            ~Aviso.cooperados.any(),
                        ),
                    ),
                ),
            )
        )
        .order_by(
            Aviso.fixado.desc(),
            _PRIORD.asc(),
            Aviso.criado_em.desc(),
        )
    )
    return q.all()

def get_avisos_for_restaurante(rest: Restaurante):
    """
    RESTAURANTE vê:
      - global
      - restaurante (broadcast ou destinado a ESTE restaurante)
    """
    q = (
        _avisos_base_query()
        .filter(
            or_(
                (Aviso.tipo == "global"),
                and_(
                    Aviso.tipo == "restaurante",
                    or_(
                        ~Aviso.restaurantes.any(),                  # broadcast
                        Aviso.restaurantes.any(Restaurante.id == rest.id),  # específico
                    ),
                ),
            )
        )
        .order_by(
            Aviso.fixado.desc(),
            _PRIORD.asc(),
            Aviso.criado_em.desc(),
        )
    )

    return q.all()

# =========================
# Rotas de mídia (fotos armazenadas no banco)
# =========================
from flask import Response

def _send_bytes_with_cache(data: bytes, mime: str, filename: str):
    """Envia bytes com Cache-Control/ETag para evitar hits repetidos no banco."""
    if not data:
        abort(404)
    rv = send_file(
        io.BytesIO(data),
        mimetype=(mime or "application/octet-stream"),
        as_attachment=False,
        download_name=filename,
        max_age=60 * 60 * 24 * 7,  # 7 dias
        conditional=True,          # habilita ETag/If-None-Match
        last_modified=None,
        etag=True,
    )
    # Cache explícito (defensivo)
    rv.headers["Cache-Control"] = "public, max-age=604800, immutable"
    return rv

@app.get("/media/coop/<int:coop_id>")
def media_coop(coop_id: int):
    """
    Serve a foto do cooperado.

    Ordem de tentativa:
    1) foto_url salva em disco persistente/static
    2) foto_bytes antigo no banco (compatibilidade)
    3) imagem padrão
    """
    try:
        c = Cooperado.query.get_or_404(coop_id)
    except OperationalError:
        db.session.rollback()
        c = Cooperado.query.get_or_404(coop_id)

    # 1) Novo padrão: arquivo salvo em disco
    foto_url = (getattr(c, "foto_url", None) or "").strip()
    if foto_url:
        try:
            return _serve_uploaded(
                foto_url,
                download_name=(c.foto_filename or f"coop_{coop_id}.jpg"),
                force_download=False,
            )
        except Exception:
            # se a URL existir no banco mas o arquivo sumiu, tenta fallback antigo
            pass

    # 2) Compatibilidade com fotos antigas salvas no banco
    if getattr(c, "foto_bytes", None):
        return _send_bytes_with_cache(
            c.foto_bytes,
            c.foto_mime or "image/jpeg",
            c.foto_filename or f"coop_{coop_id}.jpg",
        )

    # 3) Fallback final
    return redirect(url_for("static", filename="img/default.png"))


@app.get("/media/rest/<int:rest_id>")
def media_rest(rest_id: int):
    """
    Serve a foto do restaurante.

    Ordem de tentativa:
    1) foto_url salva em disco persistente/static
    2) foto_bytes antigo no banco (compatibilidade)
    3) imagem padrão
    """
    try:
        r = Restaurante.query.get_or_404(rest_id)
    except OperationalError:
        db.session.rollback()
        r = Restaurante.query.get_or_404(rest_id)

    # 1) Novo padrão: arquivo salvo em disco
    foto_url = (getattr(r, "foto_url", None) or "").strip()
    if foto_url:
        try:
            return _serve_uploaded(
                foto_url,
                download_name=(r.foto_filename or f"rest_{rest_id}.jpg"),
                force_download=False,
            )
        except Exception:
            # se a URL existir no banco mas o arquivo sumiu, tenta fallback antigo
            pass

    # 2) Compatibilidade com fotos antigas salvas no banco
    if getattr(r, "foto_bytes", None):
        return _send_bytes_with_cache(
            r.foto_bytes,
            r.foto_mime or "image/jpeg",
            r.foto_filename or f"rest_{rest_id}.jpg",
        )

    # 3) Fallback final
    return redirect(url_for("static", filename="img/default.png"))
# =========================
# Rota raiz
# =========================
@app.route("/")
def index():
    uid = session.get("user_id")
    if not uid:
        return redirect(url_for("login"))
    u = Usuario.query.get(uid)
    if not u:
        return redirect(url_for("login"))
    if u.tipo == "admin":
        return redirect(url_for("admin_dashboard"))
    if u.tipo == "cooperado":
        return redirect(url_for("portal_cooperado"))
    if u.tipo == "restaurante":
        return redirect(url_for("portal_restaurante"))
    return redirect(url_for("login"))

# =========================
# Auth
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    erro_login = None

    if request.method == "POST":
        usuario = (request.form.get("usuario") or "").strip()
        senha = request.form.get("senha") or ""

        u = Usuario.query.filter_by(usuario=usuario).first()

        # fallback: login pelo nome do restaurante
        if not u:
            r = (
                Restaurante.query.filter(Restaurante.nome.ilike(usuario)).first()
                or Restaurante.query.filter(Restaurante.nome.ilike(f"%{usuario}%")).first()
            )
            if r and r.usuario_ref:
                u = r.usuario_ref

        if u and u.check_password(senha):
            # corrige registros antigos com ativo nulo
            if getattr(u, "ativo", None) is None:
                u.ativo = True
                db.session.commit()

            if u.ativo is False:
                session.clear()
                flash("Conta desativada. Fale com o administrador.", "danger")
                return redirect(url_for("login"))

            session.clear()
            session.permanent = True
            session["user_id"] = u.id
            session["user_tipo"] = u.tipo

            if u.tipo == "admin":
                return redirect(url_for("admin_dashboard"))
            elif u.tipo == "cooperado":
                return redirect(url_for("portal_cooperado"))
            elif u.tipo == "restaurante":
                return redirect(url_for("portal_restaurante"))

            session.clear()
            flash("Tipo de usuário inválido.", "danger")
            return redirect(url_for("login"))

        erro_login = "Usuário/senha inválidos."
        flash(erro_login, "danger")

    login_tpl = os.path.join(app.template_folder or "templates", "login.html")
    if os.path.exists(login_tpl):
        return render_template("login.html", erro_login=erro_login)

    return """
    <form method="POST" style="max-width:320px;margin:80px auto;font-family:Arial">
      <h3>Login</h3>
      <input name="usuario" placeholder="Usuário" style="width:100%;padding:10px;margin:6px 0">
      <input name="senha" type="password" placeholder="Senha" style="width:100%;padding:10px;margin:6px 0">
      <button style="padding:10px 16px">Entrar</button>
    </form>
    """

@app.route("/logout")
def logout():
    session.clear()
    flash("Você saiu do sistema.", "info")
    return redirect(url_for("login"))


@app.get("/sso/entrar")
def sso_entrar():
    token = (request.args.get("token") or "").strip()
    if not token:
        return redirect(url_for("login"))

    try:
        data = sso_load(token, max_age_seconds=45)
    except SignatureExpired:
        flash("Link expirou. Clique novamente no botão.", "danger")
        return redirect(url_for("login"))
    except BadSignature:
        flash("Link inválido.", "danger")
        return redirect(url_for("login"))

    if data.get("aud") != "painel-destino":
        flash("Token com destino inválido.", "danger")
        return redirect(url_for("login"))

    tipo = (data.get("tipo") or "admin").strip().lower()
    if tipo not in ("admin", "supervisao"):
        tipo = "admin"

    u = _get_or_create_sso_user(tipo=tipo)

    session.clear()
    session.permanent = True
    session["user_id"] = u.id
    session["user_tipo"] = u.tipo

    next_url = data.get("next") or url_for("admin_dashboard")
    return redirect(next_url)
    
def _safe_float(v, default=0.0):
    try:
        if v in (None, ''):
            return float(default)
        return float(v)
    except Exception:
        return float(default)


def _receita_total_real(r: ReceitaCooperativa) -> float:
    if getattr(r, 'auto_taxa_adm', False):
        return round(_safe_float(getattr(r, 'valor_pago', 0.0)) + _safe_float(getattr(r, 'valor_multa', 0.0)) + _safe_float(getattr(r, 'valor_juros', 0.0)), 2)
    return round(_safe_float(getattr(r, 'valor_total', 0.0)), 2)


def _calc_taxa_admin_encargos(valor_principal: float, data_vencimento: date | None, data_pagamento: date | None = None, multa_percentual: float = 2.0, juros_dia_percentual: float = 0.03):
    valor_principal = round(_safe_float(valor_principal), 2)
    multa_percentual = _safe_float(multa_percentual, 2.0)
    juros_dia_percentual = _safe_float(juros_dia_percentual, 0.03)
    ref = data_pagamento or date.today()
    if not data_vencimento:
        return {
            'dias_atraso': 0,
            'valor_multa': 0.0,
            'valor_juros': 0.0,
            'valor_total': valor_principal,
        }
    dias = (ref - data_vencimento).days
    dias = dias if dias > 0 else 0
    multa = round(valor_principal * (multa_percentual / 100.0), 2) if dias > 0 else 0.0
    juros = round(valor_principal * (juros_dia_percentual / 100.0) * dias, 2) if dias > 0 else 0.0
    return {
        'dias_atraso': dias,
        'valor_multa': multa,
        'valor_juros': juros,
        'valor_total': round(valor_principal + multa + juros, 2),
    }


def _taxa_competencia_iter(data_base: date | None, months_back: int = 0):
    if not data_base:
        return []
    hoje = date.today()
    start = date(hoje.year, hoje.month, 1)
    end = date(hoje.year, 12, 1)
    cur = start
    items = []
    while cur <= end:
        last_day = (cur + relativedelta(months=1) - timedelta(days=1)).day
        day = min(data_base.day, last_day)
        venc = date(cur.year, cur.month, day)
        items.append((cur.strftime('%Y-%m'), venc))
        cur = cur + relativedelta(months=1)
    return items


def _ensure_taxas_admin_receitas(restaurantes: list[Restaurante], months_back: int = 0):
    restaurantes_validos = []
    wanted = set()
    for rest in restaurantes:
        valor = _safe_float(getattr(rest, 'taxa_admin_valor', 0.0))
        data_base = getattr(rest, 'taxa_admin_data_base', None)
        ativo_rest = bool(getattr(rest, 'ativo', True)) if hasattr(rest, 'ativo') else True
        if (not ativo_rest) or valor <= 0 or not data_base:
            continue
        comps = _taxa_competencia_iter(data_base, months_back=months_back)
        if not comps:
            continue
        restaurantes_validos.append((rest, valor, data_base, comps))
        for competencia, _ in comps:
            wanted.add((rest.id, competencia))

    if not restaurantes_validos:
        return

    rest_ids = [rest.id for rest, _, _, _ in restaurantes_validos]
    existing_rows = ReceitaCooperativa.query.filter(
        ReceitaCooperativa.auto_taxa_adm.is_(True),
        ReceitaCooperativa.restaurante_id.in_(rest_ids),
    ).all()
    existing = {(r.restaurante_id, (r.competencia or '')): r for r in existing_rows}

    changed = False
    for rest, valor, data_base, comps in restaurantes_validos:
        multa_p = _safe_float(getattr(rest, 'taxa_admin_multa_percentual', 2.0), 2.0)
        juros_p = _safe_float(getattr(rest, 'taxa_admin_juros_dia_percentual', 0.03), 0.03)
        for competencia, venc in comps:
            existente = existing.get((rest.id, competencia))
            if existente:
                if getattr(existente, 'data_vencimento', None) != venc:
                    existente.data_vencimento = venc
                    existente.data = venc
                    changed = True
                if (getattr(existente, 'descricao', None) or '') != f'Taxa administrativa - {rest.nome} - {competencia}':
                    existente.descricao = f'Taxa administrativa - {rest.nome} - {competencia}'
                    changed = True
                if abs(_safe_float(getattr(existente, 'valor_previsto', 0.0)) - valor) > 0.009:
                    existente.valor_previsto = valor
                    existente.valor_principal = valor
                    changed = True
                existente.multa_percentual = multa_p
                existente.juros_dia_percentual = juros_p
                continue

            novo = ReceitaCooperativa(
                descricao=f'Taxa administrativa - {rest.nome} - {competencia}',
                valor_total=0.0,
                data=venc,
                restaurante_id=rest.id,
                auto_taxa_adm=True,
                competencia=competencia,
                valor_previsto=valor,
                valor_principal=valor,
                valor_pago=0.0,
                valor_multa=0.0,
                valor_juros=0.0,
                data_vencimento=venc,
                status_pagamento='nao_pago',
                multa_percentual=multa_p,
                juros_dia_percentual=juros_p,
            )
            db.session.add(novo)
            changed = True

    if changed:
        db.session.commit()


def _build_taxa_admin_rows(receitas: list[ReceitaCooperativa]):
    rows = []
    total_previsto = total_recebido = total_multa = total_juros = total_em_aberto = 0.0
    for r in receitas:
        if not getattr(r, 'auto_taxa_adm', False):
            continue
        previsto = _safe_float(getattr(r, 'valor_previsto', None) or getattr(r, 'valor_principal', None) or getattr(r, 'valor_total', 0.0))
        pago = _safe_float(getattr(r, 'valor_pago', 0.0))
        status = (getattr(r, 'status_pagamento', None) or 'nao_pago').strip().lower()
        data_pag = getattr(r, 'data_pagamento', None)
        venc = getattr(r, 'data_vencimento', None) or getattr(r, 'data', None)
        calc = _calc_taxa_admin_encargos(previsto, venc, data_pag if status == 'pago' else None, _safe_float(getattr(r, 'multa_percentual', 2.0), 2.0), _safe_float(getattr(r, 'juros_dia_percentual', 0.03), 0.03))
        if status == 'pago':
            multa = _safe_float(getattr(r, 'valor_multa', 0.0))
            juros = _safe_float(getattr(r, 'valor_juros', 0.0))
        else:
            multa = calc['valor_multa']
            juros = calc['valor_juros']
        total = round(pago + ( _safe_float(getattr(r, 'valor_multa', 0.0)) if status == 'pago' else 0.0) + (_safe_float(getattr(r, 'valor_juros', 0.0)) if status == 'pago' else 0.0), 2)
        aberto = max(0.0, round(calc['valor_total'] - total, 2)) if status != 'pago' else 0.0
        rows.append({
            'obj': r,
            'id': r.id,
            'descricao': r.descricao,
            'restaurante_nome': (r.restaurante.nome if getattr(r, 'restaurante', None) else ''),
            'competencia': getattr(r, 'competencia', '') or '',
            'data_vencimento': venc,
            'data_pagamento': data_pag,
            'status': status,
            'valor_previsto': round(previsto, 2),
            'valor_pago': round(pago, 2),
            'valor_multa': round(multa, 2),
            'valor_juros': round(juros, 2),
            'valor_recebido': round(total, 2),
            'valor_em_aberto': round(aberto, 2),
            'dias_atraso': calc['dias_atraso'],
            'multa_percentual': _safe_float(getattr(r, 'multa_percentual', 2.0), 2.0),
            'juros_dia_percentual': _safe_float(getattr(r, 'juros_dia_percentual', 0.03), 0.03),
        })
        total_previsto += previsto
        total_recebido += total
        total_multa += round(multa, 2) if status == 'pago' else 0.0
        total_juros += round(juros, 2) if status == 'pago' else 0.0
        total_em_aberto += round(aberto, 2)
    rows.sort(key=lambda x: ((x['data_vencimento'] or date.min), x['restaurante_nome'].lower(), x['competencia']), reverse=True)
    return rows, {
        'previsto': round(total_previsto, 2),
        'recebido': round(total_recebido, 2),
        'multa': round(total_multa, 2),
        'juros': round(total_juros, 2),
        'em_aberto': round(total_em_aberto, 2),
    }

# =========================
# Admin Dashboard
# =========================

from flask import jsonify, request, render_template, session, flash, redirect, url_for
from sqlalchemy import func, inspect, or_
from sqlalchemy.exc import SQLAlchemyError, OperationalError, ProgrammingError
from datetime import date, timedelta
from collections import defaultdict, namedtuple
import re


@app.post("/admin/cooperados/<int:id>/toggle-status")
@admin_required
def toggle_status_cooperado(id):
    """
    Alterna o status 'ativo' do usuário vinculado ao cooperado.

    Observação crítica:
    - Se o campo/coluna 'ativo' ainda não existir no MODEL/DB, retorna erro orientando migração.
    """
    try:
        coop = db.session.get(Cooperado, id)
        if not coop or not getattr(coop, "usuario_ref", None):
            return jsonify(ok=False, error="Cooperado não encontrado"), 404

        user = coop.usuario_ref

        if not hasattr(user, "ativo"):
            return jsonify(
                ok=False,
                error="Campo 'ativo' ausente no modelo. Atualize o models.py (Usuario.ativo) e faça deploy."
            ), 500

        try:
            insp = inspect(db.engine)
            table = getattr(Usuario, "__tablename__", None)

            if not table:
                return jsonify(
                    ok=False,
                    error="Não foi possível identificar a tabela do modelo Usuario."
                ), 500

            cols = {c["name"] for c in insp.get_columns(table)}
            if "ativo" not in cols:
                return jsonify(
                    ok=False,
                    error=(
                        "Coluna 'ativo' ausente no banco. Faça a migração/ALTER TABLE em produção "
                        f"(tabela: {table})."
                    ),
                    table=table
                ), 409
        except Exception:
            pass

        atual = bool(getattr(user, "ativo", True))
        user.ativo = not atual

        db.session.commit()
        return jsonify(ok=True, ativo=bool(user.ativo))

    except (OperationalError, ProgrammingError):
        db.session.rollback()
        return jsonify(
            ok=False,
            error="Falha ao salvar: provável falta da coluna 'ativo' no banco. Faça migração/ALTER TABLE."
        ), 409
    except SQLAlchemyError:
        db.session.rollback()
        return jsonify(ok=False, error="Falha ao salvar no banco"), 500

@app.route("/admin/admins/<int:usuario_id>/toggle-status", methods=["POST"])
@admin_required
def admin_toggle_admin_status(usuario_id):
    if not is_admin_master():
        flash("Apenas o administrador master pode alterar o status de administradores.", "danger")
        return redirect(url_for("admin_dashboard", tab="config"))

    admin = Usuario.query.filter_by(id=usuario_id, tipo="admin").first_or_404()

    if admin.is_master:
        flash("O administrador master não pode ser desativado por esta tela.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    admin.ativo = not bool(admin.ativo)
    db.session.commit()

    if admin.ativo:
        flash("Administrador ativado com sucesso.", "success")
    else:
        flash("Administrador desativado com sucesso.", "success")

    return redirect(url_for("admin_dashboard", tab="config"))

@app.route("/admin/admins/<int:usuario_id>/delete", methods=["POST"])
@admin_perm_required("config", "editar")
def admin_delete_admin(usuario_id):
    if not is_admin_master():
        flash("Apenas o administrador master pode excluir administradores.", "danger")
        return redirect(url_for("admin_dashboard", tab="config"))

    admin = Usuario.query.filter_by(id=usuario_id, tipo="admin").first_or_404()

    if admin.is_master:
        flash("O administrador master não pode ser excluído.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    AdminPermissao.query.filter_by(usuario_id=admin.id).delete()
    db.session.delete(admin)
    db.session.commit()

    flash("Administrador excluído com sucesso.", "success")
    return redirect(url_for("admin_dashboard", tab="config"))
    

def _is_cooperado_ativo(coop) -> bool:
    try:
        u = getattr(coop, "usuario_ref", None)
        if u is not None and getattr(u, "ativo", None) is not None:
            return bool(u.ativo)
        return bool(getattr(coop, "ativo", True))
    except Exception:
        return True

def _is_restaurante_ativo(rest) -> bool:
    try:
        u = getattr(rest, "usuario_ref", None)
        if u is not None and getattr(u, "ativo", None) is not None:
            return bool(u.ativo)
        return bool(getattr(rest, "ativo", True))
    except Exception:
        return True

def _cooperados_todos_ordenados() -> list[Cooperado]:
    return Cooperado.query.order_by(Cooperado.nome.asc()).all()

def _cooperados_ativos_ordenados() -> list[Cooperado]:
    return [c for c in _cooperados_todos_ordenados() if _is_cooperado_ativo(c)]

def _restaurantes_todos_ordenados() -> list[Restaurante]:
    return Restaurante.query.order_by(Restaurante.nome.asc()).all()

def _restaurantes_ativos_ordenados() -> list[Restaurante]:
    return [r for r in _restaurantes_todos_ordenados() if _is_restaurante_ativo(r)]

@app.route("/admin", methods=["GET"])
@admin_required
def admin_dashboard():
    args = request.args
    active_tab = (args.get("tab") or "lancamentos").strip().lower()

    admin_logado = _usuario_logado()
    if not admin_logado:
        session.clear()
        flash("Sessão inválida. Faça login novamente.", "danger")
        return redirect(url_for("login"))

    if (admin_logado.tipo or "").strip().lower() != "admin":
        session.clear()
        flash("Acesso restrito ao administrador.", "danger")
        return redirect(url_for("login"))

    if active_tab not in ADMIN_ABAS:
        active_tab = "lancamentos"

    # monta o mapa de permissões logo no início
    if getattr(admin_logado, "is_master", False):
        admin_perms = {
            aba: {
                "ver": True,
                "criar": True,
                "editar": True,
                "excluir": True,
            }
            for aba in ADMIN_ABAS.keys()
        }
    else:
        admin_perms = get_admin_permissions_map(admin_logado.id)

        abas_liberadas = [
            aba
            for aba in ADMIN_ABAS.keys()
            if admin_perms.get(aba, {}).get("ver", False)
        ]

        if not abas_liberadas:
            session.clear()
            flash("Seu usuário admin está sem permissões liberadas. Fale com o administrador master.", "danger")
            return redirect(url_for("login"))

        # config sempre restrita ao master
        if active_tab == "config":
            flash("A aba de configurações é restrita ao administrador master.", "danger")
            return redirect(url_for("admin_dashboard", tab=abas_liberadas[0]))

        # se tentar abrir aba sem permissão, redireciona
        if active_tab not in abas_liberadas:
            flash("Você não tem permissão para acessar essa aba.", "warning")
            return redirect(url_for("admin_dashboard", tab=abas_liberadas[0]))

    def _pick_date(*keys):
        for k in keys:
            v = args.get(k)
            if v:
                d = _parse_date(v)
                if d:
                    return d
        return None

    data_inicio = _pick_date("resumo_inicio", "data_inicio")
    data_fim = _pick_date("resumo_fim", "data_fim")

    filtro_periodo_aplicado = bool(data_inicio or data_fim)

    if data_inicio and not data_fim:
        data_fim = data_inicio
    elif data_fim and not data_inicio:
        data_inicio = data_fim
    elif not data_inicio and not data_fim:
        hoje_ref = date.today()
        if active_tab == 'receitas':
            data_inicio = date(hoje_ref.year, hoje_ref.month, 1)
            data_fim = (data_inicio + relativedelta(months=1)) - timedelta(days=1)
        else:
            data_inicio = hoje_ref - timedelta(days=hoje_ref.weekday())
            data_fim = data_inicio + timedelta(days=6)

    restaurante_id = args.get("restaurante_id", type=int)
    cooperado_id = args.get("cooperado_id", type=int)
    considerar_periodo = bool(args.get("considerar_periodo"))
    dows = set(args.getlist("dow"))

    # =========================
    # Lançamentos
    # =========================
    lancamentos = []
    total_producoes = 0.0
    total_inss = 0.0
    total_sest = 0.0
    total_encargos = 0.0

    q = Lancamento.query

    if restaurante_id:
        q = q.filter(Lancamento.restaurante_id == restaurante_id)
    if cooperado_id:
        q = q.filter(Lancamento.cooperado_id == cooperado_id)
    if data_inicio:
        q = q.filter(Lancamento.data >= data_inicio)
    if data_fim:
        q = q.filter(Lancamento.data <= data_fim)

    lanc_base = q.order_by(Lancamento.data.desc(), Lancamento.id.desc()).all()

    if dows:
        lancamentos = [l for l in lanc_base if l.data and _dow(l.data) in dows]
    else:
        lancamentos = lanc_base

    if considerar_periodo and restaurante_id:
        rest = Restaurante.query.get(restaurante_id)
        if rest:
            mapa = {
                "seg-dom": {"1", "2", "3", "4", "5", "6", "7"},
                "sab-sex": {"6", "7", "1", "2", "3", "4", "5"},
                "sex-qui": {"5", "6", "7", "1", "2", "3", "4"},
            }
            permitidos = mapa.get(rest.periodo, {"1", "2", "3", "4", "5", "6", "7"})
            lancamentos = [l for l in lancamentos if l.data and _dow(l.data) in permitidos]

    total_producoes = sum((l.valor or 0.0) for l in lancamentos)
    total_inss = round(total_producoes * INSS_ALIQ, 2)
    total_sest = round(total_producoes * SEST_ALIQ, 2)
    total_encargos = round(total_inss + total_sest, 2)

    # =========================
    # Receitas / Despesas Coop
    # =========================
    receitas = []
    despesas = []
    total_receitas = 0.0
    total_despesas = 0.0

    if True:
        rq = ReceitaCooperativa.query
        dq = DespesaCooperativa.query

        if data_inicio:
            rq = rq.filter(ReceitaCooperativa.data >= data_inicio)
            dq = dq.filter(DespesaCooperativa.data >= data_inicio)
        if data_fim:
            rq = rq.filter(ReceitaCooperativa.data <= data_fim)
            dq = dq.filter(DespesaCooperativa.data <= data_fim)

        receitas = rq.order_by(
            ReceitaCooperativa.data.desc().nullslast(),
            ReceitaCooperativa.id.desc()
        ).all()

        despesas = dq.order_by(
            DespesaCooperativa.data.desc(),
            DespesaCooperativa.id.desc()
        ).all()

        total_receitas = sum(_receita_total_real(r) for r in receitas)
        total_despesas = sum((d.valor or 0.0) for d in despesas)

    # =========================
    # Receitas / Despesas Cooperado
    # =========================
    receitas_coop = []
    despesas_coop = []
    total_receitas_coop = 0.0
    total_despesas_coop = 0.0
    total_adiantamentos_coop = 0.0

    if True:
        rq2 = ReceitaCooperado.query
        dq2 = DespesaCooperado.query

        if data_inicio:
            rq2 = rq2.filter(ReceitaCooperado.data >= data_inicio)
        if data_fim:
            rq2 = rq2.filter(ReceitaCooperado.data <= data_fim)

        if data_inicio and data_fim:
            dq2 = dq2.filter(
                DespesaCooperado.data_inicio <= data_fim,
                DespesaCooperado.data_fim >= data_inicio,
            )
        elif data_inicio:
            dq2 = dq2.filter(DespesaCooperado.data_fim >= data_inicio)
        elif data_fim:
            dq2 = dq2.filter(DespesaCooperado.data_inicio <= data_fim)

        if cooperado_id:
            rq2 = rq2.filter(ReceitaCooperado.cooperado_id == cooperado_id)
            dq2 = dq2.filter(DespesaCooperado.cooperado_id == cooperado_id)

        somente_pendentes = bool((request.args.get('somente_pendentes') or '').strip())

        receitas_coop = rq2.order_by(
            ReceitaCooperado.data.desc(),
            ReceitaCooperado.id.desc()
        ).all()

        despesas_coop = dq2.order_by(
            DespesaCooperado.data_fim.desc().nullslast(),
            DespesaCooperado.id.desc()
        ).all()

        if somente_pendentes and cooperado_id:
            snap_pend = _compute_coop_debt_snapshot(cooperado_id, data_inicio, data_fim)
            pend_ids = {item['id'] for item in snap_pend['itens'] if item['status'] in ('pendente', 'parcial', 'a_descontar') and item['restante'] > 0}
            despesas_coop = [d for d in despesas_coop if d.id in pend_ids]

        total_receitas_coop = sum((r.valor or 0.0) for r in receitas_coop)
        total_despesas_coop = sum(
            (d.valor or 0.0) for d in despesas_coop
            if not getattr(d, "eh_adiantamento", False)
        )
        total_adiantamentos_coop = sum(
            (d.valor or 0.0) for d in despesas_coop
            if getattr(d, "eh_adiantamento", False)
        )

        despesa_snapshot_map = {}
        for _cid in {getattr(d, "cooperado_id", None) for d in despesas_coop if getattr(d, "cooperado_id", None)}:
            _snap = _compute_coop_debt_snapshot(_cid, data_inicio, data_fim)
            for _it in _snap["itens"]:
                despesa_snapshot_map[_it["id"]] = _it

    cfg = get_config()

    cooperados_todos = _cooperados_todos_ordenados()
    cooperados = [c for c in cooperados_todos if _is_cooperado_ativo(c)]
    restaurantes_todos = _restaurantes_todos_ordenados()
    restaurantes = [r for r in restaurantes_todos if _is_restaurante_ativo(r)]
    active_coop_ids = {c.id for c in cooperados}
    active_rest_ids = {r.id for r in restaurantes}

    lancamentos = [
        l for l in lancamentos
        if ((getattr(l, "cooperado_id", None) is None) or (l.cooperado_id in active_coop_ids))
        and ((getattr(l, "restaurante_id", None) is None) or (l.restaurante_id in active_rest_ids))
    ]
    total_producoes = sum((l.valor or 0.0) for l in lancamentos)
    total_inss = round(total_producoes * INSS_ALIQ, 2)
    total_sest = round(total_producoes * SEST_ALIQ, 2)
    total_encargos = round(total_inss + total_sest, 2)

    receitas_coop = [r for r in receitas_coop if getattr(r, "cooperado_id", None) in active_coop_ids]
    despesas_coop = [d for d in despesas_coop if getattr(d, "cooperado_id", None) in active_coop_ids]
    total_receitas_coop = sum((r.valor or 0.0) for r in receitas_coop)
    total_despesas_coop = sum((d.valor or 0.0) for d in despesas_coop if not getattr(d, "eh_adiantamento", False))
    total_adiantamentos_coop = sum((d.valor or 0.0) for d in despesas_coop if getattr(d, "eh_adiantamento", False))

    despesa_snapshot_map = {}
    for _cid in {getattr(d, "cooperado_id", None) for d in despesas_coop if getattr(d, "cooperado_id", None) in active_coop_ids}:
        _snap = _compute_coop_debt_snapshot(_cid, data_inicio, data_fim)
        for _it in _snap["itens"]:
            despesa_snapshot_map[_it["id"]] = _it

    _ensure_taxas_admin_receitas(restaurantes, months_back=0)

    # Recarrega SEMPRE as receitas/despesas após gerar taxas automáticas.
    # Assim, sem filtro manual, a aba de receitas já abre mostrando o mês atual,
    # e com filtro continua respeitando o período informado.
    rq = ReceitaCooperativa.query
    dq = DespesaCooperativa.query
    if data_inicio:
        rq = rq.filter(ReceitaCooperativa.data >= data_inicio)
        dq = dq.filter(DespesaCooperativa.data >= data_inicio)
    if data_fim:
        rq = rq.filter(ReceitaCooperativa.data <= data_fim)
        dq = dq.filter(DespesaCooperativa.data <= data_fim)

    receitas = rq.order_by(
        ReceitaCooperativa.data.desc().nullslast(),
        ReceitaCooperativa.id.desc(),
    ).all()
    despesas = dq.order_by(
        DespesaCooperativa.data.desc(),
        DespesaCooperativa.id.desc(),
    ).all()
    total_receitas = sum(_receita_total_real(r) for r in receitas)
    total_despesas = sum((d.valor or 0.0) for d in despesas)
    taxa_admin_rows, taxa_admin_totais = _build_taxa_admin_rows(receitas)
    juros_arrecadados_total = round(sum((r['valor_multa'] + r['valor_juros']) for r in taxa_admin_rows if r['status'] == 'pago'), 2)
    cooperados_map = {c.id: c for c in cooperados}

    # =========================
    # Documentos / status
    # =========================
    docinfo_map = {c.id: _build_docinfo(c) for c in cooperados}
    status_doc_por_coop = {
        c.id: {
            "cnh_ok": docinfo_map[c.id]["cnh"]["ok"],
            "placa_ok": docinfo_map[c.id]["placa"]["ok"],
        }
        for c in cooperados
    }

    # =========================
    # Escalas
    # =========================
    escalas_all = (
        db.session.query(Escala)
        .outerjoin(Cooperado, Escala.cooperado_id == Cooperado.id)
        .outerjoin(Usuario, Cooperado.usuario_id == Usuario.id)
        .filter(
            or_(
                Escala.cooperado_id.is_(None),
                Usuario.ativo.is_(True)
            )
        )
        .order_by(Escala.id.asc())
        .all()
    )

    esc_by_int = defaultdict(list)
    esc_by_str = defaultdict(list)

    for e in escalas_all:
        k_int = e.cooperado_id if e.cooperado_id is not None else 0
        esc_item = {
            "data": e.data,
            "turno": e.turno,
            "horario": e.horario,
            "contrato": e.contrato,
            "cor": getattr(e, "cor", None),
            "nome_planilha": getattr(e, "cooperado_nome", None),
        }
        esc_by_int[k_int].append(esc_item)
        esc_by_str[str(k_int)].append(esc_item)

    cont_rows = dict(
        db.session.query(Escala.cooperado_id, func.count(Escala.id))
        .outerjoin(Cooperado, Escala.cooperado_id == Cooperado.id)
        .outerjoin(Usuario, Cooperado.usuario_id == Usuario.id)
        .filter(
            or_(
                Escala.cooperado_id.is_(None),
                Usuario.ativo.is_(True)
            )
        )
        .group_by(Escala.cooperado_id)
        .all()
    )

    qtd_escalas_map = {c.id: int(cont_rows.get(c.id, 0)) for c in cooperados}
    qtd_sem_cadastro = int(cont_rows.get(None, 0))

    contratos_set = {((e.contrato or "").strip()) for e in escalas_all if (e.contrato or "").strip()}
    contratos_set.update({((r.nome or "").strip()) for r in restaurantes if (r.nome or "").strip()})
    contratos_escala_opcoes = sorted(contratos_set, key=lambda s: s.lower())

    escala_editor_rows = []
    for e in sorted(escalas_all, key=_escala_sort_key):
        coop_obj = None
        if e.cooperado_id:
            coop_obj = cooperados_map.get(e.cooperado_id)

        nome_atual = (coop_obj.nome if coop_obj else (e.cooperado_nome or "").strip())
        escala_editor_rows.append({
            "id": e.id,
            "data": e.data or "",
            "weekday_num": _escala_weekday_num(e.data),
            "weekday_label": _escala_weekday_label(e.data),
            "turno": e.turno or "",
            "horario": e.horario or "",
            "contrato": e.contrato or "",
            "cooperado_id": e.cooperado_id,
            "cooperado_nome": nome_atual or "",
            "cooperado_nome_livre": (e.cooperado_nome or "") if not coop_obj else "",
            "restaurante_id": e.restaurante_id,
            "cor": getattr(e, "cor", None),
        })

    escala_alertas_1h = _build_escala_alertas_1h(escalas_all, cooperados_map)

    # =========================
    # Gráficos
    # =========================
    sums = {}
    for l in lancamentos:
        if not l.data:
            continue
        key = l.data.strftime("%Y-%m")
        sums[key] = sums.get(key, 0.0) + (l.valor or 0.0)

    labels_ord = sorted(sums.keys())

    def _fmt_label(k: str) -> str:
        parts = k.split("-")
        if len(parts) == 2 and parts[0] and parts[1]:
            year, month = parts[0], parts[1]
            return f"{month}/{year[-2:]}"
        return k

    labels_fmt = [_fmt_label(k) for k in labels_ord]
    values = [round(sums[k], 2) for k in labels_ord]
    chart_data_lancamentos_coop = {"labels": labels_fmt, "values": values}
    chart_data_lancamentos_cooperados = {"labels": labels_fmt, "values": values}

    # =========================
    # Admin master / principal
    # =========================
    admin_user = (
        Usuario.query
        .filter_by(tipo="admin", is_master=True)
        .order_by(Usuario.id.asc())
        .first()
    )

    if not admin_user:
        admin_user = (
            Usuario.query
            .filter_by(tipo="admin")
            .order_by(Usuario.id.asc())
            .first()
        )

    # =========================
    # Folha
    # =========================
    folha_por_coop = []
    folha_inicio = None
    folha_fim = None

    if active_tab == "folha":
        folha_inicio = _parse_date(args.get("folha_inicio"))
        folha_fim = _parse_date(args.get("folha_fim"))

        if folha_inicio and not folha_fim:
            folha_fim = folha_inicio
        elif folha_fim and not folha_inicio:
            folha_inicio = folha_fim
        elif not folha_inicio and not folha_fim:
            hoje_ref = date.today()
            folha_inicio = hoje_ref - timedelta(days=hoje_ref.weekday())
            folha_fim = folha_inicio + timedelta(days=6)

        INSS_ALIQ_FOLHA = 0.04
        SEST_ALIQ_FOLHA = 0.005

        FolhaItem = namedtuple(
            "FolhaItem",
            "cooperado lancamentos receitas despesas bruto inss sest encargos outras_desp liquido"
        )

        for c in cooperados:
            l = (
                Lancamento.query.filter(
                    Lancamento.cooperado_id == c.id,
                    Lancamento.data >= folha_inicio,
                    Lancamento.data <= folha_fim,
                )
                .order_by(Lancamento.data.asc(), Lancamento.id.asc())
                .all()
            )

            r = (
                ReceitaCooperado.query.filter(
                    ReceitaCooperado.cooperado_id == c.id,
                    ReceitaCooperado.data >= folha_inicio,
                    ReceitaCooperado.data <= folha_fim,
                )
                .order_by(ReceitaCooperado.data.asc(), ReceitaCooperado.id.asc())
                .all()
            )

            d = (
                DespesaCooperado.query.filter(
                    (DespesaCooperado.cooperado_id == c.id) | (DespesaCooperado.cooperado_id.is_(None)),
                    DespesaCooperado.data_inicio <= folha_fim,
                    DespesaCooperado.data_fim >= folha_inicio,
                )
                .order_by(DespesaCooperado.data_inicio.asc(), DespesaCooperado.id.asc())
                .all()
            )

            bruto_lanc = sum((x.valor or 0) for x in l)
            inss = round(bruto_lanc * INSS_ALIQ_FOLHA, 2)
            sest = round(bruto_lanc * SEST_ALIQ_FOLHA, 2)
            encargos = round(inss + sest, 2)
            outras_desp = sum((x.valor or 0) for x in d)
            bruto_total = bruto_lanc + sum((x.valor or 0) for x in r)
            liquido = round(bruto_total - encargos - outras_desp, 2)

            for x in l:
                x.conta_inss = True
                x.isento_benef = False
                x.inss = round((x.valor or 0) * INSS_ALIQ_FOLHA, 2)
                x.sest = round((x.valor or 0) * SEST_ALIQ_FOLHA, 2)
                x.encargos = round((x.inss or 0) + (x.sest or 0), 2)

            folha_por_coop.append(
                FolhaItem(
                    cooperado=c,
                    lancamentos=l,
                    receitas=r,
                    despesas=d,
                    bruto=round(bruto_total, 2),
                    inss=inss,
                    sest=sest,
                    encargos=encargos,
                    outras_desp=round(outras_desp, 2),
                    liquido=liquido,
                )
            )

    # =========================
    # Benefícios
    # =========================
    def _tokenize(s: str):
        return [x.strip() for x in re.split(r"[;,]", s or "") if x.strip()]

    def _d(s):
        if not s:
            return None
        s = s.strip()
        try:
            if "/" in s:
                d_, m_, y_ = s.split("/")
                return date(int(y_), int(m_), int(d_))
            y_, m_, d_ = s.split("-")
            return date(int(y_), int(m_), int(d_))
        except Exception:
            return None

    b_ini = _d(request.args.get("b_ini"))
    b_fim = _d(request.args.get("b_fim"))
    coop_filter = request.args.get("coop_benef_id", type=int)

    historico_beneficios = []
    beneficios_view = []

    if True:
        if b_ini and not b_fim:
            b_fim = b_ini
        elif b_fim and not b_ini:
            b_ini = b_fim
        elif not b_ini and not b_fim:
            hoje_ref = date.today()
            b_ini = hoje_ref - timedelta(days=hoje_ref.weekday())
            b_fim = b_ini + timedelta(days=6)

        q_benef = BeneficioRegistro.query.filter(
            BeneficioRegistro.data_inicial <= b_fim,
            BeneficioRegistro.data_final >= b_ini,
        )

        historico_beneficios = q_benef.order_by(BeneficioRegistro.id.desc()).all()

        for b in historico_beneficios:
            nomes = _tokenize(b.recebedores_nomes or "")
            ids = _tokenize(b.recebedores_ids or "")

            recs = []
            for i, nome in enumerate(nomes):
                rid = None

                if i < len(ids) and str(ids[i]).isdigit():
                    try:
                        rid = int(ids[i])
                    except Exception:
                        rid = None

                if rid is not None and rid not in active_coop_ids:
                    continue

                if coop_filter and (rid is not None) and (rid != coop_filter):
                    continue

                recs.append({
                    "id": rid,
                    "nome": nome,
                })

            if coop_filter and not recs:
                continue

            beneficios_view.append({
                "id": b.id,
                "data_inicial": b.data_inicial,
                "data_final": b.data_final,
                "data_lancamento": b.data_lancamento,
                "tipo": b.tipo,
                "valor_total": b.valor_total or 0.0,
                "recebedores": recs,
            })

    # =========================
    # Trocas
    # =========================
    def _escala_desc(e):
        return _escala_label(e) if e else ""

    def _split_turno_horario(s: str) -> tuple[str, str]:
        if not s:
            return "", ""
        parts = [p.strip() for p in s.split("•")]
        if len(parts) == 2:
            return parts[0], parts[1]
        return s.strip(), ""

    def _linha_from_escala(e: Escala, saiu: str, entrou: str) -> dict:
        return {
            "dia": _escala_label(e).split(" • ")[0],
            "turno_horario": " • ".join(
                [x for x in [(e.turno or "").strip(), (e.horario or "").strip()] if x]
            ),
            "contrato": (e.contrato or "").strip(),
            "saiu": saiu,
            "entrou": entrou,
        }

    trocas_all = TrocaSolicitacao.query.order_by(TrocaSolicitacao.id.desc()).all()
    trocas_pendentes = []
    trocas_historico = []
    trocas_historico_flat = []

    for t in trocas_all:
        solicitante = Cooperado.query.get(t.solicitante_id)
        destinatario = Cooperado.query.get(t.destino_id)
        orig = Escala.query.get(t.origem_escala_id)

        linhas_afetadas = _parse_linhas_from_msg(t.mensagem) if t.status == "aprovada" else []

        if t.status == "aprovada" and not linhas_afetadas and orig and solicitante and destinatario:
            linhas_afetadas.append(_linha_from_escala(orig, saiu=solicitante.nome, entrou=destinatario.nome))

            wd_o = _weekday_from_data_str(orig.data)
            buck_o = _turno_bucket(orig.turno, orig.horario)
            candidatas = Escala.query.filter_by(cooperado_id=destinatario.id).all()
            best = None

            for e in candidatas:
                if _weekday_from_data_str(e.data) == wd_o and _turno_bucket(e.turno, e.horario) == buck_o:
                    if (orig.contrato or "").strip().lower() == (e.contrato or "").strip().lower():
                        best = e
                        break
                    if best is None:
                        best = e

            if best:
                linhas_afetadas.append(_linha_from_escala(best, saiu=destinatario.nome, entrou=solicitante.nome))

        destino_data = ""
        destino_turno = ""
        destino_horario = ""
        destino_contrato = ""

        if t.status == "aprovada" and linhas_afetadas and solicitante and destinatario:
            linha_dest = None
            for r_ in linhas_afetadas:
                if r_.get("saiu") == destinatario.nome and r_.get("entrou") == solicitante.nome:
                    linha_dest = r_
                    break

            if linha_dest:
                destino_data = linha_dest.get("dia", "")
                turno_txt, horario_txt = _split_turno_horario(linha_dest.get("turno_horario", ""))
                destino_turno = turno_txt
                destino_horario = horario_txt
                destino_contrato = linha_dest.get("contrato", "")

        if not destino_data and orig and destinatario:
            wd_o = _weekday_from_data_str(orig.data)
            buck_o = _turno_bucket(orig.turno, orig.horario)
            candidatas = Escala.query.filter_by(cooperado_id=destinatario.id).all()
            best = None

            for e in candidatas:
                if _weekday_from_data_str(e.data) == wd_o and _turno_bucket(e.turno, e.horario) == buck_o:
                    if (orig.contrato or "").strip().lower() == (e.contrato or "").strip().lower():
                        best = e
                        break
                    if best is None:
                        best = e

            if best:
                destino_data = best.data
                destino_turno = (best.turno or "").strip()
                destino_horario = (best.horario or "").strip()
                destino_contrato = (best.contrato or "").strip()

        item = {
            "id": t.id,
            "status": t.status,
            "mensagem": t.mensagem,
            "criada_em": t.criada_em,
            "aplicada_em": t.aplicada_em,
            "solicitante": solicitante,
            "destinatario": destinatario,
            "origem": orig,
            "destino": destinatario,
            "origem_desc": _escala_desc(orig),
            "origem_weekday": _weekday_from_data_str(orig.data) if orig else None,
            "origem_turno_bucket": _turno_bucket(orig.turno if orig else None, orig.horario if orig else None),
            "destino_data": destino_data,
            "destino_turno": destino_turno,
            "destino_horario": destino_horario,
            "destino_contrato": destino_contrato,
            "linhas_afetadas": linhas_afetadas,
        }

        if t.status == "aprovada" and linhas_afetadas:
            itens = []
            for r_ in linhas_afetadas:
                turno_txt, horario_txt = _split_turno_horario(r_.get("turno_horario", ""))
                itens.append(
                    {
                        "data": r_.get("dia", ""),
                        "turno": turno_txt,
                        "horario": horario_txt,
                        "contrato": r_.get("contrato", ""),
                        "saiu_nome": r_.get("saiu", ""),
                        "entrou_nome": r_.get("entrou", ""),
                    }
                )

                trocas_historico_flat.append(
                    {
                        "data": r_.get("dia", ""),
                        "turno": turno_txt,
                        "horario": horario_txt,
                        "contrato": r_.get("contrato", ""),
                        "saiu_nome": r_.get("saiu", ""),
                        "entrou_nome": r_.get("entrou", ""),
                        "aplicada_em": t.aplicada_em,
                    }
                )

            item["itens"] = itens

        if t.status == "pendente":
            trocas_pendentes.append(item)
        else:
            trocas_historico.append(item)

    admins = (
        Usuario.query
        .filter_by(tipo="admin", is_master=False)
        .order_by(Usuario.id.asc())
        .all()
    )

    admin_permissions_map = {}
    for adm in admins:
        admin_permissions_map[adm.id] = get_admin_permissions_map(adm.id)

    admins_secundarios = admins
    admins_permissoes = admin_permissions_map

    current_date = date.today()
    data_limite = date(current_date.year, 12, 31)

    escala_hist_inicio = _parse_ymd_date(args.get("escala_hist_inicio"))
    escala_hist_fim = _parse_ymd_date(args.get("escala_hist_fim"))
    trocas_hist_inicio = _parse_ymd_date(args.get("trocas_hist_inicio"))
    trocas_hist_fim = _parse_ymd_date(args.get("trocas_hist_fim"))

    escala_historico_rows = []
    escala_editor_rows_export = []
    trocas_historico_export = []
    contagem_contrato_turno = []

    if active_tab == "escalas":
        if not escala_hist_inicio and not escala_hist_fim:
            escala_hist_fim = current_date
            escala_hist_inicio = current_date - timedelta(days=30)
        elif escala_hist_inicio and not escala_hist_fim:
            escala_hist_fim = escala_hist_inicio
        elif escala_hist_fim and not escala_hist_inicio:
            escala_hist_inicio = escala_hist_fim

        escala_hist_q = _history_rows_between(EscalaHistorico.query, EscalaHistorico.snapshot_em, escala_hist_inicio, escala_hist_fim)
        escala_hist_rows_db = escala_hist_q.order_by(EscalaHistorico.snapshot_em.desc(), EscalaHistorico.id.desc()).all()
        for h in escala_hist_rows_db:
            escala_historico_rows.append({
                "snapshot_em": h.snapshot_em,
                "data": h.data or "",
                "turno": h.turno or "",
                "horario": h.horario or "",
                "contrato": h.contrato or "",
                "cooperado_nome": h.cooperado_nome or "",
                "saiu_nome": h.saiu_nome or "",
                "entrou_nome": h.entrou_nome or "",
                "origem": h.origem or "",
                "acao": h.acao or "",
            })

        hist_rows_for_current = _history_rows_between(EscalaHistorico.query.filter(EscalaHistorico.saiu_nome.isnot(None)), EscalaHistorico.snapshot_em, escala_hist_inicio, escala_hist_fim).all()
        escala_editor_rows_export = _resolve_change_columns(escala_editor_rows, hist_rows_for_current)
    else:
        escala_hist_fim = escala_hist_fim or current_date
        escala_hist_inicio = escala_hist_inicio or (current_date - timedelta(days=30))

    if active_tab == "trocas":
        if not trocas_hist_inicio and not trocas_hist_fim:
            trocas_hist_fim = current_date
            trocas_hist_inicio = current_date - timedelta(days=30)
        elif trocas_hist_inicio and not trocas_hist_fim:
            trocas_hist_fim = trocas_hist_inicio
        elif trocas_hist_fim and not trocas_hist_inicio:
            trocas_hist_inicio = trocas_hist_fim

        trocas_hist_q = _history_rows_between(TrocaHistorico.query, TrocaHistorico.aplicada_em, trocas_hist_inicio, trocas_hist_fim)
        trocas_historico_export = trocas_hist_q.order_by(TrocaHistorico.aplicada_em.desc(), TrocaHistorico.id.desc()).all()
    else:
        trocas_hist_fim = trocas_hist_fim or current_date
        trocas_hist_inicio = trocas_hist_inicio or (current_date - timedelta(days=30))


    # resumo por cooperado calculado no backend para evitar travar no JS
    resumo_coop_rows = []
    resumo_totais = {
        "prod": 0.0, "inss4": 0.0, "sest05": 0.0, "rec": 0.0,
        "des": 0.0, "adiant": 0.0, "a_receber": 0.0, "saldo_pendente": 0.0,
        "pend_programado": 0.0
    }
    for coop in cooperados:
        snap = _compute_coop_debt_snapshot(coop.id, data_inicio, data_fim)
        prod = sum((l.valor or 0.0) for l in lancamentos if getattr(l, "cooperado_id", None) == coop.id)
        rec = sum((r.valor or 0.0) for r in receitas_coop if getattr(r, "cooperado_id", None) == coop.id)
        inss4 = sum((l.valor or 0.0) * INSS_ALIQ for l in lancamentos if getattr(l, "cooperado_id", None) == coop.id)
        sest05 = sum((l.valor or 0.0) * SEST_ALIQ for l in lancamentos if getattr(l, "cooperado_id", None) == coop.id)
        des = round(snap.get("descontado_despesa", 0.0), 2)
        adiant = round(sum((d.valor or 0.0) for d in despesas_coop if getattr(d, "cooperado_id", None) == coop.id and bool(getattr(d, "eh_adiantamento", False))), 2)
        if prod or rec or des or adiant or snap["saldo_devedor"] or snap["a_descontar"]:
            _a_receber = round(max(0.0, snap["disponivel_auto_restante"]), 2)
            _saldo_pendente = round(snap["saldo_devedor"], 2)
            _pend_programado = round(snap["a_descontar"], 2)
            resumo_coop_rows.append({
                "id": coop.id,
                "nome": coop.nome,
                "prod": round(prod,2),
                "inss4": round(inss4,2),
                "sest05": round(sest05,2),
                "rec": round(rec,2),
                "des": round(des,2),
                "adiant": round(adiant,2),
                "a_receber": _a_receber,
                "aReceber": _a_receber,
                "saldo_pendente": _saldo_pendente,
                "saldoPendente": _saldo_pendente,
                "pend_programado": _pend_programado,
                "pendProgramado": _pend_programado,
            })
            resumo_totais["prod"] += prod
            resumo_totais["inss4"] += inss4
            resumo_totais["sest05"] += sest05
            resumo_totais["rec"] += rec
            resumo_totais["des"] += des
            resumo_totais["adiant"] += adiant
            resumo_totais["a_receber"] += max(0.0, snap["disponivel_auto_restante"])
            resumo_totais["saldo_pendente"] += snap["saldo_devedor"]
            resumo_totais["pend_programado"] += snap["a_descontar"]

    return render_template(
        "admin_dashboard.html",
        tab=active_tab,
        total_producoes=total_producoes,
        total_inss=total_inss,
        total_sest=total_sest,
        total_encargos=total_encargos,
        total_receitas=total_receitas,
        total_despesas=total_despesas,
        total_receitas_coop=total_receitas_coop,
        total_despesas_coop=total_despesas_coop,
        total_adiantamentos_coop=total_adiantamentos_coop,
        salario_minimo=(cfg.salario_minimo or 0.0) if cfg else 0.0,
        lancamentos=lancamentos,
        receitas=receitas,
        despesas=despesas,
        receitas_coop=receitas_coop,
        despesas_coop=despesas_coop,
        cooperados=cooperados,
        cooperados_todos=cooperados_todos,
        restaurantes=restaurantes,
        restaurantes_todos=restaurantes_todos,
        beneficios_view=beneficios_view,
        historico_beneficios=historico_beneficios,
        current_date=current_date,
        data_limite=data_limite,
        admin=admin_user,
        admin_user=admin_user,
        docinfo_map=docinfo_map,
        escalas_por_coop=esc_by_int,
        escalas_por_coop_json=esc_by_str,
        qtd_escalas_map=qtd_escalas_map,
        qtd_escalas_sem_cadastro=qtd_sem_cadastro,
        status_doc_por_coop=status_doc_por_coop,
        chart_data_lancamentos_coop=chart_data_lancamentos_coop,
        chart_data_lancamentos_cooperados=chart_data_lancamentos_cooperados,
        folha_inicio=folha_inicio,
        folha_fim=folha_fim,
        folha_por_coop=folha_por_coop,
        trocas_pendentes=trocas_pendentes,
        trocas_historico=trocas_historico,
        trocas_historico_flat=trocas_historico_flat,
        admin_perms=admin_perms,
        admin_is_master=is_admin_master(),
        ADMIN_ABAS=ADMIN_ABAS,
        admins=admins,
        admin_permissions_map=admin_permissions_map,
        admins_secundarios=admins_secundarios,
        admins_permissoes=admins_permissoes,
        filtro_periodo_aplicado=filtro_periodo_aplicado,
        escala_editor_rows=escala_editor_rows,
        escala_editor_rows_export=escala_editor_rows_export,
        contratos_escala_opcoes=contratos_escala_opcoes,
        escala_hist_inicio=escala_hist_inicio,
        escala_hist_fim=escala_hist_fim,
        trocas_hist_inicio=trocas_hist_inicio,
        trocas_hist_fim=trocas_hist_fim,
        escala_historico_rows=escala_historico_rows,
        trocas_historico_export=trocas_historico_export,
        contagem_contrato_turno=contagem_contrato_turno,
        resumo_coop_rows=resumo_coop_rows,
        resumo_totais=resumo_totais,
        despesa_snapshot_map=despesa_snapshot_map,
        taxa_admin_rows=taxa_admin_rows,
        taxa_admin_totais=taxa_admin_totais,
        juros_arrecadados_total=juros_arrecadados_total,
    )
    
# =========================
# Navegação/Export util
# =========================
@app.route("/filtrar_lancamentos")
@admin_required
def filtrar_lancamentos():
    qs = request.query_string.decode("utf-8")
    base = url_for("admin_dashboard", tab="lancamentos")
    joiner = "&" if qs else ""
    return redirect(f"{base}{joiner}{qs}")


from datetime import datetime, date

def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    return None

def _dow(d: date) -> str:
    return str((d.weekday() % 7) + 1) if d else ""

def _fmt_time(t) -> str:
    if not t:
        return ""
    if hasattr(t, "strftime"):
        try:
            return t.strftime("%H:%M")
        except Exception:
            pass
    return str(t)

@app.route("/exportar_lancamentos")
@admin_required
def exportar_lancamentos():
    import io
    from collections import defaultdict

    from flask import request, send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    # -----------------------
    # Filtros
    # -----------------------
    args = request.args
    restaurante_id = args.get("restaurante_id", type=int)
    cooperado_id   = args.get("cooperado_id", type=int)
    data_inicio    = _parse_date(args.get("data_inicio"))
    data_fim       = _parse_date(args.get("data_fim"))
    dows           = set(args.getlist("dow"))  # '0'..'6'

    q = Lancamento.query
    if restaurante_id:
        q = q.filter(Lancamento.restaurante_id == restaurante_id)
    if cooperado_id:
        q = q.filter(Lancamento.cooperado_id == cooperado_id)
    if data_inicio:
        q = q.filter(Lancamento.data >= data_inicio)
    if data_fim:
        q = q.filter(Lancamento.data <= data_fim)

    lancs = q.order_by(Lancamento.data.desc(), Lancamento.id.desc()).all()
    if dows:
        lancs = [l for l in lancs if l.data and _dow(l.data) in dows]

    # ===============================
    # Estilos
    # ===============================
    wb = Workbook()

    bold        = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="DDDDDD")

    currency_fmt = "#,##0.00"
    date_fmt     = "DD/MM/YYYY"

    def _style_header(ws, ncols: int):
        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill

    def _autosize(ws, max_col, max_row, cap=55):
        widths = [0] * (max_col + 1)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > widths[c]:
                    widths[c] = len(s)
        for c in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = min(max(10, widths[c] + 2), cap)

    # ===============================
    # ABA 1 - Lançamentos (detalhado)
    # ===============================
    ws_det = wb.active
    ws_det.title = "Lançamentos"

    # Agora exporta INSS e SEST separados + Encargos
    header_det = [
        "Restaurante", "Periodo", "Cooperado", "Descricao",
        "Valor", "Data", "HoraInicio", "HoraFim",
        "INSS", "SEST", "Encargos", "Liquido",
    ]
    ws_det.append(header_det)
    _style_header(ws_det, ncols=len(header_det))

    # ===============================
    # Estruturas de soma
    # ===============================
    totais_contrato = defaultdict(lambda: {
        "restaurante": "", "periodo": "",
        "bruto": 0.0, "inss": 0.0, "sest": 0.0, "enc": 0.0, "liq": 0.0,
    })
    totais_contrato_coop = defaultdict(lambda: {
        "restaurante": "", "periodo": "", "cooperado": "",
        "bruto": 0.0, "inss": 0.0, "sest": 0.0, "enc": 0.0, "liq": 0.0,
    })
    totais_coop = defaultdict(lambda: {
        "cooperado": "",
        "bruto": 0.0, "inss": 0.0, "sest": 0.0, "enc": 0.0, "liq": 0.0,
    })
    totais_coop_dia = defaultdict(lambda: {
        "cooperado": "", "data": None, "restaurante": "", "periodo": "",
        "bruto": 0.0, "inss": 0.0, "sest": 0.0, "enc": 0.0, "liq": 0.0,
    })

    total_geral_bruto = 0.0
    total_geral_inss  = 0.0
    total_geral_sest  = 0.0
    total_geral_enc   = 0.0
    total_geral_liq   = 0.0

    # ===============================
    # Preenche lançamentos + somatórios
    # ===============================
    for l in lancs:
        v = float(l.valor or 0.0)

        inss = v * 0.04
        sest = v * 0.005
        encargos = inss + sest
        liq = v - encargos

        rest_nome   = l.restaurante.nome if getattr(l, "restaurante", None) else ""
        rest_period = l.restaurante.periodo if getattr(l, "restaurante", None) else ""
        rest_id     = int(getattr(l, "restaurante_id", 0) or 0)

        coop_nome = l.cooperado.nome if getattr(l, "cooperado", None) else ""
        coop_id   = int(getattr(l, "cooperado_id", 0) or 0)

        row = [
            rest_nome,
            rest_period,
            coop_nome,
            (l.descricao or ""),
            v,
            l.data,
            _fmt_time(getattr(l, "hora_inicio", None)),
            _fmt_time(getattr(l, "hora_fim", None)),
            inss,
            sest,
            encargos,
            liq,
        ]
        ws_det.append(row)
        r = ws_det.max_row

        # formatos
        ws_det.cell(row=r, column=5).number_format  = currency_fmt
        ws_det.cell(row=r, column=6).number_format  = date_fmt
        ws_det.cell(row=r, column=9).number_format  = currency_fmt
        ws_det.cell(row=r, column=10).number_format = currency_fmt
        ws_det.cell(row=r, column=11).number_format = currency_fmt
        ws_det.cell(row=r, column=12).number_format = currency_fmt

        # ---- Totais por contrato
        key_contrato = (rest_id, rest_nome, rest_period)
        tc = totais_contrato[key_contrato]
        tc["restaurante"] = rest_nome
        tc["periodo"]     = rest_period
        tc["bruto"]      += v
        tc["inss"]       += inss
        tc["sest"]       += sest
        tc["enc"]        += encargos
        tc["liq"]        += liq

        # ---- Totais por contrato + cooperado
        key_contrato_coop = (rest_id, rest_nome, rest_period, coop_id, coop_nome)
        tcc = totais_contrato_coop[key_contrato_coop]
        tcc["restaurante"] = rest_nome
        tcc["periodo"]     = rest_period
        tcc["cooperado"]   = coop_nome
        tcc["bruto"]      += v
        tcc["inss"]       += inss
        tcc["sest"]       += sest
        tcc["enc"]        += encargos
        tcc["liq"]        += liq

        # ---- Totais por cooperado
        key_coop = (coop_id, coop_nome)
        tcg = totais_coop[key_coop]
        tcg["cooperado"] = coop_nome
        tcg["bruto"]    += v
        tcg["inss"]     += inss
        tcg["sest"]     += sest
        tcg["enc"]      += encargos
        tcg["liq"]      += liq

        # ---- Totais por cooperado e dia
        key_coop_dia = (coop_id, coop_nome, l.data, rest_id, rest_nome, rest_period)
        tcd = totais_coop_dia[key_coop_dia]
        tcd["cooperado"]   = coop_nome
        tcd["data"]        = l.data
        tcd["restaurante"] = rest_nome
        tcd["periodo"]     = rest_period
        tcd["bruto"]      += v
        tcd["inss"]       += inss
        tcd["sest"]       += sest
        tcd["enc"]        += encargos
        tcd["liq"]        += liq

        # ---- Total geral
        total_geral_bruto += v
        total_geral_inss  += inss
        total_geral_sest  += sest
        total_geral_enc   += encargos
        total_geral_liq   += liq

    ws_det.freeze_panes = "A2"
    ws_det.auto_filter.ref = f"A1:{get_column_letter(len(header_det))}{ws_det.max_row}"
    _autosize(ws_det, max_col=len(header_det), max_row=min(ws_det.max_row, 3000))

    # ===============================
    # ABA 2 - Totais por Contrato
    # ===============================
    ws_con = wb.create_sheet("Totais por Contrato")
    header_contrato = [
        "Restaurante", "Periodo",
        "Total Bruto", "Total INSS", "Total SEST", "Total Encargos", "Total Líquido"
    ]
    ws_con.append(header_contrato)
    _style_header(ws_con, ncols=len(header_contrato))

    soma_b = soma_inss = soma_sest = soma_enc = soma_l = 0.0
    row_idx = 2

    for _, tc in sorted(
        totais_contrato.items(),
        key=lambda x: (x[1]["restaurante"], x[1]["periodo"])
    ):
        ws_con.append([
            tc["restaurante"] or "—",
            tc["periodo"] or "—",
            tc["bruto"],
            tc["inss"],
            tc["sest"],
            tc["enc"],
            tc["liq"],
        ])
        r = row_idx
        for col in (3, 4, 5, 6, 7):
            ws_con.cell(row=r, column=col).number_format = currency_fmt

        soma_b    += tc["bruto"]
        soma_inss += tc["inss"]
        soma_sest += tc["sest"]
        soma_enc  += tc["enc"]
        soma_l    += tc["liq"]
        row_idx += 1

    if row_idx > 2:
        ws_con.append(["TOTAL GERAL", "", soma_b, soma_inss, soma_sest, soma_enc, soma_l])
        r = row_idx
        for col in (1, 3, 4, 5, 6, 7):
            cell = ws_con.cell(row=r, column=col)
            cell.font = bold
            if col != 1:
                cell.number_format = currency_fmt

    ws_con.freeze_panes = "A2"
    ws_con.auto_filter.ref = f"A1:{get_column_letter(len(header_contrato))}{ws_con.max_row}"
    _autosize(ws_con, max_col=len(header_contrato), max_row=ws_con.max_row)

    # ===============================
    # ABA 3 - Contrato x Cooperado
    # ===============================
    ws_cc = wb.create_sheet("Contrato x Cooperado")
    header_cc = [
        "Restaurante", "Periodo", "Cooperado",
        "Total Bruto", "Total INSS", "Total SEST", "Total Encargos", "Total Líquido"
    ]
    ws_cc.append(header_cc)
    _style_header(ws_cc, ncols=len(header_cc))

    row_idx = 2
    for _, tcc in sorted(
        totais_contrato_coop.items(),
        key=lambda x: (x[1]["restaurante"], x[1]["periodo"], x[1]["cooperado"])
    ):
        ws_cc.append([
            tcc["restaurante"] or "—",
            tcc["periodo"] or "—",
            tcc["cooperado"] or "—",
            tcc["bruto"],
            tcc["inss"],
            tcc["sest"],
            tcc["enc"],
            tcc["liq"],
        ])
        r = row_idx
        for col in (4, 5, 6, 7, 8):
            ws_cc.cell(row=r, column=col).number_format = currency_fmt
        row_idx += 1

    ws_cc.freeze_panes = "A2"
    ws_cc.auto_filter.ref = f"A1:{get_column_letter(len(header_cc))}{ws_cc.max_row}"
    _autosize(ws_cc, max_col=len(header_cc), max_row=ws_cc.max_row)

    # ===============================
    # ABA 4 - Cooperado por Dia
    # ===============================
    ws_cd = wb.create_sheet("Cooperado por Dia")
    header_cd = [
        "Cooperado", "Data", "Restaurante", "Periodo",
        "Total Bruto", "Total INSS", "Total SEST", "Total Encargos", "Total Líquido"
    ]
    ws_cd.append(header_cd)
    _style_header(ws_cd, ncols=len(header_cd))

    row_idx = 2
    for _, tcd in sorted(
        totais_coop_dia.items(),
        key=lambda x: (x[1]["cooperado"], x[1]["data"] or date.min, x[1]["restaurante"], x[1]["periodo"])
    ):
        ws_cd.append([
            tcd["cooperado"] or "—",
            tcd["data"],
            tcd["restaurante"] or "—",
            tcd["periodo"] or "—",
            tcd["bruto"],
            tcd["inss"],
            tcd["sest"],
            tcd["enc"],
            tcd["liq"],
        ])
        r = row_idx
        ws_cd.cell(row=r, column=2).number_format = date_fmt
        for col in (5, 6, 7, 8, 9):
            ws_cd.cell(row=r, column=col).number_format = currency_fmt
        row_idx += 1

    ws_cd.freeze_panes = "A2"
    ws_cd.auto_filter.ref = f"A1:{get_column_letter(len(header_cd))}{ws_cd.max_row}"
    _autosize(ws_cd, max_col=len(header_cd), max_row=ws_cd.max_row)

    # ===============================
    # ABA 5 - Totais por Cooperado
    # ===============================
    ws_tc = wb.create_sheet("Totais por Cooperado")
    header_tc = ["Cooperado", "Total Bruto", "Total INSS", "Total SEST", "Total Encargos", "Total Líquido"]
    ws_tc.append(header_tc)
    _style_header(ws_tc, ncols=len(header_tc))

    row_idx = 2
    for _, tcg in sorted(totais_coop.items(), key=lambda x: x[1]["cooperado"]):
        ws_tc.append([
            tcg["cooperado"] or "—",
            tcg["bruto"],
            tcg["inss"],
            tcg["sest"],
            tcg["enc"],
            tcg["liq"],
        ])
        r = row_idx
        for col in (2, 3, 4, 5, 6):
            ws_tc.cell(row=r, column=col).number_format = currency_fmt
        row_idx += 1

    if row_idx > 2:
        total_b = sum(v["bruto"] for v in totais_coop.values())
        total_i = sum(v["inss"]  for v in totais_coop.values())
        total_s = sum(v["sest"]  for v in totais_coop.values())
        total_e = sum(v["enc"]   for v in totais_coop.values())
        total_l = sum(v["liq"]   for v in totais_coop.values())
        ws_tc.append(["TOTAL GERAL", total_b, total_i, total_s, total_e, total_l])
        r = row_idx
        for col in (1, 2, 3, 4, 5, 6):
            cell = ws_tc.cell(row=r, column=col)
            cell.font = bold
            if col != 1:
                cell.number_format = currency_fmt

    ws_tc.freeze_panes = "A2"
    ws_tc.auto_filter.ref = f"A1:{get_column_letter(len(header_tc))}{ws_tc.max_row}"
    _autosize(ws_tc, max_col=len(header_tc), max_row=ws_tc.max_row)

    # ===============================
    # ABA 6 - Resumo Geral
    # ===============================
    ws_rg = wb.create_sheet("Resumo Geral")
    ws_rg["A1"] = "Total Geral Bruto"
    ws_rg["A2"] = "Total Geral INSS"
    ws_rg["A3"] = "Total Geral SEST"
    ws_rg["A4"] = "Total Geral Encargos"
    ws_rg["A5"] = "Total Geral Líquido"
    for a in ("A1", "A2", "A3", "A4", "A5"):
        ws_rg[a].font = bold

    ws_rg["B1"] = total_geral_bruto
    ws_rg["B2"] = total_geral_inss
    ws_rg["B3"] = total_geral_sest
    ws_rg["B4"] = total_geral_enc
    ws_rg["B5"] = total_geral_liq
    for b in ("B1", "B2", "B3", "B4", "B5"):
        ws_rg[b].number_format = currency_fmt

    ws_rg.column_dimensions["A"].width = 24
    ws_rg.column_dimensions["B"].width = 18

    # ===============================
    # Envio do arquivo
    # ===============================
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)

    return send_file(
        mem,
        as_attachment=True,
        download_name="lancamentos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# =========================
# CRUD Lançamentos (Admin)
# =========================
@app.route("/admin/lancamentos/add", methods=["POST"])
@admin_perm_required("lancamentos", "criar")
def admin_add_lancamento():
    f = request.form

    l = Lancamento(
        restaurante_id=f.get("restaurante_id", type=int),
        cooperado_id=f.get("cooperado_id", type=int),
        descricao=(f.get("descricao") or "").strip(),
        valor=f.get("valor", type=float),
        data=_parse_date(f.get("data")),
        hora_inicio=f.get("hora_inicio"),
        hora_fim=f.get("hora_fim"),
        qtd_entregas=f.get("qtd_entregas", type=int),
    )

    db.session.add(l)
    db.session.commit()
    flash("Lançamento inserido.", "success")
    return redirect(url_for("admin_dashboard", tab="lancamentos"))


@app.route("/admin/lancamentos/<int:id>/edit", methods=["POST"])
@admin_perm_required("lancamentos", "editar")
def admin_edit_lancamento(id):
    l = Lancamento.query.get_or_404(id)
    f = request.form

    l.restaurante_id = f.get("restaurante_id", type=int)
    l.cooperado_id = f.get("cooperado_id", type=int)
    l.descricao = (f.get("descricao") or "").strip()
    l.valor = f.get("valor", type=float)
    l.data = _parse_date(f.get("data"))
    l.hora_inicio = f.get("hora_inicio")
    l.hora_fim = f.get("hora_fim")
    l.qtd_entregas = f.get("qtd_entregas", type=int)

    db.session.commit()
    flash("Lançamento atualizado.", "success")
    return redirect(url_for("admin_dashboard", tab="lancamentos"))


@app.route("/admin/lancamentos/<int:id>/delete", methods=["GET", "POST"])
@admin_perm_required("lancamentos", "excluir")
def admin_delete_lancamento(id):
    l = Lancamento.query.get_or_404(id)

    db.session.execute(
        sa_delete(AvaliacaoCooperado).where(AvaliacaoCooperado.lancamento_id == id)
    )
    db.session.execute(
        sa_delete(AvaliacaoRestaurante).where(AvaliacaoRestaurante.lancamento_id == id)
    )

    db.session.delete(l)
    db.session.commit()
    flash("Lançamento excluído.", "success")
    return redirect(url_for("admin_dashboard", tab="lancamentos"))


# =========================
# Avaliações (Admin)
# =========================
@app.route("/admin/avaliacoes", methods=["GET"])
@admin_perm_required("avaliacoes", "ver")
def admin_avaliacoes():
    tipo_raw = (request.args.get("tipo") or "cooperado").strip().lower()
    tipo = "restaurante" if tipo_raw == "restaurante" else "cooperado"

    restaurante_id = request.args.get("restaurante_id", type=int)
    cooperado_id = request.args.get("cooperado_id", type=int)

    data_inicio = (request.args.get("data_inicio") or "").strip()
    data_fim = (request.args.get("data_fim") or "").strip()

    filtro_manual = bool(
        restaurante_id
        or cooperado_id
        or data_inicio
        or data_fim
    )

    if not filtro_manual:
        hoje = date.today()
        di = hoje - timedelta(days=hoje.weekday())
        df = di + timedelta(days=6)
        data_inicio = di.strftime("%Y-%m-%d")
        data_fim = df.strftime("%Y-%m-%d")
    else:
        di = _parse_date(data_inicio)
        df = _parse_date(data_fim)

        if di and not df:
            df = di
            data_fim = df.strftime("%Y-%m-%d")
        elif df and not di:
            di = df
            data_inicio = di.strftime("%Y-%m-%d")

    Model = AvaliacaoRestaurante if tipo == "restaurante" else AvaliacaoCooperado

    def col(*names):
        for n in names:
            if hasattr(Model, n):
                return getattr(Model, n)
        return None

    f_geral = col("estrelas_geral")

    if tipo == "restaurante":
        f_trat = col("estrelas_tratamento", "estrelas_pontualidade")
        f_amb = col("estrelas_ambiente", "estrelas_educacao")
        f_sup = col("estrelas_suporte", "estrelas_eficiencia")
    else:
        f_pont = col("estrelas_pontualidade")
        f_educ = col("estrelas_educacao")
        f_efic = col("estrelas_eficiencia")
        f_apres = col("estrelas_apresentacao")

    base = (
        db.session.query(
            Model,
            Restaurante.id.label("rest_id"),
            Restaurante.nome.label("rest_nome"),
            Cooperado.id.label("coop_id"),
            Cooperado.nome.label("coop_nome"),
        )
        .join(Restaurante, Model.restaurante_id == Restaurante.id)
        .join(Cooperado, Model.cooperado_id == Cooperado.id)
    )

    filtros = []

    if restaurante_id:
        filtros.append(Model.restaurante_id == restaurante_id)

    if cooperado_id:
        filtros.append(Model.cooperado_id == cooperado_id)

    if di:
        filtros.append(Model.criado_em >= datetime.combine(di, datetime.min.time()))

    if df:
        filtros.append(Model.criado_em <= datetime.combine(df, datetime.max.time()))

    if filtros:
        base = base.filter(and_(*filtros))

    total = base.with_entities(func.count(Model.id)).scalar() or 0

    page = request.args.get("page", type=int) or 1
    per_page = request.args.get("per_page", type=int) or 10000
    if per_page > 200:
        per_page = 200

    offset = (page - 1) * per_page

    rows = (
        base.order_by(Model.criado_em.desc())
        .limit(per_page)
        .offset(offset)
        .all()
    )

    pages = max(1, (total + per_page - 1) // per_page)

    pager = SimpleNamespace(
        page=page,
        per_page=per_page,
        total=total,
        pages=pages,
        has_prev=(page > 1),
        has_next=(page < pages),
    )

    avaliacoes = []
    for a, rest_id, rest_nome, coop_id, coop_nome in rows:
        item = {
            "criado_em": a.criado_em,
            "rest_id": rest_id,
            "rest_nome": rest_nome,
            "coop_id": coop_id,
            "coop_nome": coop_nome,
            "geral": getattr(a, "estrelas_geral", 0) or 0,
            "comentario": (getattr(a, "comentario", "") or "").strip(),
            "media": getattr(a, "media_ponderada", None),
            "sentimento": getattr(a, "sentimento", None),
            "temas": getattr(a, "temas", None),
            "alerta": bool(getattr(a, "alerta_crise", False)),
            "tratamento": 0,
            "ambiente": 0,
            "suporte": 0,
            "trat": 0,
            "amb": 0,
            "sup": 0,
            "pont": 0,
            "educ": 0,
            "efic": 0,
            "apres": 0,
        }

        if tipo == "restaurante":
            trat = getattr(a, "estrelas_tratamento", None)
            if trat is None:
                trat = getattr(a, "estrelas_pontualidade", 0)

            amb = getattr(a, "estrelas_ambiente", None)
            if amb is None:
                amb = getattr(a, "estrelas_educacao", 0)

            sup = getattr(a, "estrelas_suporte", None)
            if sup is None:
                sup = getattr(a, "estrelas_eficiencia", 0)

            item.update({
                "tratamento": trat or 0,
                "ambiente": amb or 0,
                "suporte": sup or 0,
                "trat": trat or 0,
                "amb": amb or 0,
                "sup": sup or 0,
            })
        else:
            item.update({
                "pont": getattr(a, "estrelas_pontualidade", 0) or 0,
                "educ": getattr(a, "estrelas_educacao", 0) or 0,
                "efic": getattr(a, "estrelas_eficiencia", 0) or 0,
                "apres": getattr(a, "estrelas_apresentacao", 0) or 0,
            })

        avaliacoes.append(SimpleNamespace(**item))

    def avg_or_zero(coluna):
        if coluna is None:
            return 0.0

        q = db.session.query(func.coalesce(func.avg(coluna), 0.0))
        if filtros:
            q = q.select_from(Model).filter(and_(*filtros))
        return float(q.scalar() or 0.0)

    kpis = {
        "qtd": total,
        "geral": avg_or_zero(f_geral),
    }

    if tipo == "restaurante":
        kpis.update({
            "trat": avg_or_zero(f_trat),
            "amb": avg_or_zero(f_amb),
            "sup": avg_or_zero(f_sup),
        })
    else:
        kpis.update({
            "pont": avg_or_zero(f_pont),
            "educ": avg_or_zero(f_educ),
            "efic": avg_or_zero(f_efic),
            "apres": avg_or_zero(f_apres),
        })

    ranking = []
    chart_top = {"labels": [], "values": []}

    if tipo == "restaurante":
        q_rank = (
            db.session.query(
                Restaurante.id.label("id"),
                Restaurante.nome.label("nome"),
                func.count(Model.id).label("qtd"),
                func.coalesce(func.avg(f_geral), 0.0).label("m_geral"),
                (func.coalesce(func.avg(f_trat), 0.0) if f_trat is not None else literal(0.0)).label("m_trat"),
                (func.coalesce(func.avg(f_amb), 0.0) if f_amb is not None else literal(0.0)).label("m_amb"),
                (func.coalesce(func.avg(f_sup), 0.0) if f_sup is not None else literal(0.0)).label("m_sup"),
            )
            .join(Model, Model.restaurante_id == Restaurante.id)
        )

        if filtros:
            q_rank = q_rank.filter(and_(*filtros))

        ranking_rows = q_rank.group_by(Restaurante.id, Restaurante.nome).all()

        ranking = [{
            "rest_nome": r.nome,
            "qtd": int(r.qtd or 0),
            "m_geral": float(r.m_geral or 0),
            "m_trat": float(r.m_trat or 0),
            "m_amb": float(r.m_amb or 0),
            "m_sup": float(r.m_sup or 0),
        } for r in ranking_rows]

        top = sorted(
            [x for x in ranking if x["qtd"] >= 3],
            key=lambda x: x["m_geral"],
            reverse=True,
        )[:10]

        chart_top = {
            "labels": [r["rest_nome"] for r in top],
            "values": [round(r["m_geral"], 2) for r in top],
        }

    else:
        q_rank = (
            db.session.query(
                Cooperado.id.label("id"),
                Cooperado.nome.label("nome"),
                func.count(Model.id).label("qtd"),
                func.coalesce(func.avg(f_geral), 0.0).label("m_geral"),
                (func.coalesce(func.avg(f_pont), 0.0) if f_pont is not None else literal(0.0)).label("m_pont"),
                (func.coalesce(func.avg(f_educ), 0.0) if f_educ is not None else literal(0.0)).label("m_educ"),
                (func.coalesce(func.avg(f_efic), 0.0) if f_efic is not None else literal(0.0)).label("m_efic"),
                (func.coalesce(func.avg(f_apres), 0.0) if f_apres is not None else literal(0.0)).label("m_apres"),
            )
            .join(Model, Model.cooperado_id == Cooperado.id)
        )

        if filtros:
            q_rank = q_rank.filter(and_(*filtros))

        ranking_rows = q_rank.group_by(Cooperado.id, Cooperado.nome).all()

        ranking = [{
            "coop_nome": r.nome,
            "qtd": int(r.qtd or 0),
            "m_geral": float(r.m_geral or 0),
            "m_pont": float(r.m_pont or 0),
            "m_educ": float(r.m_educ or 0),
            "m_efic": float(r.m_efic or 0),
            "m_apres": float(r.m_apres or 0),
        } for r in ranking_rows]

        top = sorted(
            [x for x in ranking if x["qtd"] >= 3],
            key=lambda x: x["m_geral"],
            reverse=True,
        )[:10]

        chart_top = {
            "labels": [r["coop_nome"] for r in top],
            "values": [round(r["m_geral"], 2) for r in top],
        }

    compat_map = {}
    for a in avaliacoes:
        key = (a.coop_id, a.rest_id)
        d = compat_map.get(key)

        if not d:
            d = {
                "coop": a.coop_nome,
                "rest": a.rest_nome,
                "sum": 0.0,
                "count": 0,
            }

        d["sum"] += (a.geral or 0)
        d["count"] += 1
        compat_map[key] = d

    compat = []
    for d in compat_map.values():
        avg = (d["sum"] / d["count"]) if d["count"] else 0.0
        compat.append({
            "coop": d["coop"],
            "rest": d["rest"],
            "avg": avg,
            "count": d["count"],
        })

    compat.sort(
        key=lambda x: (-(x["avg"] or 0), -(x["count"] or 0), x["coop"], x["rest"])
    )

    _flt = SimpleNamespace(
        restaurante_id=restaurante_id,
        cooperado_id=cooperado_id,
        data_inicio=data_inicio or "",
        data_fim=data_fim or "",
    )

    preserve = request.args.to_dict(flat=True)
    preserve.pop("page", None)

    cfg = get_config()

    admin_user = (
        Usuario.query
        .filter_by(tipo="admin", is_master=True)
        .order_by(Usuario.id.asc())
        .first()
    )
    if not admin_user:
        admin_user = (
            Usuario.query
            .filter_by(tipo="admin")
            .order_by(Usuario.id.asc())
            .first()
        )

    admin_logado = _usuario_logado()

    if admin_logado and getattr(admin_logado, "is_master", False):
        admin_perms = {
            aba: {
                "ver": True,
                "criar": True,
                "editar": True,
                "excluir": True,
            }
            for aba in ADMIN_ABAS.keys()
        }
    else:
        admin_perms = get_admin_permissions_map(admin_logado.id) if admin_logado else {}

    admins_secundarios = (
        Usuario.query
        .filter_by(tipo="admin", is_master=False)
        .order_by(Usuario.id.asc())
        .all()
    )

    return render_template(
        "admin_avaliacoes.html",
        tab="avaliacoes",
        tipo=tipo,
        avaliacoes=avaliacoes,
        kpis=kpis,
        ranking=ranking,
        chart_top=chart_top,
        compat=compat,
        _flt=_flt,
        restaurantes=Restaurante.query.order_by(Restaurante.nome).all(),
        cooperados=Cooperado.query.order_by(Cooperado.nome).all(),
        pager=pager,
        page=pager.page,
        per_page=pager.per_page,
        preserve=preserve,
        admin=admin_user,
        salario_minimo=cfg.salario_minimo or 0.0,
        admin_perms=admin_perms,
        admin_is_master=is_admin_master(),
        ADMIN_ABAS=ADMIN_ABAS,
        admins_secundarios=admins_secundarios,
    )


@app.route("/admin/avaliacoes/export")
@admin_perm_required("avaliacoes", "ver")
def admin_export_avaliacoes_csv():
    flash("Exportação em CSV ainda não foi implementada.", "warning")
    args = {
        "data_inicio": request.args.get("data_inicio") or "",
        "data_fim": request.args.get("data_fim") or "",
        "tipo": request.args.get("tipo") or "",
    }
    return redirect(url_for("admin_avaliacoes", **args))


# =========================
# CRUD Receitas/Despesas Coop (Admin)
# =========================
@app.route("/receitas/add", methods=["POST"])
@admin_perm_required("receitas", "criar")
def add_receita():
    f = request.form

    r = ReceitaCooperativa(
        descricao=(f.get("descricao") or "").strip(),
        valor_total=f.get("valor", type=float),
        data=_parse_date(f.get("data")),
    )

    db.session.add(r)
    db.session.commit()
    flash("Receita adicionada.", "success")
    return redirect(url_for("admin_dashboard", tab="receitas"))


@app.route("/receitas/<int:id>/edit", methods=["POST"])
@admin_perm_required("receitas", "editar")
def edit_receita(id):
    r = ReceitaCooperativa.query.get_or_404(id)
    f = request.form

    r.descricao = (f.get("descricao") or "").strip()
    r.valor_total = f.get("valor", type=float)
    r.data = _parse_date(f.get("data"))

    db.session.commit()
    flash("Receita atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="receitas"))


@app.route("/receitas/<int:id>/delete", methods=["GET", "POST"])
@admin_perm_required("receitas", "excluir")
def delete_receita(id):
    r = ReceitaCooperativa.query.get_or_404(id)
    db.session.delete(r)
    db.session.commit()
    flash("Receita excluída.", "success")
    return redirect(url_for("admin_dashboard", tab="receitas"))


@app.route("/receitas/taxas-admin/<int:id>/status", methods=["POST"])
@admin_perm_required("receitas", "editar")
def atualizar_taxa_admin_status(id):
    r = ReceitaCooperativa.query.get_or_404(id)
    if not getattr(r, 'auto_taxa_adm', False):
        flash("Registro não é taxa administrativa automática.", "warning")
        return redirect(url_for("admin_dashboard", tab="receitas"))

    f = request.form
    status = (f.get("status_pagamento") or "nao_pago").strip().lower()
    if status not in ("nao_pago", "parcial", "pago"):
        status = "nao_pago"

    data_pagamento = _parse_date(f.get("data_pagamento"))
    if status == 'pago' and not data_pagamento:
        data_pagamento = date.today()

    valor_previsto = _safe_float(getattr(r, 'valor_previsto', None) or getattr(r, 'valor_principal', None) or 0.0)
    valor_pago = f.get("valor_pago", type=float)
    if valor_pago is None:
        valor_pago = valor_previsto if status == 'pago' else 0.0
    valor_pago = max(0.0, _safe_float(valor_pago))
    if status == 'nao_pago':
        valor_pago = 0.0
        data_pagamento = None

    calc = _calc_taxa_admin_encargos(valor_previsto, getattr(r, 'data_vencimento', None) or getattr(r, 'data', None), data_pagamento if status == 'pago' else None, _safe_float(getattr(r, 'multa_percentual', 2.0), 2.0), _safe_float(getattr(r, 'juros_dia_percentual', 0.03), 0.03))
    r.status_pagamento = status
    r.valor_previsto = valor_previsto
    r.valor_principal = valor_previsto
    r.valor_pago = round(valor_pago, 2)
    r.data_pagamento = data_pagamento
    if status == 'pago':
        r.valor_multa = calc['valor_multa']
        r.valor_juros = calc['valor_juros']
        r.valor_total = round(r.valor_pago + r.valor_multa + r.valor_juros, 2)
        r.data = data_pagamento or r.data_vencimento or r.data
    else:
        r.valor_multa = 0.0
        r.valor_juros = 0.0
        r.valor_total = 0.0
        r.data = getattr(r, 'data_vencimento', None) or r.data

    db.session.commit()
    flash("Taxa administrativa atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="receitas"))


@app.route("/receitas/taxas-admin/lote", methods=["POST"])
@admin_perm_required("receitas", "editar")
def atualizar_taxa_admin_lote():
    ids = request.form.getlist("ids[]") or request.form.getlist("ids")
    acao = (request.form.get("acao") or '').strip().lower()
    ids_int = []
    for v in ids:
        try:
            ids_int.append(int(v))
        except Exception:
            pass
    if not ids_int:
        flash("Selecione pelo menos uma taxa administrativa.", "warning")
        return redirect(url_for("admin_dashboard", tab="receitas"))
    regs = ReceitaCooperativa.query.filter(ReceitaCooperativa.id.in_(ids_int), ReceitaCooperativa.auto_taxa_adm.is_(True)).all()
    hoje = date.today()
    for r in regs:
        valor_previsto = _safe_float(getattr(r, 'valor_previsto', None) or getattr(r, 'valor_principal', None) or 0.0)
        if acao == 'pago':
            calc = _calc_taxa_admin_encargos(valor_previsto, getattr(r, 'data_vencimento', None) or getattr(r, 'data', None), hoje, _safe_float(getattr(r, 'multa_percentual', 2.0), 2.0), _safe_float(getattr(r, 'juros_dia_percentual', 0.03), 0.03))
            r.status_pagamento = 'pago'
            r.data_pagamento = hoje
            r.valor_previsto = valor_previsto
            r.valor_principal = valor_previsto
            r.valor_pago = valor_previsto
            r.valor_multa = calc['valor_multa']
            r.valor_juros = calc['valor_juros']
            r.valor_total = round(valor_previsto + r.valor_multa + r.valor_juros, 2)
            r.data = hoje
        elif acao == 'nao_pago':
            r.status_pagamento = 'nao_pago'
            r.data_pagamento = None
            r.valor_previsto = valor_previsto
            r.valor_principal = valor_previsto
            r.valor_pago = 0.0
            r.valor_multa = 0.0
            r.valor_juros = 0.0
            r.valor_total = 0.0
            r.data = getattr(r, 'data_vencimento', None) or r.data
    db.session.commit()
    flash("Taxas administrativas atualizadas em lote.", "success")
    return redirect(url_for("admin_dashboard", tab="receitas"))


@app.route("/despesas/add", methods=["POST"])
@admin_perm_required("despesas", "criar")
def add_despesa():
    f = request.form

    d = DespesaCooperativa(
        descricao=(f.get("descricao") or "").strip(),
        valor=f.get("valor", type=float),
        data=_parse_date(f.get("data")),
    )

    db.session.add(d)
    db.session.commit()
    flash("Despesa adicionada.", "success")
    return redirect(url_for("admin_dashboard", tab="despesas"))


@app.route("/despesas/<int:id>/edit", methods=["POST"])
@admin_perm_required("despesas", "editar")
def edit_despesa(id):
    d = DespesaCooperativa.query.get_or_404(id)
    f = request.form

    d.descricao = (f.get("descricao") or "").strip()
    d.valor = f.get("valor", type=float)
    d.data = _parse_date(f.get("data"))

    db.session.commit()
    flash("Despesa atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="despesas"))


@app.route("/despesas/<int:id>/delete", methods=["GET", "POST"])
@admin_perm_required("despesas", "excluir")
def delete_despesa(id):
    d = DespesaCooperativa.query.get_or_404(id)
    db.session.delete(d)
    db.session.commit()
    flash("Despesa excluída.", "success")
    return redirect(url_for("admin_dashboard", tab="despesas"))


# =========================
# Avisos (admin + públicos)
# =========================
@app.get("/avisos", endpoint="avisos_publicos")
def avisos_publicos():
    t = session.get("user_tipo")
    if t == "cooperado":
        return redirect(url_for("portal_cooperado_avisos"))
    if t == "restaurante":
        return redirect(url_for("portal_restaurante"))
    return redirect(url_for("login"))


@app.route("/admin/avisos", methods=["GET", "POST"])
@admin_perm_required("avisos", "ver")
def admin_avisos():
    cooperados = Cooperado.query.order_by(Cooperado.nome.asc()).all()
    restaurantes = Restaurante.query.order_by(Restaurante.nome.asc()).all()

    if request.method == "POST":
        if not admin_has_perm("avisos", "criar"):
            flash("Você não tem permissão para publicar avisos.", "danger")
            return redirect(url_for("admin_avisos"))

        f = request.form

        destino_tipo = (f.get("destino_tipo") or "").strip()
        coop_alc = f.get("coop_alcance") or f.get("coop_alcance_ambos")
        rest_alc = f.get("rest_alcance") or f.get("rest_alcance_ambos")
        sel_coops = request.form.getlist("dest_cooperados[]") or request.form.getlist("dest_cooperados_ambos[]")
        sel_rests = request.form.getlist("dest_restaurantes[]") or request.form.getlist("dest_restaurantes_ambos[]")

        def _pick_msg(form):
            for key in (
                "corpo_html", "html", "mensagem_html", "conteudo_html", "descricao_html", "texto_html",
                "mensagem", "corpo", "conteudo", "descricao", "texto", "resumo", "body", "content",
            ):
                v = form.get(key)
                if v and v.strip():
                    return v.strip()
            return ""

        titulo = (f.get("titulo") or "").strip()
        msg = _pick_msg(f)
        prioridade = ((f.get("prioridade") or "normal").strip() or "normal")
        ativo = bool(f.get("ativo"))
        exigir_confirmacao = bool(f.get("exigir_confirmacao")) if hasattr(Aviso, "exigir_confirmacao") else False

        inicio_em = _parse_datetime_local(f.get("inicio_em") or f.get("agendar_inicio"))
        fim_em = _parse_datetime_local(f.get("fim_em") or f.get("agendar_fim"))

        if not titulo:
            flash("Informe o título do aviso.", "warning")
            return redirect(url_for("admin_avisos"))
        if not msg:
            flash("Informe a mensagem do aviso.", "warning")
            return redirect(url_for("admin_avisos"))
        if inicio_em and fim_em and fim_em < inicio_em:
            flash("A data final do agendamento não pode ser menor que a inicial.", "warning")
            return redirect(url_for("admin_avisos"))

        def _mk_aviso(tipo: str):
            a = Aviso(
                titulo=titulo,
                corpo=msg,
                tipo=tipo,
                prioridade=prioridade,
                fixado=False,
                ativo=ativo if hasattr(Aviso, "ativo") else True,
                criado_por_id=session.get("user_id"),
                inicio_em=inicio_em,
                fim_em=fim_em,
            )
            if hasattr(a, "exigir_confirmacao"):
                a.exigir_confirmacao = exigir_confirmacao
            return a

        avisos_para_criar = []

        if destino_tipo == "cooperados":
            if coop_alc == "selecionados":
                if not sel_coops:
                    flash("Selecione ao menos um cooperado.", "warning")
                    return redirect(url_for("admin_avisos"))
                try:
                    coop_ids = sorted({int(x) for x in sel_coops})
                except Exception:
                    flash("Seleção de cooperado inválida.", "warning")
                    return redirect(url_for("admin_avisos"))

                a = _mk_aviso("cooperado")
                a.destino_cooperado_id = None
                a.cooperados = Cooperado.query.filter(Cooperado.id.in_(coop_ids)).all()
                avisos_para_criar.append(a)
            else:
                a = _mk_aviso("cooperado")
                a.destino_cooperado_id = None
                a.cooperados = []
                avisos_para_criar.append(a)

        elif destino_tipo == "restaurantes":
            if rest_alc == "selecionados":
                if not sel_rests:
                    flash("Selecione ao menos um restaurante.", "warning")
                    return redirect(url_for("admin_avisos"))
                try:
                    ids = [int(x) for x in sel_rests]
                except Exception:
                    flash("Seleção de restaurante inválida.", "warning")
                    return redirect(url_for("admin_avisos"))

                a = _mk_aviso("restaurante")
                a.restaurantes = Restaurante.query.filter(Restaurante.id.in_(ids)).all()
                avisos_para_criar.append(a)
            else:
                a = _mk_aviso("restaurante")
                a.restaurantes = []
                avisos_para_criar.append(a)

        elif destino_tipo == "ambos":
            if coop_alc == "selecionados":
                if not sel_coops:
                    flash("Selecione ao menos um cooperado para o aviso dos cooperados.", "warning")
                    return redirect(url_for("admin_avisos"))
                try:
                    coop_ids = sorted({int(x) for x in sel_coops})
                except Exception:
                    flash("Seleção de cooperado inválida.", "warning")
                    return redirect(url_for("admin_avisos"))

                a_coop = _mk_aviso("cooperado")
                a_coop.destino_cooperado_id = None
                a_coop.cooperados = Cooperado.query.filter(Cooperado.id.in_(coop_ids)).all()
                avisos_para_criar.append(a_coop)
            else:
                a_coop = _mk_aviso("cooperado")
                a_coop.destino_cooperado_id = None
                a_coop.cooperados = []
                avisos_para_criar.append(a_coop)

            if rest_alc == "selecionados":
                if not sel_rests:
                    flash("Selecione ao menos um restaurante para o aviso dos restaurantes.", "warning")
                    return redirect(url_for("admin_avisos"))
                try:
                    ids = [int(x) for x in sel_rests]
                except Exception:
                    flash("Seleção de restaurante inválida.", "warning")
                    return redirect(url_for("admin_avisos"))

                a_rest = _mk_aviso("restaurante")
                a_rest.restaurantes = Restaurante.query.filter(Restaurante.id.in_(ids)).all()
            else:
                a_rest = _mk_aviso("restaurante")
                a_rest.restaurantes = []

            avisos_para_criar.append(a_rest)

        else:
            avisos_para_criar.append(_mk_aviso("global"))

        for a in avisos_para_criar:
            db.session.add(a)

        db.session.commit()
        flash("Aviso(s) salvo(s) com sucesso.", "success")
        return redirect(url_for("admin_avisos"))

    avisos = Aviso.query.options(selectinload(Aviso.restaurantes), selectinload(Aviso.cooperados)).order_by(Aviso.fixado.desc(), Aviso.criado_em.desc()).all()

    leituras = AvisoLeitura.query.order_by(AvisoLeitura.lido_em.desc()).all()
    leituras_por_aviso = defaultdict(list)
    for leitura in leituras:
        leituras_por_aviso[leitura.aviso_id].append(leitura)

    now_dt = datetime.utcnow()

    for a in avisos:
        destinatarios_coop, destinatarios_rest = _aviso_destinatarios(a, cooperados_all=cooperados, restaurantes_all=restaurantes)
        registros = leituras_por_aviso.get(a.id, [])

        lidos_coop = {r.cooperado_id: r for r in registros if r.cooperado_id}
        lidos_rest = {r.restaurante_id: r for r in registros if r.restaurante_id}

        if a.tipo == "global":
            a.destino_resumo = "Todos"
        elif a.tipo == "cooperado":
            if getattr(a, "cooperados", None):
                a.destino_resumo = "Cooperados selecionados"
            elif a.destino_cooperado_id:
                a.destino_resumo = "Cooperado específico"
            else:
                a.destino_resumo = "Todos os cooperados"
        elif a.tipo == "restaurante":
            if getattr(a, "restaurantes", None):
                a.destino_resumo = "Restaurantes selecionados"
            else:
                a.destino_resumo = "Todos os restaurantes"
        else:
            a.destino_resumo = a.tipo.capitalize()
        a.agendado = bool(a.inicio_em and a.inicio_em > now_dt)
        a.expirado = bool(a.fim_em and a.fim_em < now_dt)

        a.leituras_cooperados = [
            {
                "id": c.id,
                "nome": c.nome,
                "lido": c.id in lidos_coop,
                "lido_em": lidos_coop.get(c.id).lido_em if c.id in lidos_coop else None,
            }
            for c in destinatarios_coop
        ]
        a.leituras_restaurantes = [
            {
                "id": r.id,
                "nome": r.nome,
                "lido": r.id in lidos_rest,
                "lido_em": lidos_rest.get(r.id).lido_em if r.id in lidos_rest else None,
            }
            for r in destinatarios_rest
        ]
        a.total_destinatarios = len(a.leituras_cooperados) + len(a.leituras_restaurantes)
        a.total_lidos = sum(1 for x in a.leituras_cooperados if x["lido"]) + sum(1 for x in a.leituras_restaurantes if x["lido"])
        a.total_pendentes = max(a.total_destinatarios - a.total_lidos, 0)

    return render_template(
        "admin_avisos.html",
        avisos=avisos,
        cooperados=cooperados,
        cooperados_todos=cooperados_todos,
        restaurantes=restaurantes,
        restaurantes_todos=restaurantes_todos,
        agora=now_dt,
    )


@app.route("/admin/avisos/<int:aviso_id>/toggle", methods=["POST", "GET"], endpoint="admin_avisos_toggle")
@admin_perm_required("avisos", "editar")
def admin_avisos_toggle(aviso_id):
    a = Aviso.query.get_or_404(aviso_id)

    if hasattr(a, "ativo"):
        a.ativo = not bool(a.ativo)
    else:
        a.fixado = not bool(a.fixado)

    db.session.commit()
    flash("Aviso atualizado.", "success")
    return redirect(request.referrer or url_for("admin_avisos"))


@app.route("/admin/avisos/<int:aviso_id>/excluir", methods=["POST"], endpoint="admin_avisos_excluir")
@admin_perm_required("avisos", "excluir")
def admin_avisos_excluir(aviso_id):
    a = Aviso.query.get_or_404(aviso_id)

    try:
        AvisoLeitura.query.filter_by(aviso_id=aviso_id).delete(synchronize_session=False)
    except Exception:
        pass

    try:
        if hasattr(a, "restaurantes"):
            a.restaurantes.clear()
    except Exception:
        pass

    db.session.delete(a)
    db.session.commit()
    flash("Aviso excluído.", "success")
    return redirect(url_for("admin_avisos"))


@app.route("/avisos/<int:aviso_id>/lido", methods=["POST", "GET"])
def marcar_aviso_lido_universal(aviso_id: int):
    if "user_id" not in session:
        return redirect(url_for("login")) if request.method == "GET" else ("", 401)

    user_id = session.get("user_id")
    user_tipo = session.get("user_tipo")
    Aviso.query.get_or_404(aviso_id)

    def _ok_response():
        if request.method == "POST":
            return ("", 204)
        return redirect(request.referrer or url_for("portal_cooperado_avisos"))

    if user_tipo == "cooperado":
        coop = Cooperado.query.filter_by(usuario_id=user_id).first()
        if not coop:
            return ("", 403) if request.method == "POST" else redirect(url_for("login"))

        if not AvisoLeitura.query.filter_by(aviso_id=aviso_id, cooperado_id=coop.id).first():
            db.session.add(
                AvisoLeitura(
                    aviso_id=aviso_id,
                    cooperado_id=coop.id,
                    lido_em=datetime.utcnow(),
                )
            )
            db.session.commit()

        return _ok_response()

    if user_tipo == "restaurante":
        rest = Restaurante.query.filter_by(usuario_id=user_id).first()
        if not rest:
            return ("", 403) if request.method == "POST" else redirect(url_for("login"))

        if not AvisoLeitura.query.filter_by(aviso_id=aviso_id, restaurante_id=rest.id).first():
            db.session.add(
                AvisoLeitura(
                    aviso_id=aviso_id,
                    restaurante_id=rest.id,
                    lido_em=datetime.utcnow(),
                )
            )
            db.session.commit()

        return _ok_response()

    return ("", 403) if request.method == "POST" else redirect(url_for("login"))


# =========================
# CRUD Cooperados / Restaurantes / Senhas (Admin)
# =========================
@app.route("/cooperados/add", methods=["POST"])
@admin_perm_required("cooperados", "criar")
def add_cooperado():
    f = request.form
    nome = (f.get("nome") or "").strip()
    usuario_login = (f.get("usuario") or "").strip()
    senha = f.get("senha") or ""
    telefone = (f.get("telefone") or "").strip()
    foto = request.files.get("foto")
    taxa_admin_valor = f.get("taxa_admin_valor", type=float) or 0.0
    taxa_admin_data_base = _parse_date(f.get("taxa_admin_data_base"))
    taxa_admin_multa_percentual = f.get("taxa_admin_multa_percentual", type=float) or 2.0
    taxa_admin_juros_dia_percentual = f.get("taxa_admin_juros_dia_percentual", type=float) or 0.03

    if Usuario.query.filter_by(usuario=usuario_login).first():
        flash("Usuário já existente.", "warning")
        return redirect(url_for("admin_dashboard", tab="cooperados"))

    u = Usuario(usuario=usuario_login, tipo="cooperado", senha_hash="")
    u.set_password(senha)
    db.session.add(u)
    db.session.flush()

    c = Cooperado(
        nome=nome,
        usuario_id=u.id,
        telefone=telefone,
        ultima_atualizacao=datetime.now(),
    )
    db.session.add(c)
    db.session.flush()

    if foto and foto.filename:
        _save_foto_to_db(c, foto, is_cooperado=True)

    db.session.commit()
    flash("Cooperado cadastrado.", "success")
    return redirect(url_for("admin_dashboard", tab="cooperados"))


@app.route("/cooperados/<int:id>/edit", methods=["POST"])
@admin_perm_required("cooperados", "editar")
def edit_cooperado(id):
    c = Cooperado.query.get_or_404(id)
    f = request.form

    c.nome = (f.get("nome") or "").strip()
    c.usuario_ref.usuario = (f.get("usuario") or "").strip()
    c.telefone = (f.get("telefone") or "").strip()

    foto = request.files.get("foto")
    if foto and foto.filename:
        _save_foto_to_db(c, foto, is_cooperado=True)

    c.ultima_atualizacao = datetime.now()
    db.session.commit()
    flash("Cooperado atualizado.", "success")
    return redirect(url_for("admin_dashboard", tab="cooperados"))


@app.route("/cooperados/<int:id>/delete", methods=["POST"])
@admin_perm_required("cooperados", "excluir")
def delete_cooperado(id):
    c = Cooperado.query.get_or_404(id)
    u = c.usuario_ref

    try:
        escala_ids = [
            eid for (eid,) in db.session.query(Escala.id)
            .filter(Escala.cooperado_id == id)
            .all()
        ]

        if escala_ids:
            db.session.execute(
                sa_delete(TrocaSolicitacao)
                .where(TrocaSolicitacao.origem_escala_id.in_(escala_ids))
            )
            db.session.execute(
                sa_delete(Escala)
                .where(Escala.id.in_(escala_ids))
            )

        db.session.execute(
            sa_delete(TrocaSolicitacao)
            .where(or_(
                TrocaSolicitacao.solicitante_id == id,
                TrocaSolicitacao.destino_id == id,
            ))
        )

        db.session.execute(
            sa_delete(AvaliacaoCooperado).where(AvaliacaoCooperado.cooperado_id == id)
        )
        db.session.execute(
            sa_delete(AvaliacaoRestaurante).where(AvaliacaoRestaurante.cooperado_id == id)
        )
        db.session.execute(
            sa_delete(Lancamento).where(Lancamento.cooperado_id == id)
        )
        db.session.execute(
            sa_delete(ReceitaCooperado).where(ReceitaCooperado.cooperado_id == id)
        )
        db.session.execute(
            sa_delete(DespesaCooperado).where(DespesaCooperado.cooperado_id == id)
        )
        db.session.execute(
            sa_delete(AvisoLeitura).where(AvisoLeitura.cooperado_id == id)
        )

        db.session.delete(c)
        if u:
            db.session.delete(u)

        db.session.commit()
        flash("Cooperado excluído.", "success")

    except IntegrityError as e:
        db.session.rollback()
        current_app.logger.exception(e)
        flash("Não foi possível excluir: existem vínculos ativos.", "danger")

    return redirect(url_for("admin_dashboard", tab="cooperados"))


@app.route("/cooperados/<int:id>/reset_senha", methods=["POST"])
@admin_perm_required("cooperados", "editar")
def reset_senha_cooperado(id):
    c = Cooperado.query.get_or_404(id)
    ns = request.form.get("nova_senha") or ""
    cs = request.form.get("confirmar_senha") or ""

    if ns != cs:
        flash("As senhas não conferem.", "warning")
        return redirect(url_for("admin_dashboard", tab="cooperados"))

    c.usuario_ref.set_password(ns)
    db.session.commit()
    flash("Senha do cooperado atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="cooperados"))


@app.route("/restaurantes/add", methods=["POST"])
@admin_perm_required("restaurantes", "criar")
def add_restaurante():
    f = request.form
    nome = (f.get("nome") or "").strip()
    periodo = f.get("periodo", "seg-dom")
    usuario_login = (f.get("usuario") or "").strip()
    senha = f.get("senha", "")
    foto = request.files.get("foto")

    taxa_admin_valor = f.get("taxa_admin_valor", type=float) or 0.0
    taxa_admin_data_base = _parse_date(f.get("taxa_admin_data_base"))
    taxa_admin_multa_percentual = f.get("taxa_admin_multa_percentual", type=float) or 2.0
    taxa_admin_juros_dia_percentual = f.get("taxa_admin_juros_dia_percentual", type=float) or 0.03
    status_raw = (f.get("ativo") or "1").strip().lower()
    ativo_rest = status_raw in ("1", "true", "ativo", "on", "sim")

    if Usuario.query.filter_by(usuario=usuario_login).first():
        flash("Usuário já existente.", "warning")
        return redirect(url_for("admin_dashboard", tab="restaurantes"))

    u = Usuario(usuario=usuario_login, tipo="restaurante", senha_hash="")
    u.set_password(senha)
    db.session.add(u)
    db.session.flush()

    r = Restaurante(
        nome=nome,
        periodo=periodo,
        usuario_id=u.id,
        taxa_admin_valor=taxa_admin_valor,
        taxa_admin_data_base=taxa_admin_data_base,
        taxa_admin_multa_percentual=taxa_admin_multa_percentual,
        taxa_admin_juros_dia_percentual=taxa_admin_juros_dia_percentual,
        ativo=ativo_rest,
    )
    db.session.add(r)
    db.session.flush()

    if foto and foto.filename:
        _save_foto_to_db(r, foto, is_cooperado=False)

    db.session.commit()
    _ensure_taxas_admin_receitas([r], months_back=0)
    flash("Estabelecimento cadastrado.", "success")
    return redirect(url_for("admin_dashboard", tab="restaurantes"))


@app.route("/restaurantes/<int:id>/edit", methods=["POST"])
@admin_perm_required("restaurantes", "editar")
def edit_restaurante(id):
    r = Restaurante.query.get_or_404(id)
    f = request.form

    r.nome = (f.get("nome") or "").strip()
    r.periodo = f.get("periodo", "seg-dom")
    r.usuario_ref.usuario = (f.get("usuario") or "").strip()
    r.taxa_admin_valor = f.get("taxa_admin_valor", type=float) or 0.0
    r.taxa_admin_data_base = _parse_date(f.get("taxa_admin_data_base"))
    r.taxa_admin_multa_percentual = f.get("taxa_admin_multa_percentual", type=float) or 2.0
    r.taxa_admin_juros_dia_percentual = f.get("taxa_admin_juros_dia_percentual", type=float) or 0.03
    status_raw = (f.get('ativo') or '1').strip().lower()
    ativo_rest = status_raw in ('1','true','ativo','on','sim')
    if hasattr(r, 'ativo'):
        r.ativo = ativo_rest
    if getattr(r, 'usuario_ref', None) is not None and hasattr(r.usuario_ref, 'ativo'):
        r.usuario_ref.ativo = ativo_rest

    foto = request.files.get("foto")
    if foto and foto.filename:
        _save_foto_to_db(r, foto, is_cooperado=False)

    db.session.commit()
    _ensure_taxas_admin_receitas([r], months_back=0)
    flash("Estabelecimento atualizado.", "success")
    return redirect(url_for("admin_dashboard", tab="restaurantes"))


@app.route("/restaurantes/<int:id>/delete", methods=["POST"])
@admin_perm_required("restaurantes", "excluir")
def delete_restaurante(id):
    r = Restaurante.query.get_or_404(id)
    u = r.usuario_ref

    try:
        escala_ids = [
            e.id for e in Escala.query.with_entities(Escala.id)
            .filter(Escala.restaurante_id == id)
            .all()
        ]

        if escala_ids:
            db.session.execute(
                sa_delete(TrocaSolicitacao)
                .where(TrocaSolicitacao.origem_escala_id.in_(escala_ids))
            )
            db.session.execute(
                sa_delete(Escala)
                .where(Escala.restaurante_id == id)
            )

        db.session.execute(
            sa_delete(Lancamento).where(Lancamento.restaurante_id == id)
        )

        db.session.delete(r)
        if u:
            db.session.delete(u)

        db.session.commit()
        flash("Estabelecimento excluído.", "success")

    except IntegrityError as e:
        db.session.rollback()
        current_app.logger.exception(e)
        flash("Não foi possível excluir: existem vínculos ativos.", "danger")

    return redirect(url_for("admin_dashboard", tab="restaurantes"))


@app.route("/restaurantes/<int:id>/reset_senha", methods=["POST"])
@admin_perm_required("restaurantes", "editar")
def reset_senha_restaurante(id):
    r = Restaurante.query.get_or_404(id)
    ns = request.form.get("nova_senha") or ""
    cs = request.form.get("confirmar_senha") or ""

    if ns != cs:
        flash("As senhas não conferem.", "warning")
        return redirect(url_for("admin_dashboard", tab="restaurantes"))

    r.usuario_ref.set_password(ns)
    db.session.commit()
    flash("Senha do restaurante atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="restaurantes"))


@app.route("/rest/alterar-senha", methods=["POST"], endpoint="rest_alterar_senha")
@role_required("restaurante")
def alterar_senha_rest():
    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first_or_404()
    user = rest.usuario_ref

    atual = (request.form.get("senha_atual") or "").strip()
    nova = (request.form.get("senha_nova") or "").strip()
    conf = (request.form.get("senha_conf") or "").strip()

    if not (nova and conf):
        flash("Preencha todos os campos.", "warning")
        return redirect(url_for("portal_restaurante", view="config"))

    if nova != conf:
        flash("A confirmação não confere com a nova senha.", "warning")
        return redirect(url_for("portal_restaurante", view="config"))

    if len(nova) < 6:
        flash("A nova senha deve ter pelo menos 6 caracteres.", "warning")
        return redirect(url_for("portal_restaurante", view="config"))

    if user.senha_hash and not atual:
        flash("Informe a senha atual.", "warning")
        return redirect(url_for("portal_restaurante", view="config"))

    if user.senha_hash and not check_password_hash(user.senha_hash, atual):
        flash("Senha atual incorreta.", "danger")
        return redirect(url_for("portal_restaurante", view="config"))

    user.senha_hash = generate_password_hash(nova)
    db.session.commit()
    flash("Senha alterada com sucesso!", "success")
    return redirect(url_for("portal_restaurante", view="config"))



# =========================
# Backup / Restauração XLSX
# =========================
def _xlsx_cell(v):
    if isinstance(v, (datetime, date)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    if isinstance(v, bool):
        return 1 if v else 0
    return "" if v is None else v


def _sheet_from_rows(wb: Workbook, title: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append([_xlsx_cell(v) for v in row])
    for idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)
    return ws


def _coerce_bool(v, default=False):
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("1", "true", "sim", "yes", "y", "on")


def _coerce_int(v, default=None):
    if v in (None, ""):
        return default
    try:
        return int(v)
    except Exception:
        try:
            return int(float(v))
        except Exception:
            return default


def _coerce_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return _parse_date(str(v))


def _sync_pk_sequence(model):
    try:
        table = model.__tablename__
        pk = sa_inspect(model).primary_key[0].name
        max_id = db.session.execute(sa_text(f'SELECT COALESCE(MAX({pk}), 0) FROM {table}')).scalar() or 0
        if _is_sqlite():
            try:
                db.session.execute(sa_text("UPDATE sqlite_sequence SET seq = :seq WHERE name = :name"), {"seq": max_id, "name": table})
                db.session.commit()
            except Exception:
                db.session.rollback()
        else:
            db.session.execute(sa_text("SELECT setval(pg_get_serial_sequence(:table, :pk), :value, true)"), {"table": table, "pk": pk, "value": max_id if max_id > 0 else 1})
            db.session.commit()
    except Exception:
        db.session.rollback()


def _excel_safe_sheet_name(name: str) -> str:
    safe = re.sub(r'[:\/?*\[\]]', '_', str(name or '').strip())[:31].strip()
    return safe or 'Planilha'


def _serialize_backup_value(value):
    if value is None:
        return None
    if isinstance(value, bytes):
        import base64
        return '__bytes_base64__:' + base64.b64encode(value).decode('ascii')
    if isinstance(value, datetime):
        return value.isoformat(sep=' ', timespec='seconds')
    if isinstance(value, date):
        return value.isoformat()
    return value


def _deserialize_backup_value(raw, column):
    if raw in (None, ''):
        return None

    try:
        pytype = getattr(column.type, 'python_type', None)
    except Exception:
        pytype = None

    if pytype is bytes and isinstance(raw, str) and raw.startswith('__bytes_base64__:'):
        import base64
        try:
            return base64.b64decode(raw.split(':', 1)[1].encode('ascii'))
        except Exception:
            return None

    if pytype is bool:
        return _coerce_bool(raw, False)
    if pytype is int:
        return _coerce_int(raw)
    if pytype is float:
        try:
            return float(raw)
        except Exception:
            return None
    if pytype is date:
        return _coerce_date(raw)
    if pytype is datetime:
        if isinstance(raw, datetime):
            return raw
        if isinstance(raw, date):
            return datetime.combine(raw, dtime.min)
        try:
            return datetime.fromisoformat(str(raw).strip().replace('T', ' '))
        except Exception:
            return None
    if pytype is bytes and isinstance(raw, (bytes, bytearray)):
        return bytes(raw)
    return raw


def _sync_table_pk_sequence(table):
    try:
        pk_cols = list(table.primary_key.columns)
        if len(pk_cols) != 1:
            return
        pk = pk_cols[0]
        try:
            pytype = pk.type.python_type
        except Exception:
            pytype = None
        if pytype is not int:
            return

        table_name = table.name
        pk_name = pk.name
        max_id = db.session.execute(sa_text(f'SELECT COALESCE(MAX({pk_name}), 0) FROM {table_name}')).scalar() or 0
        if _is_sqlite():
            try:
                db.session.execute(sa_text("UPDATE sqlite_sequence SET seq = :seq WHERE name = :name"), {"seq": max_id, "name": table_name})
                db.session.commit()
            except Exception:
                db.session.rollback()
        else:
            db.session.execute(
                sa_text("SELECT setval(pg_get_serial_sequence(:table, :pk), :value, true)"),
                {"table": table_name, "pk": pk_name, "value": max_id if max_id > 0 else 1},
            )
            db.session.commit()
    except Exception:
        db.session.rollback()


def _backup_tables_in_order():
    return list(db.metadata.sorted_tables)


def _backup_workbook_bytes() -> io.BytesIO:
    wb = Workbook()
    ws0 = wb.active
    wb.remove(ws0)

    tables = _backup_tables_in_order()
    for table in tables:
        headers = [c.name for c in table.columns]
        rows_db = db.session.execute(table.select().order_by(*list(table.primary_key.columns))).mappings().all()
        rows = [[_serialize_backup_value(row.get(col)) for col in headers] for row in rows_db]
        _sheet_from_rows(wb, _excel_safe_sheet_name(table.name), headers, rows)

    meta = wb.create_sheet(title='instrucoes')
    meta['A1'] = 'Backup completo COOPEX'
    meta['A2'] = 'Este arquivo exporta todas as tabelas do banco em abas separadas.'
    meta['A3'] = 'A importação restaura os dados presentes nas abas reconhecidas, sem apagar dados ao exportar.'
    meta['A4'] = 'Não altere os nomes das abas nem os cabeçalhos das colunas.'
    meta['A5'] = 'Total de abas de dados: {}'.format(len(tables))
    meta.column_dimensions['A'].width = 120

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    out = []
    for row in rows[1:]:
        if row is None:
            continue
        item = {}
        has_data = False
        for i, header in enumerate(headers):
            if not header:
                continue
            val = row[i] if i < len(row) else None
            if val not in (None, ''):
                has_data = True
            item[header] = val
        if has_data:
            out.append(item)
    return out


def _import_backup_workbook(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    tables = _backup_tables_in_order()
    table_map = {table.name: table for table in tables}
    sheet_to_table = { _excel_safe_sheet_name(name): name for name in table_map.keys() }

    available = []
    for sheet_name in wb.sheetnames:
        table_name = sheet_to_table.get(sheet_name)
        if table_name:
            available.append(table_name)

    if not available:
        raise ValueError('Nenhuma aba de tabela reconhecida foi encontrada no arquivo.')

    try:
        for table in reversed(tables):
            if table.name in available:
                db.session.execute(table.delete())
        db.session.flush()

        for table in tables:
            if table.name not in available:
                continue
            ws = wb[_excel_safe_sheet_name(table.name)]
            rows = _sheet_rows(ws)
            if not rows:
                continue
            payload = []
            valid_cols = {c.name: c for c in table.columns}
            for row in rows:
                item = {}
                for key, raw in row.items():
                    col = valid_cols.get(key)
                    if not col:
                        continue
                    item[key] = _deserialize_backup_value(raw, col)
                if item:
                    payload.append(item)
            if payload:
                db.session.execute(table.insert(), payload)

        db.session.flush()
        for table in tables:
            if table.name in available:
                _sync_table_pk_sequence(table)
    except Exception:
        db.session.rollback()
        raise

@app.route("/admin/backup/exportar", methods=["GET"])
@admin_perm_required("config", "ver")
def exportar_backup_admin_xlsx():
    if not is_admin_master():
        flash("Apenas o administrador master pode exportar o backup completo.", "danger")
        return redirect(url_for("admin_dashboard", tab="config"))

    arquivo = _backup_workbook_bytes()
    nome = f"backup_coopex_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=nome,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/admin/backup/importar", methods=["POST"])
@admin_perm_required("config", "editar")
def importar_backup_admin_xlsx():
    if not is_admin_master():
        flash("Apenas o administrador master pode importar backup completo.", "danger")
        return redirect(url_for("admin_dashboard", tab="config"))

    arquivo = request.files.get("arquivo_backup")
    if not arquivo or not getattr(arquivo, "filename", ""):
        flash("Selecione um arquivo XLSX para importar.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    if not arquivo.filename.lower().endswith(".xlsx"):
        flash("Envie um arquivo no formato .xlsx.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    try:
        _import_backup_workbook(arquivo)
        db.session.commit()
        flash("Backup importado com sucesso.", "success")
    except Exception as e:
        db.session.rollback()
        try:
            current_app.logger.exception(e)
        except Exception:
            pass
        flash(f"Falha ao importar backup: {e}", "danger")

    return redirect(url_for("admin_dashboard", tab="config"))

# =========================
# Configurações / Admins
# =========================
@app.route("/config/update", methods=["POST"])
@admin_perm_required("config", "editar")
def update_config():
    cfg = get_config()
    cfg.salario_minimo = request.form.get("salario_minimo", type=float) or 0.0
    db.session.commit()
    flash("Configuração atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="config"))


@app.route("/admin/alterar_admin", methods=["POST"])
@admin_perm_required("config", "editar")
def alterar_admin():
    admin = Usuario.query.filter_by(tipo="admin", is_master=True).first()

    if not admin:
        flash("Administrador master não encontrado.", "danger")
        return redirect(url_for("admin_dashboard", tab="config"))

    novo_usuario = (request.form.get("usuario") or "").strip()
    nova = (request.form.get("nova_senha") or "").strip()
    confirmar = (request.form.get("confirmar_senha") or "").strip()

    if not novo_usuario:
        flash("Informe o usuário do administrador.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    existente = Usuario.query.filter(
        Usuario.usuario == novo_usuario,
        Usuario.id != admin.id
    ).first()
    if existente:
        flash("Já existe outro usuário com esse login.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    admin.usuario = novo_usuario

    if nova or confirmar:
        if nova != confirmar:
            flash("As senhas não conferem.", "warning")
            return redirect(url_for("admin_dashboard", tab="config"))
        admin.set_password(nova)

    db.session.commit()
    flash("Conta do administrador atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="config"))


@app.route("/admin/admins/add", methods=["POST"])
@admin_perm_required("config", "editar")
def add_admin_secundario():
    if not is_admin_master():
        flash("Apenas o administrador master pode criar outros administradores.", "danger")
        return redirect(url_for("admin_dashboard", tab="config"))

    nome = (request.form.get("nome") or "").strip()
    usuario = (request.form.get("usuario") or "").strip()
    senha = (request.form.get("senha") or "").strip()
    confirmar_senha = (request.form.get("confirmar_senha") or "").strip()
    ativo = str(request.form.get("ativo") or "1").strip() == "1"

    if not nome or not usuario or not senha or not confirmar_senha:
        flash("Preencha nome, usuário, senha e confirmação.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    if senha != confirmar_senha:
        flash("As senhas não conferem.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    if Usuario.query.filter_by(usuario=usuario).first():
        flash("Já existe um usuário com esse login.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    u = Usuario(
        nome=nome,
        usuario=usuario,
        tipo="admin",
        senha_hash="",
        is_master=False,
        ativo=ativo,
    )
    u.set_password(senha)

    db.session.add(u)
    db.session.flush()

    for aba in ADMIN_ABAS.keys():
        db.session.add(
            AdminPermissao(
                usuario_id=u.id,
                aba=aba,
                pode_ver=False,
                pode_criar=False,
                pode_editar=False,
                pode_excluir=False,
            )
        )

    db.session.commit()
    flash("Administrador criado com sucesso.", "success")
    return redirect(url_for("admin_dashboard", tab="config"))


@app.route("/admin/admins/<int:usuario_id>/reset-password", methods=["POST"])
@admin_perm_required("config", "editar")
def admin_reset_admin_password(usuario_id):
    if not is_admin_master():
        flash("Apenas o administrador master pode redefinir senhas de administradores.", "danger")
        return redirect(url_for("admin_dashboard", tab="config"))

    admin = Usuario.query.filter_by(id=usuario_id, tipo="admin").first_or_404()

    if admin.is_master:
        flash("A senha do administrador master não pode ser redefinida por esta tela.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    nova_senha = (request.form.get("nova_senha") or "").strip()
    confirmar_senha = (request.form.get("confirmar_senha") or "").strip()

    if not nova_senha or not confirmar_senha:
        flash("Preencha a nova senha e a confirmação.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    if nova_senha != confirmar_senha:
        flash("As senhas não conferem.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    admin.set_password(nova_senha)
    db.session.commit()

    flash("Senha do administrador redefinida com sucesso.", "success")
    return redirect(url_for("admin_dashboard", tab="config"))


@app.route(
    "/admin/admins/<int:usuario_id>/permissoes",
    methods=["POST"],
    endpoint="admin_salvar_permissoes"
)
@admin_perm_required("config", "editar")
def salvar_permissoes_admin(usuario_id):
    if not is_admin_master():
        flash("Apenas o administrador master pode alterar permissões.", "danger")
        return redirect(url_for("admin_dashboard", tab="config"))

    admin = Usuario.query.filter_by(id=usuario_id, tipo="admin").first_or_404()

    if admin.is_master:
        flash("As permissões do administrador master não podem ser limitadas por esta tela.", "warning")
        return redirect(url_for("admin_dashboard", tab="config"))

    for aba in ADMIN_ABAS.keys():
        perm = AdminPermissao.query.filter_by(usuario_id=admin.id, aba=aba).first()

        if not perm:
            perm = AdminPermissao(
                usuario_id=admin.id,
                aba=aba,
                pode_ver=False,
                pode_criar=False,
                pode_editar=False,
                pode_excluir=False,
            )
            db.session.add(perm)

        perm.pode_ver = bool(request.form.get(f"perm_{aba}_ver"))
        perm.pode_criar = bool(request.form.get(f"perm_{aba}_criar"))
        perm.pode_editar = bool(request.form.get(f"perm_{aba}_editar"))
        perm.pode_excluir = bool(request.form.get(f"perm_{aba}_excluir"))

    db.session.commit()
    flash("Permissões atualizadas com sucesso.", "success")
    return redirect(url_for("admin_dashboard", tab="config"))


# =========================
# Receitas/Despesas Cooperado (Admin)
# =========================
@app.route("/coop/receitas/add", methods=["POST"])
@admin_perm_required("coop_receitas", "criar")
def add_receita_coop():
    f = request.form

    coop = _cooperado_ativo_por_id(f.get("cooperado_id", type=int))
    if not coop:
        flash("Selecione um cooperado ativo.", "warning")
        return redirect(url_for("admin_dashboard", tab="coop_receitas"))

    rc = ReceitaCooperado(
        cooperado_id=coop.id,
        descricao=(f.get("descricao") or "").strip(),
        valor=f.get("valor", type=float),
        data=_parse_date(f.get("data")),
    )

    db.session.add(rc)
    db.session.commit()
    flash("Receita do cooperado adicionada.", "success")
    return redirect(url_for("admin_dashboard", tab="coop_receitas"))


@app.route("/coop/receitas/<int:id>/edit", methods=["POST"])
@admin_perm_required("coop_receitas", "editar")
def edit_receita_coop(id):
    rc = ReceitaCooperado.query.get_or_404(id)
    f = request.form

    coop = _cooperado_ativo_por_id(f.get("cooperado_id", type=int))
    if not coop:
        flash("Selecione um cooperado ativo.", "warning")
        return redirect(url_for("admin_dashboard", tab="coop_receitas"))

    rc.cooperado_id = coop.id
    rc.descricao = (f.get("descricao") or "").strip()
    rc.valor = f.get("valor", type=float)
    rc.data = _parse_date(f.get("data"))

    db.session.commit()
    flash("Receita do cooperado atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="coop_receitas"))


@app.route("/coop/receitas/<int:id>/delete", methods=["GET", "POST"])
@admin_perm_required("coop_receitas", "excluir")
def delete_receita_coop(id):
    rc = ReceitaCooperado.query.get_or_404(id)
    db.session.delete(rc)
    db.session.commit()
    flash("Receita do cooperado excluída.", "success")
    return redirect(url_for("admin_dashboard", tab="coop_receitas"))


def _competencia_ref(data_base, competencia_semana):
    base = data_base or date.today()
    comp = (competencia_semana or '').strip().lower()
    if comp in ('passada', 'semana_passada'):
        base = base - timedelta(days=7)
    elif comp in ('proxima', 'proxima_semana'):
        base = base + timedelta(days=7)
    return base

def _competencia_label(comp):
    comp = (comp or '').strip().lower()
    if comp in ('passada', 'semana_passada'):
        return 'semana_passada'
    if comp in ('proxima', 'proxima_semana'):
        return 'proxima_semana'
    return 'esta_semana'

def _despesa_due_date(dc):
    base = dc.data_fim or dc.data or date.today()
    comp = _competencia_label(getattr(dc, 'competencia_desconto', 'esta_semana'))
    if comp == 'semana_passada':
        return base - timedelta(days=7)
    if comp == 'proxima_semana':
        return base + timedelta(days=7)
    return base



def _admin_redirect_with_filters(default_tab="coop_despesas"):
    args = {
        "tab": request.form.get("tab") or request.args.get("tab") or default_tab,
        "data_inicio": request.form.get("data_inicio") or request.args.get("data_inicio") or "",
        "data_fim": request.form.get("data_fim") or request.args.get("data_fim") or "",
        "restaurante_id": request.form.get("restaurante_id") or request.args.get("restaurante_id") or "",
        "cooperado_id": request.form.get("cooperado_id") or request.args.get("cooperado_id") or "",
    }
    if request.form.get("somente_pendentes") or request.args.get("somente_pendentes"):
        args["somente_pendentes"] = "1"
    return redirect(url_for("admin_dashboard", **args))


def _cooperado_ativo_por_id(cid: int | None) -> Cooperado | None:
    try:
        cid = int(cid or 0)
    except Exception:
        return None
    coop = Cooperado.query.get(cid)
    if not coop or not _is_cooperado_ativo(coop):
        return None
    return coop

def _cooperados_ativos_por_ids(ids) -> list[Cooperado]:
    vistos = set()
    coops = []
    for raw in ids or []:
        try:
            cid = int(raw)
        except Exception:
            continue
        if cid in vistos:
            continue
        coop = _cooperado_ativo_por_id(cid)
        if coop is None:
            continue
        vistos.add(cid)
        coops.append(coop)
    coops.sort(key=lambda c: (c.nome or '').lower())
    return coops

def _restaurante_ativo_por_id(rid: int | None) -> Restaurante | None:
    try:
        rid = int(rid or 0)
    except Exception:
        return None
    rest = Restaurante.query.get(rid)
    if not rest or not _is_restaurante_ativo(rest):
        return None
    return rest

def _compute_coop_debt_snapshot(coop_id, di, df):
    """
    Regras do abatimento automático no resumo do cooperado:
    - Produção líquida da semana = produção bruta - INSS - SEST.
    - Receita entra inteira como crédito da semana.
    - O crédito da semana abate primeiro ADIANTAMENTOS vencidos/abertos.
    - Depois abate DESPESAS vencidas/abertas.
    - Crédito positivo de semanas anteriores NÃO quita dívida de semanas futuras.
    - Dívida pendente carrega para as semanas seguintes até zerar.
    """
    as_of = df or date.today()

    def _week_start(d):
        return d - timedelta(days=d.weekday())

    q_prod = (
        Lancamento.query
        .filter(Lancamento.cooperado_id == coop_id)
        .filter(Lancamento.data <= as_of)
        .order_by(Lancamento.data.asc(), Lancamento.id.asc())
    )
    q_rec = (
        ReceitaCooperado.query
        .filter(ReceitaCooperado.cooperado_id == coop_id)
        .filter(ReceitaCooperado.data <= as_of)
        .order_by(ReceitaCooperado.data.asc(), ReceitaCooperado.id.asc())
    )
    q_desp = (
        DespesaCooperado.query
        .filter(DespesaCooperado.cooperado_id == coop_id)
        .order_by(DespesaCooperado.data_fim.asc().nullslast(), DespesaCooperado.id.asc())
    )

    prods = q_prod.all()
    recs = q_rec.all()
    despesas = q_desp.all()

    itens_map = {}
    creditos_por_semana = {}
    semanas = set()

    for p in prods:
        if not p.data:
            continue
        ws = _week_start(p.data)
        semanas.add(ws)
        liquido = round(float(p.valor or 0.0) * (1.0 - INSS_ALIQ - SEST_ALIQ), 2)
        if liquido > 0:
            creditos_por_semana[ws] = round(creditos_por_semana.get(ws, 0.0) + liquido, 2)

    for r in recs:
        if not r.data:
            continue
        ws = _week_start(r.data)
        semanas.add(ws)
        valor = round(float(r.valor or 0.0), 2)
        if valor > 0:
            creditos_por_semana[ws] = round(creditos_por_semana.get(ws, 0.0) + valor, 2)

    for dc in despesas:
        valor_total = round(float(dc.valor or 0.0), 2)
        if valor_total <= 0:
            continue
        due_date = _despesa_due_date(dc)
        due_week = _week_start(due_date)
        semanas.add(due_week)
        itens_map[dc.id] = {
            'id': dc.id,
            'data': dc.data,
            'data_inicio': dc.data_inicio,
            'data_fim': dc.data_fim,
            'due_date': due_date,
            'due_week': due_week,
            'descricao': dc.descricao or '',
            'valor': valor_total,
            'eh_adiantamento': bool(getattr(dc, 'eh_adiantamento', False)),
            'beneficio_id': getattr(dc, 'beneficio_id', None),
            'competencia_desconto': getattr(dc, 'competencia_desconto', None),
            'pago_auto': 0.0,
            'pago_manual': 0.0,
            'restante': valor_total,
            'status': 'a_descontar' if due_date > as_of else 'aberta',
        }

    semanas_processadas = sorted(ws for ws in semanas if ws <= _week_start(as_of))
    if not semanas_processadas and itens_map:
        semanas_processadas = sorted({_week_start(as_of)})

    abertas = []
    livre_sem_divida = 0.0

    def _ordenar_abertas():
        abertas.sort(key=lambda item: (0 if item['eh_adiantamento'] else 1, item['due_week'], item['due_date'], item['id']))

    for semana in semanas_processadas:
        for item in itens_map.values():
            if item['due_week'] == semana and item['due_date'] <= as_of and item['restante'] > 0:
                if item not in abertas:
                    item['status'] = 'aberta'
                    abertas.append(item)
        _ordenar_abertas()

        credito_semana = round(creditos_por_semana.get(semana, 0.0), 2)
        restante_credito = credito_semana

        while restante_credito > 0.0001 and abertas:
            item = abertas[0]
            falta = round(item['restante'], 2)
            abat = min(falta, restante_credito)
            item['pago_auto'] = round(item['pago_auto'] + abat, 2)
            item['restante'] = round(max(0.0, item['restante'] - abat), 2)
            restante_credito = round(max(0.0, restante_credito - abat), 2)
            if item['restante'] <= 0.0001:
                item['restante'] = 0.0
                item['status'] = 'quitada'
                abertas.pop(0)
            else:
                item['status'] = 'parcial'
                _ordenar_abertas()

        livre_sem_divida = round(restante_credito if not abertas else 0.0, 2)

    total_vencido_pendente = 0.0
    total_programado = 0.0
    total_descontado_despesa = 0.0
    total_descontado_adiant = 0.0
    itens = []

    for item in sorted(itens_map.values(), key=lambda x: (x['due_week'], x['due_date'], 0 if x['eh_adiantamento'] else 1, x['id'])):
        due_date = item['due_date']
        restante = round(item['restante'], 2)
        pago_auto = round(item['pago_auto'], 2)
        valor_total = round(item['valor'], 2)

        if due_date > as_of:
            item['status'] = 'a_descontar'
            total_programado += restante
        else:
            if restante <= 0.0001:
                item['status'] = 'quitada'
            elif pago_auto > 0:
                item['status'] = 'parcial'
            else:
                item['status'] = 'aberta'
            total_vencido_pendente += restante
            if item['eh_adiantamento']:
                total_descontado_adiant += round(valor_total - restante, 2)
            else:
                total_descontado_despesa += round(valor_total - restante, 2)

        itens.append(item)

    if di or df:
        lower = di if di else None
        upper = df if df else None
        itens_visiveis = []
        for item in itens:
            due_date = item['due_date']
            if lower and due_date < lower and item['restante'] <= 0:
                continue
            if upper and due_date > upper and item['restante'] <= 0:
                continue
            itens_visiveis.append(item)
    else:
        itens_visiveis = itens

    return {
        'itens': itens_visiveis,
        'saldo_devedor': round(total_vencido_pendente, 2),
        'a_descontar': round(total_programado, 2),
        'descontado_despesa': round(total_descontado_despesa, 2),
        'descontado_adiant': round(total_descontado_adiant, 2),
        'disponivel_auto_restante': round(max(0.0, livre_sem_divida), 2),
    }



@app.route("/coop/despesas/add", methods=["POST"])
@admin_perm_required("coop_despesas", "criar")
def add_despesa_coop():
    f = request.form
    ids = f.getlist("cooperado_ids[]")
    excluir_ids = f.getlist("cooperado_excluir_ids[]")
    ratear_todos = bool(f.get("ratear_todos_ativos")) or ('all' in ids)

    descricao = (f.get("descricao") or "").strip()
    valor_total = f.get("valor", type=float) or 0.0
    d = _parse_date(f.get("data"))
    eh_adiantamento = bool(f.get("eh_adiantamento"))
    competencia_semana = (f.get("competencia_desconto") or f.get("competencia_semana") or "esta_semana").strip().lower()

    if not d:
        d = date.today()

    if ratear_todos:
        coops_base = _cooperados_ativos_ordenados()
        excluidos = {c.id for c in _cooperados_ativos_por_ids(excluir_ids)}
        coops = [c for c in coops_base if c.id not in excluidos]
    else:
        coops = _cooperados_ativos_por_ids(ids)

    if not coops:
        flash("Selecione pelo menos um cooperado ativo.", "warning")
        return _admin_redirect_with_filters("coop_despesas")

    data_comp = _competencia_ref(d, competencia_semana)
    di_comp, df_comp = semana_bounds(data_comp)

    qtd = len(coops)
    valor_unit = round((valor_total / qtd), 2) if qtd > 0 else 0.0
    acumulado = 0.0

    for idx, coop in enumerate(coops, start=1):
        valor_item = valor_unit
        if idx == qtd:
            valor_item = round(valor_total - acumulado, 2)
        acumulado = round(acumulado + valor_item, 2)
        db.session.add(
            DespesaCooperado(
                cooperado_id=coop.id,
                descricao=descricao,
                valor=valor_item,
                data=df_comp,
                data_inicio=di_comp,
                data_fim=df_comp,
                eh_adiantamento=eh_adiantamento,
                competencia_desconto=competencia_semana,
            )
        )

    db.session.commit()
    flash("Despesa(s) lançada(s).", "success")
    return _admin_redirect_with_filters("coop_despesas")


@app.route("/coop/despesas/<int:id>/edit", methods=["POST"])
@admin_perm_required("coop_despesas", "editar")
def edit_despesa_coop(id):
    dc = DespesaCooperado.query.get_or_404(id)
    f = request.form

    coop = _cooperado_ativo_por_id(f.get("cooperado_id", type=int))
    if not coop:
        flash("Selecione um cooperado ativo.", "warning")
        return _admin_redirect_with_filters("coop_despesas")

    dc.cooperado_id = coop.id
    dc.descricao = (f.get("descricao") or "").strip()
    dc.valor = f.get("valor", type=float)
    data_edit = _parse_date(f.get("data")) or dc.data or date.today()
    competencia_semana = (f.get("competencia_desconto") or f.get("competencia_semana") or "esta_semana").strip().lower()
    data_comp = _competencia_ref(data_edit, competencia_semana)
    di_comp, df_comp = semana_bounds(data_comp)
    dc.data = df_comp
    dc.data_inicio = di_comp
    dc.data_fim = df_comp
    dc.eh_adiantamento = bool(f.get("eh_adiantamento"))
    dc.competencia_desconto = competencia_semana

    db.session.commit()
    flash("Despesa do cooperado atualizada.", "success")
    return _admin_redirect_with_filters("coop_despesas")


@app.route("/coop/despesas/<int:id>/delete", methods=["GET", "POST"])
@admin_perm_required("coop_despesas", "excluir")
def delete_despesa_coop(id):
    dc = DespesaCooperado.query.get_or_404(id)
    db.session.delete(dc)
    db.session.commit()
    flash("Despesa do cooperado excluída.", "success")
    return _admin_redirect_with_filters("coop_despesas")

# =========================
# Benefícios — Editar / Excluir (Admin)
# =========================

from datetime import date

def _split_field(form, key_list, key_str):
    """
    Lê uma lista do form (key[]) ou uma string separada por , ; (key_str).
    Retorna lista de strings já stripadas e sem vazios.
    """
    vals = form.getlist(key_list) if key_list.endswith("[]") else []
    if not vals:
        raw = (form.get(key_str) or "").replace(",", ";")
        vals = [x.strip() for x in raw.split(";") if x.strip()]
    return vals

def _ensure_periodo_ok(di: date | None, df: date | None) -> tuple[date | None, date | None]:
    if di and df and df < di:
        di, df = df, di  # inverte para garantir di <= df
    return di, df

TIPO_MAP = {
    "hosp": "hospitalar", "hospitalar": "hospitalar",
    "farm": "farmaceutico","farmacêutico":"farmaceutico","farmaceutico":"farmaceutico",
    "alim": "alimentar",  "alimentar": "alimentar",
}

@app.post("/beneficios/<int:id>/edit")
@admin_perm_required("beneficios", "editar")
def edit_beneficio(id):
    """
    Atualiza um registro de benefício existente.
    Espera (via form):
      - data_inicial (YYYY-MM-DD ou DD/MM/YYYY)
      - data_final   (YYYY-MM-DD ou DD/MM/YYYY)
      - data_lancamento (opcional)
      - tipo  (hospitalar|farmaceutico|alimentar ou siglas: hosp|farm|alim)
      - valor_total (float)
      - recebedores_ids[]  ou recebedores_ids  (string: "1;2;3")
      - recebedores_nomes[] ou recebedores_nomes (string: "Ana;Bia;…")
    """
    b = BeneficioRegistro.query.get_or_404(id)
    f = request.form

    # --- Datas ---
    di = _parse_date(f.get("data_inicial"))
    df = _parse_date(f.get("data_final"))
    di, df = _ensure_periodo_ok(di, df)

    if di: b.data_inicial = di
    if df: b.data_final   = df

    dl = _parse_date(f.get("data_lancamento"))
    if dl:
        b.data_lancamento = dl

    # --- Tipo ---
    tipo_in = (f.get("tipo") or "").strip().lower()
    if tipo_in in TIPO_MAP:
        b.tipo = TIPO_MAP[tipo_in]

    # --- Valor total ---
    val_raw = f.get("valor_total")
    if val_raw is not None and val_raw != "":
        try:
            b.valor_total = float(str(val_raw).replace(",", "."))
        except ValueError:
            flash("Valor total inválido.", "warning")
            return redirect(url_for("admin_dashboard", tab="beneficios"))

    # --- Recebedores ---
    ids_list   = _split_field(f, "recebedores_ids[]",   "recebedores_ids")
    nomes_list = _split_field(f, "recebedores_nomes[]", "recebedores_nomes")

    # Se vierem só IDs, tenta resolver nomes
    if ids_list and not nomes_list:
        ids_int = [int(x) for x in ids_list if str(x).isdigit()]
        if ids_int:
            coops = _cooperados_ativos_por_ids(ids_int)
            m = {str(c.id): c.nome for c in coops}
            nomes_list = [m.get(str(i), "") for i in ids_int]

    # Sanitiza; mantém alinhamento por índice
    ids_sane   = [str(int(x)) for x in ids_list if str(x).isdigit()]
    nomes_sane = [n for n in nomes_list if n is not None]

    # Alinha tamanhos (corta o excedente do maior)
    n = min(len(ids_sane), len(nomes_sane)) if ids_sane and nomes_sane else max(len(ids_sane), len(nomes_sane))
    ids_sane   = ids_sane[:n]
    nomes_sane = (nomes_sane[:n] if nomes_sane else [""] * n)

    b.recebedores_ids   = ";".join(ids_sane)
    b.recebedores_nomes = ";".join(nomes_sane)

    db.session.commit()
    flash("Benefício atualizado.", "success")
    return redirect(url_for("admin_dashboard", tab="beneficios"))


# 1) Excluir 1 (via modal, com hidden)
@app.post("/beneficios/delete-one", endpoint="excluir_beneficio_one")
@admin_perm_required("beneficios", "excluir")
def excluir_beneficio_one():
    bid = request.form.get("beneficio_id", type=int)
    if not bid:
        flash("ID inválido.", "warning")
        return redirect(url_for("admin_dashboard", tab="beneficios"))
    b = BeneficioRegistro.query.get_or_404(bid)
    db.session.delete(b)
    db.session.commit()
    flash("Registro de benefício excluído.", "info")
    return redirect(url_for("admin_dashboard", tab="beneficios"))

# 2) Excluir vários (bulk)
@app.post("/beneficios/delete-bulk", endpoint="excluir_beneficio_bulk")
@admin_perm_required("beneficios", "excluir")
def excluir_beneficio_bulk():
    ids = {int(x) for x in request.form.getlist("ids[]") if str(x).isdigit()}
    if not ids:
        flash("Selecione ao menos um benefício.", "warning")
        return redirect(url_for("admin_dashboard", tab="beneficios"))
    qs = BeneficioRegistro.query.filter(BeneficioRegistro.id.in_(ids)).all()
    for b in qs:
        db.session.delete(b)
    db.session.commit()
    flash(f"{len(qs)} registro(s) excluído(s).", "info")
    return redirect(url_for("admin_dashboard", tab="beneficios"))

# =========================
# Benefícios — Criar/Ratear (Admin)
# =========================
@app.post("/beneficios/ratear", endpoint="ratear_beneficios")
@admin_perm_required("beneficios", "criar")
def ratear_beneficios():
    """
    Aceita tanto o formato simples legado quanto o formulário atual da aba Benefícios
    (hospitalar / farmacêutico / alimentar), criando os registros recebidos e lançando
    as despesas para os pagantes do rateio.
    """
    f = request.form

    def _to_float_br(v):
        s = str(v or "").strip()
        if not s:
            return 0.0
        s = s.replace("R$", "").replace(" ", "")
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    def _redirect_back():
        return redirect(url_for(
            "admin_dashboard",
            tab="beneficios",
            data_inicio=(f.get("data_inicio") or ""),
            data_fim=(f.get("data_fim") or ""),
            restaurante_id=(f.get("restaurante_id") or ""),
            cooperado_id=(f.get("cooperado_id") or ""),
        ))

    di = _parse_date(f.get("data_inicial"))
    df = _parse_date(f.get("data_final"))
    if di and df and df < di:
        di, df = df, di

    if not di or not df:
        flash("Preencha a data inicial e a data final.", "warning")
        return _redirect_back()

    dl = _parse_date(f.get("data_lancamento")) or date.today()

    def _coops_by_ids(ids):
        return _cooperados_ativos_por_ids(ids)

    def _add_beneficio(tipo: str, valor_total: float, recebedores_ids: list[int], isentos_ids: list[int]):
        recebedores = _coops_by_ids(recebedores_ids)
        if valor_total <= 0 or not recebedores:
            return False

        rec_ids = [c.id for c in recebedores]
        rec_nomes = [c.nome for c in recebedores]

        b = BeneficioRegistro(
            data_inicial=di,
            data_final=df,
            data_lancamento=dl,
            tipo=tipo,
            valor_total=round(valor_total, 2),
            recebedores_ids=";".join(str(i) for i in rec_ids),
            recebedores_nomes=";".join(rec_nomes),
        )
        db.session.add(b)
        db.session.flush()

        bloqueados = set(rec_ids) | {int(x) for x in isentos_ids if str(x).isdigit()}
        pagantes = [c for c in _cooperados_ativos_ordenados() if c.id not in bloqueados]
        valor_por_pagante = round(valor_total / len(pagantes), 2) if pagantes else 0.0

        if valor_por_pagante > 0:
            for coop in pagantes:
                db.session.add(DespesaCooperado(
                    cooperado_id=coop.id,
                    descricao=f"Rateio benefício {tipo}",
                    valor=valor_por_pagante,
                    data=dl,
                    data_inicio=di,
                    data_fim=df,
                    beneficio_id=b.id,
                    eh_adiantamento=False,
                ))
        return True

    # compatibilidade com fluxo simples legado
    tipo_in = (f.get("tipo") or "").strip().lower()
    ids_list = _split_field(f, "recebedores_ids[]", "recebedores_ids")
    nomes_list = _split_field(f, "recebedores_nomes[]", "recebedores_nomes")
    if tipo_in or ids_list or nomes_list or f.get("valor_total"):
        ids_sane = [int(x) for x in ids_list if str(x).isdigit()]
        if not ids_sane:
            flash("Selecione pelo menos um recebedor.", "warning")
            return _redirect_back()
        tipo = TIPO_MAP.get(tipo_in, tipo_in or "alimentar")
        valor_total = _to_float_br(f.get("valor_total"))
        _add_beneficio(tipo, valor_total, ids_sane, [])
        db.session.commit()
        flash("Benefício registrado com sucesso.", "success")
        return _redirect_back()

    tipos_cfg = [
        ("hosp", "hospitalar"),
        ("farm", "farmaceutico"),
        ("alim", "alimentar"),
    ]
    criou = 0
    for prefixo, tipo in tipos_cfg:
        recebedores_ids = [int(x) for x in f.getlist(f"{prefixo}_beneficiarios[]") if str(x).isdigit()]
        isentos_ids = [int(x) for x in f.getlist(f"{prefixo}_isencoes[]") if str(x).isdigit()]
        valor_total = _to_float_br(f.get(f"{prefixo}_valor"))
        if valor_total <= 0:
            valor_unit = _to_float_br(f.get(f"{prefixo}_valor_unit"))
            valor_total = round(valor_unit * len(recebedores_ids), 2)
        if _add_beneficio(tipo, valor_total, recebedores_ids, isentos_ids):
            criou += 1

    if not criou:
        flash("Informe ao menos um benefício com valor e recebedores.", "warning")
        return _redirect_back()

    db.session.commit()
    flash("Benefícios registrados com sucesso.", "success")
    return _redirect_back()

# =========================
# Escalas/Trocas — Exportações e histórico
# =========================
def _xlsx_finish_and_send(wb, filename, *, fast=False):
    import io
    from openpyxl.styles import Font, PatternFill, Alignment

    if not fast:
        for ws in wb.worksheets:
            try:
                header = next(ws.iter_rows(min_row=1, max_row=1))
            except Exception:
                header = []
            for cell in header:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(fill_type='solid', fgColor='DCE6F1')

            max_scan_rows = 400
            try:
                max_col = ws.max_column or 0
                max_row = ws.max_row or 0
            except Exception:
                max_col = 0
                max_row = 0
            if max_col:
                widths = [10] * max_col
                scan_to = min(max_row, max_scan_rows)
                for row in ws.iter_rows(min_row=1, max_row=scan_to):
                    for idx, cell in enumerate(row):
                        try:
                            widths[idx] = min(max(widths[idx], len(str(cell.value or '')) + 2), 36)
                        except Exception:
                            pass
                from openpyxl.utils import get_column_letter
                for idx, width in enumerate(widths, start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.get("/admin/escalas/exportar_atual")
@admin_perm_required("escalas", "ver")
def admin_exportar_escalas_atual():
    from openpyxl import Workbook

    hist_ini = _parse_ymd_date(request.args.get("escala_hist_inicio")) or (date.today() - timedelta(days=30))
    hist_fim = _parse_ymd_date(request.args.get("escala_hist_fim")) or date.today()

    latest = {}
    hist_q = _history_rows_between(
        db.session.query(
            EscalaHistorico.data,
            EscalaHistorico.turno,
            EscalaHistorico.horario,
            EscalaHistorico.contrato,
            EscalaHistorico.cooperado_nome,
            EscalaHistorico.saiu_nome,
            EscalaHistorico.entrou_nome,
            EscalaHistorico.snapshot_em,
            EscalaHistorico.id,
        ).filter(EscalaHistorico.saiu_nome.isnot(None)),
        EscalaHistorico.snapshot_em,
        hist_ini,
        hist_fim,
    ).order_by(EscalaHistorico.snapshot_em.desc(), EscalaHistorico.id.desc())
    for h in hist_q.all():
        k = ((h.data or '').strip(), (h.turno or '').strip(), (h.horario or '').strip(), (h.contrato or '').strip(), (h.cooperado_nome or '').strip())
        if k not in latest:
            latest[k] = h

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title='Escala atual')
    ws.append(['Data', 'Turno', 'Horário', 'Nº', 'Contrato', 'Cooperado atual', 'Quem foi retirado', 'Quem entrou'])

    escalas_all = db.session.query(
        Escala.data,
        Escala.turno,
        Escala.horario,
        Escala.contrato,
        Escala.cooperado_nome,
        Escala.cooperado_id,
    ).order_by(Escala.data.asc(), Escala.turno.asc(), Escala.contrato.asc(), Escala.horario.asc(), Escala.id.asc()).all()

    coop_ids = sorted({int(e.cooperado_id) for e in escalas_all if getattr(e, 'cooperado_id', None)})
    cooperados_map = {}
    if coop_ids:
        cooperados_map = {c.id: c.nome for c in Cooperado.query.filter(Cooperado.id.in_(coop_ids)).all()}

    slot_counts = defaultdict(int)
    for e in escalas_all:
        nome = cooperados_map.get(getattr(e, 'cooperado_id', None)) or (e.cooperado_nome or '')
        group_key = ((e.data or '').strip(), (e.turno or '').strip(), (e.contrato or '').strip())
        slot_counts[group_key] += 1
        h = latest.get(((e.data or '').strip(), (e.turno or '').strip(), (e.horario or '').strip(), (e.contrato or '').strip(), nome.strip()))
        ws.append([e.data or '', e.turno or '', e.horario or '', slot_counts[group_key], e.contrato or '', nome, (h.saiu_nome if h else ''), (h.entrou_nome if h else '')])

    return _xlsx_finish_and_send(wb, 'escala_atual_com_alteracoes.xlsx', fast=True)


@app.get("/admin/escalas/exportar_historico")
@admin_perm_required("escalas", "ver")
def admin_exportar_escalas_historico():
    from openpyxl import Workbook
    ini = _parse_ymd_date(request.args.get("escala_hist_inicio")) or (date.today() - timedelta(days=30))
    fim = _parse_ymd_date(request.args.get("escala_hist_fim")) or date.today()
    hist = _history_rows_between(
        db.session.query(
            EscalaHistorico.snapshot_em,
            EscalaHistorico.origem,
            EscalaHistorico.acao,
            EscalaHistorico.data,
            EscalaHistorico.turno,
            EscalaHistorico.horario,
            EscalaHistorico.contrato,
            EscalaHistorico.cooperado_nome,
            EscalaHistorico.saiu_nome,
            EscalaHistorico.entrou_nome,
        ),
        EscalaHistorico.snapshot_em,
        ini,
        fim,
    ).order_by(EscalaHistorico.snapshot_em.desc(), EscalaHistorico.id.desc())
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title='Histórico escalas')
    ws.append(['Registrado em', 'Origem', 'Ação', 'Data', 'Turno', 'Horário', 'Contrato', 'Cooperado atual', 'Quem saiu', 'Quem entrou'])
    for h in hist.all():
        ws.append([h.snapshot_em.strftime('%d/%m/%Y %H:%M') if h.snapshot_em else '', h.origem or '', h.acao or '', h.data or '', h.turno or '', h.horario or '', h.contrato or '', h.cooperado_nome or '', h.saiu_nome or '', h.entrou_nome or ''])
    return _xlsx_finish_and_send(wb, 'historico_escalas.xlsx', fast=True)


@app.get("/admin/trocas/exportar_historico")
@admin_perm_required("escalas", "ver")
def admin_exportar_trocas_historico():
    from openpyxl import Workbook
    ini = _parse_ymd_date(request.args.get("trocas_hist_inicio")) or (date.today() - timedelta(days=30))
    fim = _parse_ymd_date(request.args.get("trocas_hist_fim")) or date.today()
    hist = _history_rows_between(
        db.session.query(
            TrocaHistorico.aplicada_em,
            TrocaHistorico.tipo,
            TrocaHistorico.solicitante_nome,
            TrocaHistorico.destino_nome,
            TrocaHistorico.data,
            TrocaHistorico.turno,
            TrocaHistorico.horario,
            TrocaHistorico.contrato,
            TrocaHistorico.saiu_nome,
            TrocaHistorico.entrou_nome,
        ),
        TrocaHistorico.aplicada_em,
        ini,
        fim,
    ).order_by(TrocaHistorico.aplicada_em.desc(), TrocaHistorico.id.desc())
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title='Histórico trocas')
    ws.append(['Aplicada em', 'Tipo', 'Solicitante', 'Destino', 'Data', 'Turno', 'Horário', 'Contrato', 'Saiu', 'Entrou'])
    for h in hist.all():
        ws.append([h.aplicada_em.strftime('%d/%m/%Y %H:%M') if h.aplicada_em else '', h.tipo or '', h.solicitante_nome or '', h.destino_nome or '', h.data or '', h.turno or '', h.horario or '', h.contrato or '', h.saiu_nome or '', h.entrou_nome or ''])
    return _xlsx_finish_and_send(wb, 'historico_trocas.xlsx', fast=True)


# =========================
# Escalas — Upload (substituição TOTAL sempre)
# =========================
@app.route("/escalas/upload", methods=["POST"])
@admin_perm_required("escalas", "criar")
def upload_escala():
    from datetime import datetime, date
    import os, re as _re, unicodedata as _u, difflib as _dif

    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("Envie um arquivo .xlsx válido.", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    # salva o arquivo (o nome não influencia a lógica)
    path = os.path.join(UPLOAD_DIR, secure_filename(file.filename))
    file.save(path)

    # abre com openpyxl
    try:
        import openpyxl
    except Exception:
        flash("Arquivo salvo, mas falta a biblioteca 'openpyxl' (pip install openpyxl).", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"Erro ao abrir a planilha: {e}", "danger")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    # ------- helpers -------
    def _norm_local(s: str) -> str:
        s = _u.normalize("NFD", str(s or "").strip().lower())
        s = "".join(ch for ch in s if _u.category(ch) != "Mn")
        return _re.sub(r"[^a-z0-9]+", " ", s).strip()

    def _norm_login_local(s: str) -> str:
        s = _u.normalize("NFD", s or "")
        s = "".join(ch for ch in s if _u.category(ch) != "Mn")
        return _re.sub(r"\s+", "", s.lower().strip())

    def to_css_color_local(v: str) -> str:
        t = str(v or "").strip()
        if not t: return ""
        if _re.fullmatch(r"[0-9a-fA-F]{8}", t):
            a = int(t[0:2], 16) / 255.0
            r = int(t[2:4], 16); g = int(t[4:6], 16); b = int(t[6:8], 16)
            return f"rgba({r},{g},{b},{a:.3f})"
        if _re.fullmatch(r"[0-9a-fA-F]{6}", t):
            return f"#{t}"
        if _re.fullmatch(r"#?[0-9a-fA-F]{6,8}", t):
            if not t.startswith("#"): t = f"#{t}"
            if len(t) == 9:
                a = int(t[1:3], 16) / 255.0
                r = int(t[3:5], 16); g = int(t[5:7], 16); b = int(t[7:9], 16)
                return f"rgba({r},{g},{b},{a:.3f})"
            return t
        m = _re.fullmatch(r"\s*(\d{1,3})\s*[,;]\s*(\d{1,3})\s*[,;]\s*(\d{1,3})\s*", t)
        if m:
            r, g, b = [max(0, min(255, int(x))) for x in m.groups()]
            return f"rgb({r},{g},{b})"
        mapa = {"azul":"blue","vermelho":"red","verde":"green","amarelo":"yellow",
                "cinza":"gray","preto":"black","branco":"white","laranja":"orange","roxo":"purple"}
        return mapa.get(t.lower(), t)

    def fmt_data_cell(v) -> str:
        if v is None or str(v).strip() == "": return ""
        if isinstance(v, datetime): return v.date().strftime("%d/%m/%Y")
        if isinstance(v, date):     return v.strftime("%d/%m/%Y")
        s = str(v).strip()
        m = _re.fullmatch(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
        if m:
            y, mth, d = map(int, m.groups())
            try: return date(y, mth, d).strftime("%d/%m/%Y")
            except Exception: return s
        return s

    # ------- cabeçalhos (detecção automática) -------
    def _score_header_row(cells):
        aliases = [
            "data","dia","data do plantao","turno","horario","horário","hora","periodo","período",
            "contrato","restaurante","unidade","local",
            "login","usuario","usuário","username","user","nome de usuario","nome de usuário",
            "nome","nome do cooperado","cooperado","motoboy","entregador",
            "cor","cores","cor da celula","cor celula",
        ]
        aliases_norm = {_norm_local(a) for a in aliases}
        score = 0
        seen = set()
        for c in cells:
            key = _norm_local(str(getattr(c, "value", "") or ""))
            if not key:
                continue
            score += 1
            for a in aliases_norm:
                if a and (a == key or a in key or key in a):
                    if (key, a) not in seen:
                        score += 2
                        seen.add((key, a))
        return score

    header_row_idx = 1
    best_score = -1
    last_row_to_check = min(ws.max_row, 10)
    for i in range(1, last_row_to_check + 1):
        row_cells = list(ws[i])
        s = _score_header_row(row_cells)
        if s > best_score:
            best_score = s
            header_row_idx = i

    headers_norm = { _norm_local(str(c.value or "")) : j for j, c in enumerate(ws[header_row_idx], start=1) }

    def find_col(*aliases):
        al = [_norm_local(a) for a in aliases]
        for a in al:
            if a in headers_norm: return headers_norm[a]
        for k_norm, j in headers_norm.items():
            for a in al:
                if a and a in k_norm: return j
        return None

    col_data     = find_col("data", "dia", "data do plantao")
    col_turno    = find_col("turno")
    col_horario  = find_col("horario", "horário", "hora", "periodo", "período")
    col_contrato = find_col("contrato", "restaurante", "unidade", "local")
    col_login    = find_col("login", "usuario", "usuário", "username", "user", "nome de usuario", "nome de usuário")
    col_nome     = find_col("nome", "nome do cooperado", "cooperado", "motoboy", "entregador")
    col_cor      = find_col("cor","cores","cor da celula","cor celula")

    app.logger.info(f"[ESCALAS] header_row={header_row_idx} headers_norm={headers_norm}")

    if not col_login and not col_nome:
        flash("Não encontrei a coluna de LOGIN nem a de NOME do cooperado na planilha.", "danger")
        app.logger.warning(f"[ESCALAS] Falha header: headers_norm={headers_norm} (linha {header_row_idx})")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    # ------- cache entidades -------
    restaurantes = Restaurante.query.order_by(Restaurante.nome).all()
    cooperados   = Cooperado.query.order_by(Cooperado.nome).all()

    def match_restaurante_id(contrato_txt: str) -> int | None:
        a = _norm_local(contrato_txt)
        if not a: return None
        for r in restaurantes:
            b = _norm_local(r.nome)
            if a == b or a in b or b in a: return r.id
        try:
            nomes_norm = [_norm_local(r.nome) for r in restaurantes]
            close = _dif.get_close_matches(a, nomes_norm, n=1, cutoff=0.87)
            if close:
                alvo = close[0]
                for r in restaurantes:
                    if _norm_local(r.nome) == alvo: return r.id
        except Exception:
            pass
        return None

    def match_cooperado_by_login(login_txt: str) -> Cooperado | None:
        key = _norm_login_local(login_txt)
        if not key: return None
        for c in cooperados:
            login = getattr(c, "usuario_ref", None)
            login_val = getattr(login, "usuario", "") if login else ""
            if _norm_login_local(login_val) == key:
                return c
        return None

    # helper global por nome (já existe no seu arquivo)
    from_here_match_by_name = _match_cooperado_by_name

    # ------- parse linhas -------
    linhas_novas, total_linhas_planilha = [], 0
    start_row = header_row_idx + 1

    for i in range(start_row, ws.max_row + 1):
        login_txt = str(ws.cell(i, col_login).value).strip() if col_login else ""
        nome_txt  = str(ws.cell(i, col_nome ).value).strip() if col_nome  else ""

        # ignora linhas totalmente vazias
        if not login_txt and not nome_txt:
            vals = [
                (ws.cell(i, col_data).value     if col_data     else None),
                (ws.cell(i, col_turno).value    if col_turno    else None),
                (ws.cell(i, col_horario).value  if col_horario  else None),
                (ws.cell(i, col_contrato).value if col_contrato else None),
            ]
            if all((v is None or str(v).strip() == "") for v in vals):
                continue

        total_linhas_planilha += 1

        data_v     = ws.cell(i, col_data).value     if col_data     else None
        turno_v    = ws.cell(i, col_turno).value    if col_turno    else None
        horario_v  = ws.cell(i, col_horario).value  if col_horario  else None
        contrato_v = ws.cell(i, col_contrato).value if col_contrato else None
        cor_v      = ws.cell(i, col_cor).value      if col_cor      else None

        contrato_txt = (str(contrato_v).strip() if contrato_v is not None else "")
        rest_id      = match_restaurante_id(contrato_txt)

        coop_match = match_cooperado_by_login(login_txt) if login_txt else None
        if not coop_match and nome_txt:
            coop_match = from_here_match_by_name(nome_txt, cooperados)

        nome_fallback = (nome_txt or login_txt)

        linhas_novas.append({
            "cooperado_id":   (coop_match.id if coop_match else None),
            "cooperado_nome": (None if coop_match else nome_fallback),
            "data":           fmt_data_cell(data_v),
            "turno":          (str(turno_v).strip() if turno_v is not None else ""),
            "horario":        (str(horario_v).strip() if horario_v is not None else ""),
            "contrato":       contrato_txt,
            "cor":            to_css_color_local(cor_v),
            "restaurante_id": rest_id,
        })

    if not linhas_novas:
        app.logger.warning(f"[ESCALAS] Nenhuma linha importada. header_row={header_row_idx} headers_norm={headers_norm}")
        flash("Nada importado: nenhum registro válido encontrado. Verifique a linha dos cabeçalhos e os nomes das colunas.", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    # ------- SUBSTITUIÇÃO TOTAL -------
    try:
        from sqlalchemy import text as sa_text, delete as sa_delete

        # Remove todas as escalas antigas de forma segura (sem violar FKs de trocas)
        if _is_sqlite():
            # Em dev/local com SQLite não há TRUNCATE CASCADE — apaga dependentes e depois escalas
            db.session.execute(sa_delete(TrocaSolicitacao))
            db.session.execute(sa_delete(Escala))
        else:
            # Em produção (Postgres): TRUNCATE com CASCADE limpa escalas e dependentes de uma vez
            db.session.execute(sa_text("TRUNCATE TABLE escalas RESTART IDENTITY CASCADE"))

        # insere novas linhas
        for row in linhas_novas:
            db.session.add(Escala(**row))

        # marca atualização para cooperados reconhecidos
        ids_reconhecidos = {int(r["cooperado_id"]) for r in linhas_novas if r.get("cooperado_id")}
        for cid in ids_reconhecidos:
            c = Cooperado.query.get(cid)
            if c:
                c.ultima_atualizacao = datetime.now()

        db.session.commit()

        try:
            _prune_histories()
            _snapshot_escalas_atual(
                grupo_ref=str(uuid.uuid4()),
                origem="upload",
                acao="snapshot",
                admin_usuario_id=session.get("user_id"),
                snapshot_em=datetime.utcnow(),
            )
            db.session.commit()
        except Exception:
            db.session.rollback()

        flash(f"Escala substituída com sucesso. {len(linhas_novas)} linha(s) importada(s) (de {total_linhas_planilha}).", "success")

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Erro ao importar a escala")
        flash(f"Erro ao importar a escala: {e}", "danger")

    return redirect(url_for("admin_dashboard", tab="escalas"))


# =========================
# Edição rápida da escala semanal
# =========================
@app.post("/admin/escalas/<int:escala_id>/salvar")
@admin_perm_required("escalas", "editar")
def admin_escala_salvar(escala_id):
    e = Escala.query.get_or_404(escala_id)

    contrato_txt = (request.form.get("contrato") or "").strip()
    cooperado_id_raw = (request.form.get("cooperado_id") or "").strip()
    cooperado_nome_livre = (request.form.get("cooperado_nome_livre") or "").strip()
    redirect_day = (request.form.get("redirect_day") or "semana").strip().lower()
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest" or (request.form.get("ajax") == "1")

    def _reply_ok(message):
        if is_ajax:
            return jsonify({"ok": True, "message": message})
        flash(message, "success")
        return redirect(url_for("admin_dashboard", tab="escalas", escala_dia=redirect_day))

    def _reply_error(message, code=400):
        if is_ajax:
            return jsonify({"ok": False, "message": message}), code
        flash(message, "danger")
        return redirect(url_for("admin_dashboard", tab="escalas", escala_dia=redirect_day))

    try:
        before_nome = _safe_coop_nome_by_id(e.cooperado_id) or (e.cooperado_nome or "")

        # contrato pode ser alterado linha por linha; vazio mantém o valor atual
        if contrato_txt:
            e.contrato = contrato_txt
            e.restaurante_id = _match_restaurante_id(contrato_txt)

        if cooperado_id_raw:
            try:
                coop_id = int(cooperado_id_raw)
            except Exception:
                return _reply_error("Cooperado inválido.")

            coop = (
                Cooperado.query
                .join(Usuario, Cooperado.usuario_id == Usuario.id)
                .filter(Cooperado.id == coop_id, Usuario.ativo.is_(True))
                .first()
            )
            if not coop:
                return _reply_error("Cooperado inválido ou inativo.")

            e.cooperado_id = coop.id
            e.cooperado_nome = None
        else:
            # retirar cooperado da linha
            e.cooperado_id = None
            e.cooperado_nome = cooperado_nome_livre or None

        after_nome = ""
        if e.cooperado_id:
            after_nome = _safe_coop_nome_by_id(e.cooperado_id)
        if not after_nome:
            after_nome = e.cooperado_nome or ""

        # primeiro persiste a alteração principal da escala
        db.session.commit()

        # histórico é melhor-esforço; se falhar, não desfaz a alteração da linha
        try:
            _prune_histories()
            _log_escala_historico(
                origem="edicao_manual",
                acao="substituicao",
                escala_ref_id=e.id,
                grupo_ref=str(uuid.uuid4()),
                admin_usuario_id=session.get("user_id"),
                data=e.data or "",
                turno=e.turno or "",
                horario=e.horario or "",
                contrato=e.contrato or "",
                cooperado_id=e.cooperado_id,
                cooperado_nome=after_nome,
                saiu_nome=(before_nome or "") if (before_nome or "") != (after_nome or "") else None,
                entrou_nome=(after_nome or "") if (before_nome or "") != (after_nome or "") else None,
                snapshot_em=datetime.utcnow(),
            )
            db.session.commit()
        except Exception as hist_err:
            app.logger.warning("Falha ao gravar histórico de escala id=%s: %s", e.id, hist_err)
            db.session.rollback()

        return _reply_ok("Linha da escala atualizada com sucesso.")
    except Exception as err:
        db.session.rollback()
        app.logger.exception("Falha ao salvar linha da escala id=%s", escala_id)
        return _reply_error(f"Erro ao salvar a linha da escala: {err}", 500)


@app.get("/admin/api/escalas/alertas_1h")
@admin_perm_required("escalas", "ver")
def admin_api_escala_alertas_1h():
    admin_logado = _usuario_logado()
    if not admin_logado or (admin_logado.tipo or "").strip().lower() != "admin":
        return jsonify({"ok": False, "message": "Não autorizado."}), 403

    escalas_all = (
        db.session.query(Escala)
        .outerjoin(Cooperado, Escala.cooperado_id == Cooperado.id)
        .outerjoin(Usuario, Cooperado.usuario_id == Usuario.id)
        .filter(
            or_(
                Escala.cooperado_id.is_(None),
                Usuario.ativo.is_(True)
            )
        )
        .order_by(Escala.id.asc())
        .all()
    )

    cooperados = (
        Cooperado.query
        .join(Usuario, Cooperado.usuario_id == Usuario.id)
        .filter(Usuario.ativo.is_(True))
        .order_by(Cooperado.nome)
        .all()
    )
    cooperados_map = {c.id: c for c in cooperados}
    alertas = _build_escala_alertas_1h(escalas_all, cooperados_map)
    return jsonify({
        "ok": True,
        "now": _brasil_now().strftime("%Y-%m-%dT%H:%M:%S"),
        "total": len(alertas),
        "alertas": alertas,
    })


# =========================
# Ações de exclusão de escalas
# =========================
@app.post("/escalas/purge_all")
@admin_perm_required("escalas", "excluir")
def escalas_purge_all():
    res = db.session.execute(sa_delete(Escala))
    db.session.commit()
    flash(f"Todas as escalas foram excluídas ({res.rowcount or 0}).", "info")
    return redirect(url_for("admin_dashboard", tab="escalas"))

@app.post("/escalas/purge_cooperado/<int:coop_id>")
@admin_perm_required("escalas", "excluir")
def escalas_purge_cooperado(coop_id):
    res = db.session.execute(sa_delete(Escala).where(Escala.cooperado_id == coop_id))
    db.session.commit()
    flash(f"Escalas do cooperado removidas ({res.rowcount or 0}).", "info")
    return redirect(url_for("admin_dashboard", tab="escalas"))

@app.post("/escalas/purge_restaurante/<int:rest_id>")
@admin_perm_required("escalas", "excluir")
def escalas_purge_restaurante(rest_id):
    res = db.session.execute(sa_delete(Escala).where(Escala.restaurante_id == rest_id))
    db.session.commit()
    flash(f"Escalas do restaurante #{rest_id} excluídas ({res.rowcount or 0}).", "info")
    return redirect(url_for("admin_dashboard", tab="escalas"))

# =========================
# Trocas (Admin aprovar/recusar)
# =========================
@app.post("/admin/trocas/<int:id>/aprovar")
@admin_perm_required("escalas", "editar")
def admin_aprovar_troca(id):
    t = TrocaSolicitacao.query.get_or_404(id)
    if t.status != "pendente":
        flash("Esta solicitação já foi tratada.", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    orig_e = Escala.query.get(t.origem_escala_id)
    if not orig_e:
        flash("Plantão de origem inválido.", "danger")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    solicitante = Cooperado.query.get(t.solicitante_id)
    destinatario = Cooperado.query.get(t.destino_id)
    if not solicitante or not destinatario:
        flash("Cooperado(s) inválido(s) na solicitação.", "danger")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    wd_o = _weekday_from_data_str(orig_e.data)
    buck_o = _turno_bucket(orig_e.turno, orig_e.horario)
    minhas = (Escala.query
              .filter_by(cooperado_id=destinatario.id)
              .order_by(Escala.id.asc()).all())
    candidatas = [e for e in minhas
                  if _weekday_from_data_str(e.data) == wd_o
                  and _turno_bucket(e.turno, e.horario) == buck_o]

    if len(candidatas) != 1:
        if len(candidatas) == 0:
            flash("Destino não possui plantões compatíveis (mesmo dia da semana e mesmo turno).", "danger")
        else:
            flash("Mais de um plantão compatível encontrado para o destino. Aprove pelo portal do cooperado (onde é possível escolher).", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    dest_e = candidatas[0]

    linhas = [
        {
            "dia": _escala_label(orig_e).split(" • ")[0],
            "turno_horario": " • ".join([x for x in [(orig_e.turno or "").strip(), (orig_e.horario or "").strip()] if x]),
            "contrato": (orig_e.contrato or "").strip(),
            "saiu": solicitante.nome,
            "entrou": destinatario.nome,
        },
        {
            "dia": _escala_label(dest_e).split(" • ")[0],
            "turno_horario": " • ".join([x for x in [(dest_e.turno or "").strip(), (dest_e.horario or "").strip()] if x]),
            "contrato": (dest_e.contrato or "").strip(),
            "saiu": destinatario.nome,
            "entrou": solicitante.nome,
        }
    ]
    afetacao_json = {"linhas": linhas}

    solicitante_id = orig_e.cooperado_id
    destino_id = dest_e.cooperado_id
    orig_e.cooperado_id, dest_e.cooperado_id = destino_id, solicitante_id

    t.status = "aprovada"
    t.aplicada_em = datetime.utcnow()
    prefix = "" if not (t.mensagem and t.mensagem.strip()) else (t.mensagem.rstrip() + "\n")
    t.mensagem = prefix + "__AFETACAO_JSON__:" + json.dumps(afetacao_json, ensure_ascii=False)

    when = datetime.utcnow()
    _prune_histories()
    _log_troca_historico_rows(t.id, linhas, solicitante=solicitante, destinatario=destinatario, tipo="troca", when=when)
    for linha in linhas:
        turno_txt = linha.get("turno_horario", "")
        turno_part = turno_txt.split("•")[0].strip() if turno_txt else ""
        horario_part = turno_txt.split("•", 1)[1].strip() if "•" in turno_txt else ""
        _log_escala_historico(
            origem="troca_aprovada",
            acao="troca",
            troca_ref_id=t.id,
            grupo_ref=str(uuid.uuid4()),
            admin_usuario_id=session.get("user_id"),
            data=linha.get("dia", ""),
            turno=turno_part,
            horario=horario_part,
            contrato=linha.get("contrato", ""),
            cooperado_nome=linha.get("entrou", ""),
            saiu_nome=linha.get("saiu", ""),
            entrou_nome=linha.get("entrou", ""),
            snapshot_em=when,
        )
    db.session.commit()
    flash("Troca aprovada e aplicada com sucesso!", "success")
    return redirect(url_for("admin_dashboard", tab="escalas"))

@app.post("/admin/trocas/<int:id>/recusar")
@admin_perm_required("escalas", "editar")
def admin_recusar_troca(id):
    t = TrocaSolicitacao.query.get_or_404(id)
    if t.status != "pendente":
        flash("Esta solicitação já foi tratada.", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))
    t.status = "recusada"
    db.session.commit()
    flash("Solicitação recusada.", "info")
    return redirect(url_for("admin_dashboard", tab="escalas"))


# --- Admin tool: aplicar ON DELETE CASCADE nas FKs (Postgres) ---
@app.get("/admin/tools/apply_fk_cascade")
@admin_required
def apply_fk_cascade():
    """
    Aplica/garante ON DELETE CASCADE nas FKs relevantes (Postgres).
    Tudo está dentro de uma string SQL, evitando SyntaxError no deploy.
    """
    from sqlalchemy import text as sa_text

    sql = """
BEGIN;

-- =========================
-- AVALIAÇÕES (já existia)
-- =========================
-- ajusta FK de avaliacoes.lancamento_id
ALTER TABLE public.avaliacoes
  DROP CONSTRAINT IF EXISTS avaliacoes_lancamento_id_fkey;
ALTER TABLE public.avaliacoes
  ADD CONSTRAINT avaliacoes_lancamento_id_fkey
  FOREIGN KEY (lancamento_id)
  REFERENCES public.lancamentos (id)
  ON DELETE CASCADE;

-- cria/garante CASCADE para avaliacoes_restaurante.lancamento_id
DO $do$
BEGIN
    IF NOT EXISTS (
        SELECT 1
        FROM information_schema.table_constraints
        WHERE constraint_name = 'av_rest_lancamento_id_fkey'
          AND table_name = 'avaliacoes_restaurante'
          AND table_schema = 'public'
    ) THEN
        ALTER TABLE public.avaliacoes_restaurante
          ADD CONSTRAINT av_rest_lancamento_id_fkey
          FOREIGN KEY (lancamento_id)
          REFERENCES public.lancamentos (id)
          ON DELETE CASCADE;
    ELSE
        -- garante o CASCADE (drop/add)
        EXECUTE $$ALTER TABLE public.avaliacoes_restaurante
                 DROP CONSTRAINT IF EXISTS av_rest_lancamento_id_fkey$$;
        EXECUTE $$ALTER TABLE public.avaliacoes_restaurante
                 ADD CONSTRAINT av_rest_lancamento_id_fkey
                 FOREIGN KEY (lancamento_id)
                 REFERENCES public.lancamentos (id)
                 ON DELETE CASCADE$$;
    END IF;
END
$do$;

-- =========================
-- ESCALAS
-- =========================
-- cooperado_id -> cooperados(id) ON DELETE CASCADE
ALTER TABLE public.escalas
  DROP CONSTRAINT IF EXISTS escalas_cooperado_id_fkey;
ALTER TABLE public.escalas
  ADD CONSTRAINT escalas_cooperado_id_fkey
  FOREIGN KEY (cooperado_id)
  REFERENCES public.cooperados (id)
  ON DELETE CASCADE;

-- restaurante_id -> restaurantes(id) ON DELETE CASCADE
ALTER TABLE public.escalas
  DROP CONSTRAINT IF EXISTS escalas_restaurante_id_fkey;
ALTER TABLE public.escalas
  ADD CONSTRAINT escalas_restaurante_id_fkey
  FOREIGN KEY (restaurante_id)
  REFERENCES public.restaurantes (id)
  ON DELETE CASCADE;

-- =========================
-- TROCAS
-- =========================
-- solicitante_id -> cooperados(id) ON DELETE CASCADE
ALTER TABLE public.trocas
  DROP CONSTRAINT IF EXISTS trocas_solicitante_id_fkey;
ALTER TABLE public.trocas
  ADD CONSTRAINT trocas_solicitante_id_fkey
  FOREIGN KEY (solicitante_id)
  REFERENCES public.cooperados (id)
  ON DELETE CASCADE;

-- destino_id -> cooperados(id) ON DELETE CASCADE
ALTER TABLE public.trocas
  DROP CONSTRAINT IF EXISTS trocas_destino_id_fkey;
ALTER TABLE public.trocas
  ADD CONSTRAINT trocas_destino_id_fkey
  FOREIGN KEY (destino_id)
  REFERENCES public.cooperados (id)
  ON DELETE CASCADE;

-- origem_escala_id -> escalas(id) ON DELETE CASCADE
ALTER TABLE public.trocas
  DROP CONSTRAINT IF EXISTS trocas_origem_escala_id_fkey;
ALTER TABLE public.trocas
  ADD CONSTRAINT trocas_origem_escala_id_fkey
  FOREIGN KEY (origem_escala_id)
  REFERENCES public.escalas (id)
  ON DELETE CASCADE;

COMMIT;
"""

    try:
        if _is_sqlite():
            flash("SQLite local: esta operação é específica de Postgres (sem efeito aqui).", "warning")
            return redirect(url_for("admin_dashboard", tab="config"))

        db.session.execute(sa_text(sql))
        db.session.commit()
        flash("FKs com ON DELETE CASCADE aplicadas com sucesso.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao aplicar FKs: {e}", "danger")
    return redirect(url_for("admin_dashboard", tab="config"))


# =========================
# Documentos (Admin)
# =========================
@app.route("/documentos/<int:coop_id>", methods=["GET", "POST"])
@admin_required
def editar_documentos(coop_id):
    c = Cooperado.query.get_or_404(coop_id)

    if request.method == "POST":
        f = request.form
        c.cnh_numero = f.get("cnh_numero")
        c.placa = f.get("placa")
        def parse_date_local(s):
            try:
                return datetime.strptime(s, "%Y-%m-%d").date() if s else None
            except Exception:
                return None
        c.cnh_validade = parse_date_local(f.get("cnh_validade"))
        c.placa_validade = parse_date_local(f.get("placa_validade"))
        c.ultima_atualizacao = datetime.now()
        db.session.commit()
        flash("Documentos atualizados.", "success")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    tpl = os.path.join("templates", "editar_documentos.html")
    hoje = date.today()
    prazo_final = date(hoje.year, 12, 31)
    if os.path.exists(tpl):
        docinfo = {
            "prazo_final": prazo_final,
            "dias_ate_prazo": max(0, (prazo_final - hoje).days),
            "cnh": {
                "numero": c.cnh_numero,
                "validade": c.cnh_validade,
                "prox_validade": _prox_ocorrencia_anual(c.cnh_validade),
                "ok": (c.cnh_validade is not None and c.cnh_validade >= hoje),
                "modo": "auto",
            },
            "placa": {
                "numero": c.placa,
                "validade": c.placa_validade,
                "prox_validade": _prox_ocorrencia_anual(c.placa_validade),
                "ok": (c.placa_validade is not None and c.placa_validade >= hoje),
                "modo": "auto",
            }
        }
        return render_template("editar_documentos.html", cooperado=c, docinfo=docinfo)

    return f"""
    <div style="max-width:560px;margin:30px auto;font-family:Arial">
      <h3>Documentos — {c.nome}</h3>
      <form method="POST">
        <label>CNH (número)</label><br>
        <input name="cnh_numero" value="{c.cnh_numero or ''}" style="width:100%;padding:8px"><br><br>
        <label>Validade CNH</label><br>
        <input type="date" name="cnh_validade" value="{c.cnh_validade.strftime('%Y-%m-%d') if c.cnh_validade else ''}" style="width:100%;padding:8px"><br><br>
        <label>Placa</label><br>
        <input name="placa" value="{c.placa or ''}" style="width:100%;padding:8px"><br><br>
        <label>Validade da Placa</label><br>
        <input type="date" name="placa_validade" value="{c.placa_validade.strftime('%Y-%m-%d') if c.placa_validade else ''}" style="width:100%;padding:8px"><br><br>
        <button style="padding:10px 16px">Salvar</button>
        <a href="{url_for('admin_dashboard', tab='escalas')}" style="margin-left:10px">Voltar</a>
      </form>
    </div>
    """

# =========================
# PORTAL COOPERADO
# =========================
@app.route("/portal/cooperado")
@role_required("cooperado")
def portal_cooperado():
    u_id = session.get("user_id")
    coop = Cooperado.query.filter_by(usuario_id=u_id).first()
    if not coop:
        return "<p style='font-family:Arial;margin:40px'>Seu usuário não está vinculado a um cooperado. Avise o administrador.</p>"

    try:
        coop.usuario = coop.usuario_ref.usuario
    except Exception:
        coop.usuario = ""

    # ---------- FILTRO POR DATA (padrão = HOJE) ----------
    di = _parse_date(request.args.get("data_inicio"))
    df = _parse_date(request.args.get("data_fim"))

    if di and not df:
        df = di
    if df and not di:
        di = df
    elif not di and not df:
        hoje = date.today()
        di = hoje - timedelta(days=hoje.weekday())   # segunda
        df = di + timedelta(days=6)                  # domingo

    def in_range(qs, col):
        return qs.filter(col >= di, col <= df)

    # =========================
    # Produções (Lançamentos)
    # =========================
    ql = in_range(Lancamento.query.filter_by(cooperado_id=coop.id), Lancamento.data)
    producoes = ql.order_by(Lancamento.data.desc(), Lancamento.id.desc()).all()

    ids = [l.id for l in producoes]
    minhas = {}

    if ids:
        rows = (
            db.session.query(
                AvaliacaoRestaurante.lancamento_id,
                AvaliacaoRestaurante.estrelas_geral
            )
            .filter(
                AvaliacaoRestaurante.lancamento_id.in_(ids),
                AvaliacaoRestaurante.cooperado_id == coop.id
            )
            .all()
        )
        minhas = {lid: nota for lid, nota in rows}

    for l in producoes:
        l.minha_avaliacao = minhas.get(l.id)

    # =========================
    # Receitas / Despesas
    # =========================
    qr = in_range(ReceitaCooperado.query.filter_by(cooperado_id=coop.id), ReceitaCooperado.data)
    receitas_coop = qr.order_by(ReceitaCooperado.data.desc(), ReceitaCooperado.id.desc()).all()

    qd = DespesaCooperado.query.filter_by(cooperado_id=coop.id)

    if di and df:
        qd = qd.filter(
            DespesaCooperado.data_inicio <= df,
            DespesaCooperado.data_fim >= di,
        )
    elif di:
        qd = qd.filter(DespesaCooperado.data_fim >= di)
    elif df:
        qd = qd.filter(DespesaCooperado.data_inicio <= df)

    despesas_coop = qd.order_by(
        DespesaCooperado.data_fim.desc().nullslast(),
        DespesaCooperado.id.desc()
    ).all()

    # =========================
    # Totais
    # =========================
    total_bruto = (
        sum((l.valor or 0.0) for l in producoes)
        + sum((r.valor or 0.0) for r in receitas_coop)
    )

    inss_valor = sum((l.valor or 0.0) * 0.04 for l in producoes)
    sest_valor = sum((l.valor or 0.0) * 0.005 for l in producoes)

    encargos_valor = inss_valor + sest_valor

    debt_snapshot = _compute_coop_debt_snapshot(coop.id, di, df)

    # só o que venceu entra automaticamente no período; o futuro fica em "a descontar"
    total_descontos = sum((it['pago_manual'] + it['pago_auto']) for it in debt_snapshot['itens'])
    total_liquido = max(0.0, total_bruto - encargos_valor - total_descontos)
    saldo_devedor = debt_snapshot['saldo_devedor']
    total_a_descontar = debt_snapshot['a_descontar']
    despesas_detalhadas = debt_snapshot['itens']


       # =====================================================
    # MÉTRICAS DA VIDA (NÃO DEPENDE DO FILTRO DE DATA)
    # =====================================================

    total_entregas_vida = (
        db.session.query(func.count(Lancamento.id))
        .filter(Lancamento.cooperado_id == coop.id)
        .scalar()
        or 0
    )

    nota_vida = (
        db.session.query(func.avg(AvaliacaoCooperado.estrelas_geral))
        .filter(AvaliacaoCooperado.cooperado_id == coop.id)
        .scalar()
    )

    nota_vida = float(nota_vida or 5.0)

# =====================================================
# TOTAL DE ENTREGAS DA VIDA (SEM FILTRO DE DATA)
# =====================================================

    total_entregas_vida = db.session.query(func.count(Lancamento.id))\
        .filter(Lancamento.cooperado_id == coop.id)\
        .scalar() or 0


    # =========================
    # Config / Complemento
    # =========================
    cfg = get_config()
    salario_minimo = cfg.salario_minimo or 0.0
    inss_complemento = salario_minimo * 0.20

    today = date.today()

    def dias_para_3112():
        alvo = date(today.year, 12, 31)
        if today > alvo:
            alvo = date(today.year + 1, 12, 31)
        return (alvo - today).days

    doc_cnh = {
        "numero": coop.cnh_numero,
        "vencimento": coop.cnh_validade,
        "ok": (coop.cnh_validade is not None and coop.cnh_validade >= today),
        "dias_para_prazo": dias_para_3112(),
    }
    doc_placa = {
        "numero": coop.placa,
        "vencimento": coop.placa_validade,
        "ok": (coop.placa_validade is not None and coop.placa_validade >= today),
        "dias_para_prazo": dias_para_3112(),
    }

    # ---------- ESCALA (dedupe + ordenação cronológica robusta) ----------
    raw_escala = (Escala.query
                 .filter_by(cooperado_id=coop.id)
                 .order_by(Escala.id.asc())
                 .all())

    import unicodedata as _u, re as _re
    def _norm_c(s: str) -> str:
        s = _u.normalize("NFD", str(s or "").lower())
        s = "".join(ch for ch in s if _u.category(ch) != "Mn")
        return _re.sub(r"[^a-z0-9]+", " ", s).strip()

    def _score(e):
        h = (e.horario or "").strip()
        return (1 if h else 0, len(h), e.id)

    def _to_date_from_str(s: str):
        m = _re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', str(s or ''))
        if not m:
            return None
        d_, mth, y = map(int, m.groups())
        if y < 100:
            y += 2000
        try:
            return date(y, mth, d_)
        except Exception:
            return None

    def _mins(h):
        m = _re.search(r'(\d{1,2}):(\d{2})', str(h or ''))
        if not m:
            return 24*60 + 59
        hh, mm = map(int, m.groups())
        return hh*60 + mm

    def _bucket_idx(turno, horario):
        b = (_turno_bucket(turno, horario) or "").lower()
        if "dia" in b:
            return 1
        if "noite" in b:
            return 2
        mins = _mins(horario)
        return 2 if (mins >= 17*60 or mins <= 6*60) else 1

    best = {}
    for e in raw_escala:
        key = (_norm_c(e.data), _norm_c(e.turno), _norm_c(e.contrato))
        cur = best.get(key)
        if not cur or _score(e) > _score(cur):
            best[key] = e

    cand = list(best.values())
    for e in cand:
        d = _to_date_from_str(e.data) or date.min
        mins = _mins(e.horario or "")
        bidx = _bucket_idx(e.turno, e.horario)
        e._ord = (d.toordinal(), bidx, mins, (e.contrato or ""), e.id)

    minha_escala = sorted(cand, key=lambda x: x._ord)

    for e in minha_escala:
        dt = _to_date_from_str(e.data)
        if dt is None:
            status = 'unknown'
        else:
            if dt < today:
                status = 'past'
            elif dt == today:
                status = 'today'
            elif dt == today + timedelta(days=1):
                status = 'tomorrow'
            else:
                status = 'future'
        e.status = status
        e.status_color = (
            '#ef4444' if status == 'past' else
            '#22c55e' if status == 'today' else
            '#3b82f6' if status in ('tomorrow', 'future') else
            'transparent'
        )

    minha_escala_json = []
    for e in minha_escala:
        minha_escala_json.append({
            "id": e.id,
            "data": e.data or "",
            "turno": e.turno or "",
            "horario": e.horario or "",
            "contrato": e.contrato or "",
            "weekday": _weekday_from_data_str(e.data),
            "turno_bucket": _turno_bucket(e.turno, e.horario),
        })

    # ---------- Trocas ----------
    coops = (Cooperado.query
             .filter(Cooperado.id != coop.id)
             .order_by(Cooperado.nome.asc())
             .all())
    cooperados_json = [
        {"id": c.id, "nome": c.nome, "foto_url": (c.foto_url or "")}
        for c in coops
    ]
    cooperados_escalas_map = {}
    for c in coops:
        escalas_c = (Escala.query.filter_by(cooperado_id=c.id).order_by(Escala.data.asc(), Escala.id.asc()).all())
        cooperados_escalas_map[str(c.id)] = [
            {
                "id": e.id,
                "data": e.data or "",
                "turno": e.turno or "",
                "horario": e.horario or "",
                "contrato": e.contrato or "",
                "weekday": _weekday_from_data_str(e.data),
                "turno_bucket": _turno_bucket(e.turno, e.horario),
            }
            for e in escalas_c
        ]

    def _escala_desc(e: Escala | None) -> str:
        return _escala_label(e)

    rx = (TrocaSolicitacao.query
          .filter(TrocaSolicitacao.destino_id == coop.id)
          .order_by(TrocaSolicitacao.id.desc())
          .all())

    trocas_recebidas_pendentes = []
    trocas_recebidas_historico = []
    for t in rx:
        solicitante = Cooperado.query.get(t.solicitante_id)
        orig = Escala.query.get(t.origem_escala_id)

        mensagem_limpa = _strip_afetacao_blob(t.mensagem)
        linhas_afetadas = _parse_linhas_from_msg(t.mensagem) if t.status == "aprovada" else []

        item = {
            "id": t.id,
            "status": t.status,
            "mensagem": mensagem_limpa,
            "criada_em": t.criada_em,
            "aplicada_em": t.aplicada_em,
            "solicitante": solicitante,
            "origem": orig,
            "origem_desc": _escala_desc(orig),
            "linhas_afetadas": linhas_afetadas,
            "origem_weekday": _weekday_from_data_str(orig.data) if orig else None,
            "origem_turno_bucket": _turno_bucket(orig.turno if orig else None, orig.horario if orig else None),
        }

        (trocas_recebidas_pendentes if t.status == "pendente" else trocas_recebidas_historico).append(item)

    ex = (TrocaSolicitacao.query
          .filter(TrocaSolicitacao.solicitante_id == coop.id)
          .order_by(TrocaSolicitacao.id.desc())
          .all())

    trocas_enviadas = []
    for t in ex:
        destino = Cooperado.query.get(t.destino_id)
        orig = Escala.query.get(t.origem_escala_id)
        mensagem_limpa = _strip_afetacao_blob(t.mensagem)
        linhas_afetadas = _parse_linhas_from_msg(t.mensagem) if t.status == "aprovada" else []
        trocas_enviadas.append({
            "id": t.id,
            "status": t.status,
            "mensagem": mensagem_limpa,
            "criada_em": t.criada_em,
            "aplicada_em": t.aplicada_em,
            "destino": destino,
            "origem": orig,
            "origem_desc": _escala_desc(orig),
            "linhas_afetadas": linhas_afetadas,
        })


    # return TEM que ficar aqui (fora do for)
    return render_template(
        "painel_cooperado.html",
        cooperado=coop,
        producoes=producoes,
        receitas_coop=receitas_coop,
        despesas_coop=despesas_coop,
        total_bruto=total_bruto,
        inss_valor=inss_valor,
        sest_senat_valor=sest_valor,
        total_descontos=total_descontos,
        total_liquido=total_liquido,
        inss_complemento=inss_complemento,
        salario_minimo=salario_minimo,
        current_year=today.year,
        doc_cnh=doc_cnh,
        doc_placa=doc_placa,
        minha_escala=minha_escala,
        minha_escala_json=minha_escala_json,
        cooperados_json=cooperados_json,
        cooperados_escalas_map=cooperados_escalas_map,
        trocas_recebidas_pendentes=trocas_recebidas_pendentes,
        trocas_recebidas_historico=trocas_recebidas_historico,
        trocas_enviadas=trocas_enviadas,
        
        # MÉTRICAS DA VIDA
        nota_vida=nota_vida,
        total_entregas_vida=total_entregas_vida,
        data_inicio=di,
        data_fim=df,
        saldo_devedor=saldo_devedor,
        total_a_descontar=total_a_descontar,
        despesas_detalhadas=despesas_detalhadas,
    )

# === AVALIAR RESTAURANTE (cooperado -> restaurante)
# Duas rotas para a MESMA função e MESMO endpoint (o do template):
@app.post("/coop/avaliar/restaurante/<int:lanc_id>")
@app.post("/producoes/<int:lanc_id>/avaliar")
@role_required("cooperado")
def producoes_avaliar(lanc_id):
    # 1) Cooperado logado
    u_id = session.get("user_id")
    coop = Cooperado.query.filter_by(usuario_id=u_id).first_or_404()

    # 2) Lançamento existe e é dele
    lanc = Lancamento.query.get_or_404(lanc_id)
    if lanc.cooperado_id != coop.id:
        abort(403)

    # 3) Já existe avaliação DESTE cooperado para ESTE lançamento?
    ja = (AvaliacaoRestaurante.query
          .filter_by(lancamento_id=lanc.id, cooperado_id=coop.id)
          .first())
    if ja:
        flash("Você já avaliou esta produção.", "info")
        return _portal_cooperado_redirect_tab("producoes")

    # -------- Helpers locais --------
    def _clamp_star_local(v):
        try:
            n = int(float(v))
        except Exception:
            return None
        return n if 1 <= n <= 5 else None

    def _get(v):
        return request.form.get(v)

    # 4) Campos do form — SOMENTE 3 dimensões, com retrocompat:
    amb  = _clamp_star_local(_get("av_ambiente")    or _get("av_apresentacao"))  # retro
    trat = _clamp_star_local(_get("av_tratamento")  or _get("av_educacao"))      # retro
    sup  = _clamp_star_local(_get("av_suporte")     or _get("av_eficiencia"))    # retro

    # Se vier 'nota' / 'av_geral', usa como fallback nas faltantes
    nota = _clamp_star_local(_get("nota") or _get("av_geral"))
    if nota is not None:
        if amb  is None: amb  = nota
        if trat is None: trat = nota
        if sup  is None: sup  = nota

    # Validação
    if not (amb and trat and sup):
        flash("Selecione notas (1..5) para Ambiente, Tratamento e Suporte.", "warning")
        return _portal_cooperado_redirect_tab("producoes")

    # Média (geral e ponderada iguais, com arredondamentos diferentes)
    media = (amb + trat + sup) / 3.0
    estrelas_geral  = round(media, 1)
    media_ponderada = round(media, 2)

    comentario = (_get("av_comentario") or "").strip() or None

    # Derivados opcionais — mantém compat se suas funções existirem
    try:
        senti = _analise_sentimento(comentario) if comentario else None
    except Exception:
        senti = None
    try:
        temas = "; ".join(_identifica_temas(comentario)) if comentario else None
    except Exception:
        temas = None
    try:
        crise = _sinaliza_crise(estrelas_geral, comentario)
    except Exception:
        crise = False

    a = AvaliacaoRestaurante(
        restaurante_id=lanc.restaurante_id,
        cooperado_id=coop.id,
        lancamento_id=lanc.id,

        estrelas_ambiente=amb,
        estrelas_tratamento=trat,
        estrelas_suporte=sup,
        estrelas_geral=estrelas_geral,
        media_ponderada=media_ponderada,

        comentario=comentario,
        sentimento=senti,
        temas=temas,
        alerta_crise=crise,
        # criado_em => default no modelo
    )
    db.session.add(a)

    try:
        db.session.commit()
        flash("Avaliação do restaurante registrada.", "success")
    except IntegrityError:
        db.session.rollback()
        flash("Avaliação já registrada para este lançamento.", "info")

    return _portal_cooperado_redirect_tab("producoes")


# === ALIAS DO PAINEL: /painel/cooperado  -> redireciona para o endpoint oficial
@app.get("/painel/cooperado")
@role_required("cooperado")
def coop_dashboard_alias():
    return redirect(url_for("coop_dashboard"))
    

def _portal_cooperado_redirect_tab(tab: str = "resumo", **extra):
    params = {"active_tab": tab}
    for k, v in (extra or {}).items():
        if v is not None and v != "":
            params[k] = v
    return redirect(url_for("portal_cooperado", **params))


@app.route("/escala/solicitar_troca", methods=["POST"])
@role_required("cooperado")
def solicitar_troca():
    u_id = session.get("user_id")
    me = Cooperado.query.filter_by(usuario_id=u_id).first()
    if not me:
        abort(403)

    from_escala_id = request.form.get("from_escala_id", type=int)
    to_cooperado_id = request.form.get("to_cooperado_id", type=int)
    destino_escala_id = request.form.get("destino_escala_id", type=int)
    mensagem = (request.form.get("mensagem") or "").strip()

    if not from_escala_id or not to_cooperado_id:
        flash("Selecione o seu turno e o cooperado de destino.", "warning")
        return _portal_cooperado_redirect_tab("trocas")

    origem = Escala.query.get(from_escala_id)
    if not origem or origem.cooperado_id != me.id:
        flash("Turno inválido para solicitação.", "danger")
        return _portal_cooperado_redirect_tab("trocas")

    destino = Cooperado.query.get(to_cooperado_id)
    if not destino or destino.id == me.id:
        flash("Cooperado de destino inválido.", "danger")
        return _portal_cooperado_redirect_tab("trocas")

    def _safe_str(v):
        return str(v or "").strip()

    def _norm_horario(v):
        txt = _safe_str(v)
        m = re.findall(r"(\d{1,2}):(\d{2})", txt)
        if m:
            return " | ".join(f"{int(h):02d}:{int(mm):02d}" for h, mm in m)
        return _norm(txt)

    def _escala_signature(e: Escala | None):
        if not e:
            return None
        return {
            "weekday": _weekday_from_data_str(getattr(e, "data", None)),
            "bucket": _turno_bucket(getattr(e, "turno", None), getattr(e, "horario", None)),
            "horario": _norm_horario(getattr(e, "horario", None)),
            "contrato": _norm(getattr(e, "contrato", "") or ""),
        }

    def _same_turno(sig_a, sig_b) -> bool:
        return bool(
            sig_a and sig_b
            and sig_a["weekday"] == sig_b["weekday"]
            and sig_a["bucket"] == sig_b["bucket"]
        )

    nova_sig = _escala_signature(origem)
    if not nova_sig:
        flash("Não foi possível identificar esse turno para a troca.", "danger")
        return _portal_cooperado_redirect_tab("trocas")

    escalas_destino_compativeis = []
    for e_dest in Escala.query.filter_by(cooperado_id=destino.id).order_by(Escala.id.asc()).all():
        sig_dest = _escala_signature(e_dest)
        if sig_dest and _same_turno(sig_dest, nova_sig):
            escalas_destino_compativeis.append(e_dest)

    escala_destino_escolhida = None
    modo_passagem = False
    if escalas_destino_compativeis:
        if destino_escala_id:
            escala_destino_escolhida = next((e for e in escalas_destino_compativeis if int(e.id) == int(destino_escala_id)), None)
            if not escala_destino_escolhida:
                flash("O turno escolhido do cooperado não é compatível com essa troca.", "warning")
                return _portal_cooperado_redirect_tab("trocas")
        elif len(escalas_destino_compativeis) == 1:
            escala_destino_escolhida = escalas_destino_compativeis[0]
        else:
            flash("Esse cooperado possui mais de um turno compatível. Escolha qual turno dele será usado na troca.", "info")
            return _portal_cooperado_redirect_tab("trocas")
    else:
        modo_passagem = True

    contratos_destino = {
        _norm(getattr(e, "contrato", "") or "")
        for e in escalas_destino_compativeis
        if (getattr(e, "contrato", "") or "").strip()
    }
    ids_destino_compativeis = {int(e.id) for e in escalas_destino_compativeis if e.id}

    pendentes_mesma_dupla = (
        TrocaSolicitacao.query
        .filter(TrocaSolicitacao.status == "pendente")
        .filter(
            or_(
                and_(TrocaSolicitacao.solicitante_id == me.id, TrocaSolicitacao.destino_id == destino.id),
                and_(TrocaSolicitacao.solicitante_id == destino.id, TrocaSolicitacao.destino_id == me.id),
            )
        )
        .order_by(TrocaSolicitacao.id.desc())
        .all()
    )

    for t_exist in pendentes_mesma_dupla:
        esc_exist = Escala.query.get(t_exist.origem_escala_id)
        if not esc_exist:
            continue
        exist_sig = _escala_signature(esc_exist)
        if not exist_sig or not _same_turno(nova_sig, exist_sig):
            continue

        contrato_exist = _norm(getattr(esc_exist, "contrato", "") or "")
        mesma_origem = int(t_exist.origem_escala_id or 0) == int(origem.id)

        troca_reversa_equivalente = (
            t_exist.solicitante_id == destino.id and (
                mesma_origem
                or int(getattr(esc_exist, "id", 0) or 0) == int(getattr(escala_destino_escolhida, "id", 0) or 0)
                or int(getattr(esc_exist, "id", 0) or 0) in ids_destino_compativeis
                or (contrato_exist and contrato_exist in contratos_destino)
                or modo_passagem
            )
        )
        troca_mesmo_sentido = (
            t_exist.solicitante_id == me.id and (
                mesma_origem
                or int(getattr(esc_exist, "id", 0) or 0) == int(getattr(escala_destino_escolhida, "id", 0) or 0)
                or (contrato_exist and contrato_exist == _norm(getattr(origem, "contrato", "") or ""))
                or modo_passagem
            )
        )

        if troca_reversa_equivalente:
            flash("Já existe uma solicitação dessa mesma troca. Vá na aba Trocas e aceite a solicitação que já foi enviada.", "info")
            return _portal_cooperado_redirect_tab("trocas")

        if troca_mesmo_sentido:
            flash("Já existe uma solicitação pendente para esse mesmo turno com esse cooperado. Abra a aba Trocas para acompanhar.", "warning")
            return _portal_cooperado_redirect_tab("trocas")

    t = TrocaSolicitacao(
        solicitante_id=me.id,
        destino_id=destino.id,
        origem_escala_id=origem.id,
        mensagem=mensagem or None,
        status="pendente",
    )
    db.session.add(t)
    db.session.commit()

    if modo_passagem:
        flash("Solicitação de passagem de turno enviada com sucesso. Agora o cooperado precisa aceitar.", "success")
    else:
        flash("Solicitação de troca enviada com sucesso.", "success")
    return _portal_cooperado_redirect_tab("trocas")

@app.post("/trocas/<int:troca_id>/aceitar")
@role_required("cooperado")
def aceitar_troca(troca_id):
    u_id = session.get("user_id")
    me = Cooperado.query.filter_by(usuario_id=u_id).first()
    t = TrocaSolicitacao.query.get_or_404(troca_id)

    if not me or t.destino_id != me.id:
        abort(403)

    if t.status != "pendente":
        flash("Esta solicitação já foi tratada.", "warning")
        return _portal_cooperado_redirect_tab("trocas")

    destino_escala_id = request.form.get("destino_escala_id", type=int)
    orig_e = Escala.query.get(t.origem_escala_id)

    if not orig_e:
        flash("Turno de origem inválido.", "danger")
        return _portal_cooperado_redirect_tab("trocas")

    minhas = Escala.query.filter_by(cooperado_id=me.id).order_by(Escala.id.asc()).all()
    wd_o = _weekday_from_data_str(orig_e.data)
    buck_o = _turno_bucket(orig_e.turno, orig_e.horario)
    candidatas = [
        e for e in minhas
        if _weekday_from_data_str(e.data) == wd_o
        and _turno_bucket(e.turno, e.horario) == buck_o
    ]

    dest_e = None
    if destino_escala_id:
        dest_e = Escala.query.get(destino_escala_id)
        if not dest_e or dest_e.cooperado_id != me.id:
            flash("Seleção de turno inválida.", "danger")
            return _portal_cooperado_redirect_tab("trocas")
    elif len(candidatas) == 1:
        dest_e = candidatas[0]
    elif len(candidatas) > 1:
        flash("Escolha na aba Trocas qual dos seus turnos compatíveis deseja usar.", "warning")
        return _portal_cooperado_redirect_tab("trocas")

    solicitante = Cooperado.query.get(t.solicitante_id)
    destinatario = me

    if dest_e is None:
        linhas = [{
            "dia": _escala_label(orig_e).split(" • ")[0],
            "turno_horario": " • ".join([x for x in [(orig_e.turno or "").strip(), (orig_e.horario or "").strip()] if x]),
            "contrato": (orig_e.contrato or "").strip(),
            "saiu": solicitante.nome if solicitante else "",
            "entrou": destinatario.nome,
        }]
        afetacao_json = {"linhas": linhas}
        orig_e.cooperado_id = destinatario.id
        orig_e.cooperado_nome = None
        t.status = "aprovada"
        t.aplicada_em = datetime.utcnow()
        prefix = "" if not (t.mensagem and t.mensagem.strip()) else (t.mensagem.rstrip() + "\n")
        t.mensagem = prefix + "__AFETACAO_JSON__:" + json.dumps(afetacao_json, ensure_ascii=False)
        when = datetime.utcnow()
        _prune_histories()
        _log_troca_historico_rows(t.id, linhas, solicitante=solicitante, destinatario=destinatario, tipo="passagem", when=when)
        for linha in linhas:
            turno_txt = linha.get("turno_horario", "")
            turno_part = turno_txt.split("•")[0].strip() if turno_txt else ""
            horario_part = turno_txt.split("•", 1)[1].strip() if "•" in turno_txt else ""
            _log_escala_historico(
                origem="passagem_aprovada",
                acao="passagem",
                escala_ref_id=orig_e.id,
                troca_ref_id=t.id,
                grupo_ref=str(uuid.uuid4()),
                data=linha.get("dia", ""),
                turno=turno_part,
                horario=horario_part,
                contrato=linha.get("contrato", ""),
                cooperado_id=destinatario.id,
                cooperado_nome=linha.get("entrou", ""),
                saiu_nome=linha.get("saiu", ""),
                entrou_nome=linha.get("entrou", ""),
                snapshot_em=when,
            )
        db.session.commit()
        flash("Turno passado com sucesso!", "success")
        return _portal_cooperado_redirect_tab("trocas")

    wd_dest = _weekday_from_data_str(dest_e.data)
    buck_dest = _turno_bucket(dest_e.turno, dest_e.horario)
    if wd_o is None or wd_dest is None or wd_o != wd_dest or buck_o != buck_dest:
        flash("Troca incompatível: precisa ser no mesmo dia da semana e no mesmo turno.", "danger")
        return _portal_cooperado_redirect_tab("trocas")

    linhas = [
        {
            "dia": _escala_label(orig_e).split(" • ")[0],
            "turno_horario": " • ".join([x for x in [(orig_e.turno or "").strip(), (orig_e.horario or "").strip()] if x]),
            "contrato": (orig_e.contrato or "").strip(),
            "saiu": solicitante.nome if solicitante else "",
            "entrou": destinatario.nome,
        },
        {
            "dia": _escala_label(dest_e).split(" • ")[0],
            "turno_horario": " • ".join([x for x in [(dest_e.turno or "").strip(), (dest_e.horario or "").strip()] if x]),
            "contrato": (dest_e.contrato or "").strip(),
            "saiu": destinatario.nome,
            "entrou": solicitante.nome if solicitante else "",
        }
    ]
    afetacao_json = {"linhas": linhas}

    solicitante_id = orig_e.cooperado_id
    destino_id = dest_e.cooperado_id
    orig_e.cooperado_id = destino_id
    dest_e.cooperado_id = solicitante_id
    if orig_e.cooperado_id:
        orig_e.cooperado_nome = None
    if dest_e.cooperado_id:
        dest_e.cooperado_nome = None

    t.status = "aprovada"
    t.aplicada_em = datetime.utcnow()
    prefix = "" if not (t.mensagem and t.mensagem.strip()) else (t.mensagem.rstrip() + "\n")
    t.mensagem = prefix + "__AFETACAO_JSON__:" + json.dumps(afetacao_json, ensure_ascii=False)

    when = datetime.utcnow()
    _prune_histories()
    _log_troca_historico_rows(t.id, linhas, solicitante=solicitante, destinatario=destinatario, tipo="troca", when=when)
    for linha in linhas:
        turno_txt = linha.get("turno_horario", "")
        turno_part = turno_txt.split("•")[0].strip() if turno_txt else ""
        horario_part = turno_txt.split("•", 1)[1].strip() if "•" in turno_txt else ""
        _log_escala_historico(
            origem="troca_aprovada",
            acao="troca",
            troca_ref_id=t.id,
            grupo_ref=str(uuid.uuid4()),
            data=linha.get("dia", ""),
            turno=turno_part,
            horario=horario_part,
            contrato=linha.get("contrato", ""),
            cooperado_nome=linha.get("entrou", ""),
            saiu_nome=linha.get("saiu", ""),
            entrou_nome=linha.get("entrou", ""),
            snapshot_em=when,
        )
    db.session.commit()
    flash("Troca aplicada com sucesso!", "success")
    return _portal_cooperado_redirect_tab("trocas")

@app.post("/trocas/<int:troca_id>/recusar")
@role_required("cooperado")
def recusar_troca(troca_id):
    u_id = session.get("user_id")
    me = Cooperado.query.filter_by(usuario_id=u_id).first()
    t = TrocaSolicitacao.query.get_or_404(troca_id)

    if not me or t.destino_id != me.id:
        abort(403)

    if t.status != "pendente":
        flash("Esta solicitação já foi tratada.", "warning")
        return _portal_cooperado_redirect_tab("trocas")

    t.status = "recusada"
    db.session.commit()

    flash("Solicitação recusada.", "info")
    return _portal_cooperado_redirect_tab("trocas")



# =========================
# PORTAL RESTAURANTE
# =========================
@app.route("/portal/restaurante")
@role_required("restaurante")
def portal_restaurante():
    from datetime import date, timedelta, datetime
    import re
    from werkzeug.routing import BuildError

    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first()
    if not rest:
        return (
            "<p style='font-family:Arial;margin:40px'>"
            "Seu usuário não está vinculado a um estabelecimento. Avise o administrador."
            "</p>"
        )

    # Abas/visões
    view = (request.args.get("view", "lancar") or "lancar").strip().lower()

    # ---- helper mês YYYY-MM
    def _parse_yyyy_mm_local(s: str):
        if not s:
            return None, None
        m = re.fullmatch(r"(\d{4})-(\d{2})", s.strip())
        if not m:
            return None, None
        y = int(m.group(1))
        mth = int(m.group(2))
        try:
            di_ = date(y, mth, 1)
            if mth == 12:
                df_ = date(y + 1, 1, 1) - timedelta(days=1)
            else:
                df_ = date(y, mth + 1, 1) - timedelta(days=1)
            return di_, df_
        except Exception:
            return None, None

    # -------------------- FILTRO DE PERÍODO --------------------
    di = _parse_date(request.args.get("data_inicio"))
    df = _parse_date(request.args.get("data_fim"))

    mes = (request.args.get("mes") or "").strip()
    periodo_desc = None

    if mes:
        di_mes, df_mes = _parse_yyyy_mm_local(mes)
        if di_mes and df_mes:
            di, df = di_mes, df_mes
            periodo_desc = "mês"

    if not di or not df:
        wd_map = {"seg-dom": 0, "sab-sex": 5, "sex-qui": 4}
        start_wd = wd_map.get(getattr(rest, "periodo", None), 0)
        hoje = date.today()
        delta = (hoje.weekday() - start_wd) % 7
        di_auto = hoje - timedelta(days=delta)
        df_auto = di_auto + timedelta(days=6)
        di = di or di_auto
        df = df or df_auto
        periodo_desc = periodo_desc or getattr(rest, "periodo", "seg-dom")
    else:
        periodo_desc = periodo_desc or "personalizado"

    # -------------------- HELPER: contrato do restaurante --------------------
    def contrato_bate_restaurante(contrato: str, rest_nome: str) -> bool:
        a = " ".join(_normalize_name(contrato or ""))
        b = " ".join(_normalize_name(rest_nome or ""))
        if not a or not b:
            return False
        return a == b or a in b or b in a

    # -------------------- ESCALA (Quem trabalha) --------------------
    ref = _parse_date(request.args.get("ref")) or date.today()
    modo = request.args.get("modo", "semana")

    if modo == "dia":
        dias_list = [ref]
    else:
        semana_inicio = ref - timedelta(days=ref.weekday())
        dias_list = [semana_inicio + timedelta(days=i) for i in range(7)]

    escalas_all = (
        db.session.query(Escala)
        .outerjoin(Cooperado, Escala.cooperado_id == Cooperado.id)
        .outerjoin(Usuario, Cooperado.usuario_id == Usuario.id)
        .filter(
            or_(
                Escala.cooperado_id.is_(None),
                Usuario.ativo.is_(True)
            )
        )
        .order_by(Escala.id.asc())
        .all()
    )

    eff_map = _carry_forward_contrato(escalas_all)

    escalas_rest = [
        e for e in escalas_all
        if contrato_bate_restaurante(eff_map.get(e.id, e.contrato or ""), rest.nome)
    ]

    if not escalas_rest:
        escalas_rest = [
            e for e in escalas_all
            if (e.contrato or "").strip() == (rest.nome or "").strip()
        ]

    agenda = {d: [] for d in dias_list}
    seen = {d: set() for d in dias_list}

    for e in escalas_rest:
        dt = _parse_data_escala_str(e.data)
        wd = _weekday_from_data_str(e.data)

        for d in dias_list:
            hit = (dt and dt == d) or (wd and wd == ((d.weekday() % 7) + 1))
            if not hit:
                continue

            coop = None
            if e.cooperado_id:
                coop = (
                    Cooperado.query
                    .join(Usuario, Cooperado.usuario_id == Usuario.id)
                    .filter(
                        Cooperado.id == e.cooperado_id,
                        Usuario.ativo.is_(True)
                    )
                    .first()
                )

            nome_fallback = (e.cooperado_nome or "").strip()
            nome_show = (coop.nome if coop else nome_fallback) or "—"
            contrato_eff = (eff_map.get(e.id, e.contrato or "") or "").strip()

            key = (
                (coop.id if coop else _norm(nome_show)),
                _norm(e.turno),
                _norm(e.horario),
                _norm(contrato_eff),
            )

            if key in seen[d]:
                break

            seen[d].add(key)

            agenda[d].append({
                "coop": coop,
                "cooperado_nome": nome_fallback or None,
                "nome_planilha": nome_show,
                "turno": (e.turno or "").strip(),
                "horario": (e.horario or "").strip(),
                "contrato": contrato_eff,
                "cor": (e.cor or "").strip(),
            })

            break

    for d in dias_list:
        agenda[d].sort(
            key=lambda x: (
                (x["contrato"] or "").lower(),
                (x.get("nome_planilha") or (x["coop"].nome if x["coop"] else "")).lower(),
            )
        )

    # -------------------- COOPERADOS ESCALADOS NO PERÍODO / HOJE --------------------
    hoje = date.today()

    ids_escalados_periodo = set()
    ids_escalados_hoje = set()
    nomes_escalados_sem_cadastro = set()

    for d in dias_list:
        for item in agenda.get(d, []):
            coop_item = item.get("coop")

            if coop_item and coop_item.id:
                ids_escalados_periodo.add(coop_item.id)

                if d == hoje:
                    ids_escalados_hoje.add(coop_item.id)
            else:
                nome_pl = (item.get("nome_planilha") or "").strip()
                if nome_pl:
                    nomes_escalados_sem_cadastro.add(nome_pl)

    cooperados_escalados = (
        Cooperado.query
        .join(Usuario, Cooperado.usuario_id == Usuario.id)
        .filter(
            Usuario.ativo.is_(True),
            Cooperado.id.in_(ids_escalados_periodo) if ids_escalados_periodo else literal(False)
        )
        .order_by(Cooperado.nome)
        .all()
    )

    # todos ativos, para busca manual no lançamento
    cooperados_ativos = (
        Cooperado.query
        .join(Usuario, Cooperado.usuario_id == Usuario.id)
        .filter(Usuario.ativo.is_(True))
        .order_by(Cooperado.nome)
        .all()
    )

    # marca quem está escalado no período e quem está escalado hoje
    for c in cooperados_ativos:
        c.escalado = c.id in ids_escalados_periodo
        c.escalado_hoje = c.id in ids_escalados_hoje

    # lista exibida no painel "lancar":
    # primeiro os escalados de hoje, depois os demais
    cooperados = sorted(
        cooperados_ativos,
        key=lambda c: (
            0 if getattr(c, "escalado_hoje", False) else 1,
            (c.nome or "").lower()
        )
    )
    # -------------------- LANÇAMENTOS / TOTAIS POR PERÍODO --------------------
    total_bruto = 0.0
    total_qtd = 0
    total_entregas = 0
    total_inss = 0.0
    total_sest = 0.0

    for c in cooperados:
        q = (
            Lancamento.query
            .filter_by(restaurante_id=rest.id, cooperado_id=c.id)
            .filter(Lancamento.data >= di, Lancamento.data <= df)
            .order_by(Lancamento.data.desc(), Lancamento.id.desc())
        )

        c.lancamentos = q.all()
        c.total_periodo = sum((l.valor or 0.0) for l in c.lancamentos)
        c.inss_periodo = sum((l.valor or 0.0) * 0.04 for l in c.lancamentos)
        c.sest_periodo = sum((l.valor or 0.0) * 0.005 for l in c.lancamentos)
        c.encargos_periodo = c.inss_periodo + c.sest_periodo
        c.liquido_periodo = c.total_periodo - c.encargos_periodo

        total_bruto += c.total_periodo
        total_qtd += len(c.lancamentos)
        total_entregas += sum((l.qtd_entregas or 0) for l in c.lancamentos)
        total_inss += c.inss_periodo
        total_sest += c.sest_periodo

    total_encargos = total_inss + total_sest
    total_liquido = total_bruto - total_encargos

    # -------------------- LISTA DE LANÇAMENTOS --------------------
    lancamentos_periodo = []
    total_lanc_valor = 0.0
    total_lanc_entregas = 0

    if view == "lancamentos":
        q = (
            db.session.query(Lancamento, Cooperado)
            .join(Cooperado, Cooperado.id == Lancamento.cooperado_id)
            .filter(
                Lancamento.restaurante_id == rest.id,
                Lancamento.data >= di,
                Lancamento.data <= df,
            )
            .order_by(Lancamento.data.asc(), Lancamento.id.asc())
        )

        for lanc, coop in q.all():
            item = {
                "id": lanc.id,
                "data": lanc.data.strftime("%d/%m/%Y") if lanc.data else "",
                "hora_inicio": (
                    lanc.hora_inicio if isinstance(lanc.hora_inicio, str)
                    else (lanc.hora_inicio.strftime("%H:%M") if lanc.hora_inicio else "")
                ),
                "hora_fim": (
                    lanc.hora_fim if isinstance(lanc.hora_fim, str)
                    else (lanc.hora_fim.strftime("%H:%M") if lanc.hora_fim else "")
                ),
                "qtd_entregas": lanc.qtd_entregas or 0,
                "valor": float(lanc.valor or 0.0),
                "descricao": (lanc.descricao or ""),
                "cooperado_id": coop.id,
                "cooperado_nome": coop.nome,
                "contrato_nome": rest.nome,
            }
            lancamentos_periodo.append(item)

        total_lanc_valor = sum(x["valor"] for x in lancamentos_periodo)
        total_lanc_entregas = sum(x["qtd_entregas"] for x in lancamentos_periodo)

    # -------------------- PENDÊNCIAS DE LANÇAMENTO DO DIA --------------------
    pendencias_lancamento = []
    hoje = date.today()
    agora = datetime.now()

    def _hora_inicial_min(horario_txt: str) -> int | None:
        m = re.search(r"(\d{1,2}):(\d{2})", str(horario_txt or ""))
        if not m:
            return None
        return int(m.group(1)) * 60 + int(m.group(2))

    def _hora_final_min(horario_txt: str) -> int | None:
        txt = str(horario_txt or "")
        pares = re.findall(r"(\d{1,2}):(\d{2})", txt)
        if len(pares) >= 2:
            hh, mm = pares[-1]
            return int(hh) * 60 + int(mm)

        m = re.search(r"\b(?:as|às|a)\s*(\d{1,2}):(\d{2})", txt.lower())
        if m:
            return int(m.group(1)) * 60 + int(m.group(2))

        return None

    minutos_agora = agora.hour * 60 + agora.minute

    escalas_hoje = agenda.get(hoje, [])
    for item in escalas_hoje:
        coop = item.get("coop")
        if not coop:
            continue

        horario_txt = (item.get("horario") or "").strip()
        turno_txt = (item.get("turno") or "").strip()
        contrato_txt = (item.get("contrato") or rest.nome).strip()

        hora_ini = _hora_inicial_min(horario_txt)
        hora_fim = _hora_final_min(horario_txt)

        if hora_fim is None and hora_ini is not None:
            if "noite" in turno_txt.lower():
                hora_fim = 23 * 60 + 59
            else:
                hora_fim = hora_ini + 240

        if hora_fim is None:
            continue

        if minutos_agora < hora_fim:
            continue

        lanc_do_dia = (
            Lancamento.query
            .filter(
                Lancamento.restaurante_id == rest.id,
                Lancamento.cooperado_id == coop.id,
                Lancamento.data == hoje,
            )
            .all()
        )

        existe_mesmo_horario = False
        for lanc in lanc_do_dia:
            hi = (lanc.hora_inicio or "").strip() if isinstance(lanc.hora_inicio, str) else (
                lanc.hora_inicio.strftime("%H:%M") if lanc.hora_inicio else ""
            )
            hf = (lanc.hora_fim or "").strip() if isinstance(lanc.hora_fim, str) else (
                lanc.hora_fim.strftime("%H:%M") if lanc.hora_fim else ""
            )

            if hi and hf and horario_txt:
                if hi in horario_txt and hf in horario_txt:
                    existe_mesmo_horario = True
                    break

            if hi and not hf and horario_txt and hi in horario_txt:
                existe_mesmo_horario = True
                break

        if not existe_mesmo_horario:
            pendencias_lancamento.append({
                "cooperado_id": coop.id,
                "cooperado_nome": coop.nome,
                "turno": turno_txt or "—",
                "horario": horario_txt or "—",
                "contrato": contrato_txt or "—",
                "data": hoje.strftime("%d/%m/%Y"),
            })

    pendencias_lancamento.sort(
        key=lambda x: (
            x["cooperado_nome"].lower(),
            x["horario"].lower(),
            x["turno"].lower(),
        )
    )

    # -------------------- URLs auxiliares --------------------
    try:
        url_lancar_producao = url_for("lancar_producao")
    except BuildError:
        url_lancar_producao = "/restaurante/lancar_producao"

    has_editar_lanc = ("editar_lancamento" in app.view_functions)

    # -------------------- Render --------------------
    return render_template(
        "restaurante_dashboard.html",
        rest=rest,
        cooperados=cooperados,
        cooperados_escalados=cooperados_escalados,
        pendencias_lancamento=pendencias_lancamento,
        filtro_inicio=di,
        filtro_fim=df,
        filtro_mes=(mes or ""),
        periodo_desc=periodo_desc,
        total_bruto=total_bruto,
        total_inss=total_inss,
        total_sest=total_sest,
        total_encargos=total_encargos,
        total_liquido=total_liquido,
        total_qtd=total_qtd,
        total_entregas=total_entregas,
        view=view,
        agenda=agenda,
        dias_list=dias_list,
        ref_data=ref,
        modo=modo,
        lancamentos_periodo=(lancamentos_periodo if view == "lancamentos" else []),
        total_lanc_valor=total_lanc_valor,
        total_lanc_entregas=total_lanc_entregas,
        url_lancar_producao=url_lancar_producao,
        has_editar_lanc=has_editar_lanc,
    )
    # =====================================================
    # COOPERADOS ESCALADOS HOJE PARA ESTE RESTAURANTE
    # =====================================================
    hoje = date.today()

    escalados_hoje = []
    escalados_ids = set()

    escalas_hoje_rest = agenda.get(hoje, []) if "agenda" in locals() else []

    for item in escalas_hoje_rest:
        coop_obj = item.get("coop")
        if coop_obj and coop_obj.id not in escalados_ids:
            escalados_ids.add(coop_obj.id)
            escalados_hoje.append(coop_obj)

    escalados_hoje = sorted(escalados_hoje, key=lambda c: (c.nome or "").lower())

    # lista completa apenas para busca manual
    cooperados_busca_manual = (
        Cooperado.query
        .join(Usuario, Cooperado.usuario_id == Usuario.id)
        .filter(Usuario.ativo.is_(True))
        .order_by(Cooperado.nome)
        .all()
    )

    # =====================================================
    # HELPERS DE HORÁRIO / PENDÊNCIAS
    # =====================================================
    def _parse_hora_min(hs: str | None):
        s = (hs or "").strip().lower()
        if not s:
            return None

        s = s.replace("h", ":")

        m = re.search(r"(\d{1,2})(?::(\d{2}))?", s)
        if not m:
            return None

        hh = int(m.group(1))
        mm = int(m.group(2) or 0)

        if hh < 0 or hh > 23 or mm < 0 or mm > 59:
            return None

        return hh * 60 + mm

    def _extrair_inicio_fim_intervalo(horario_txt: str | None):
        s = (horario_txt or "").strip().lower()
        if not s:
            return (None, None)

        s = s.replace("às", "as")
        s = s.replace("á", "a")

        partes = re.split(r"\s+as\s+|\s*-\s*|\s*a\s*", s)
        partes = [p.strip() for p in partes if p.strip()]

        if len(partes) >= 2:
            ini = _parse_hora_min(partes[0])
            fim = _parse_hora_min(partes[1])
            return ini, fim

        unico = _parse_hora_min(s)
        return unico, None

    # =====================================================
    # LANÇAMENTOS PENDENTES DO DIA
    # =====================================================
    agora = datetime.now()
    agora_min = agora.hour * 60 + agora.minute

    lancamentos_hoje_rest = (
        Lancamento.query
        .filter(
            Lancamento.restaurante_id == rest.id,
            Lancamento.data == hoje
        )
        .order_by(Lancamento.cooperado_id.asc(), Lancamento.id.asc())
        .all()
    )

    lancs_por_coop_hoje = defaultdict(list)
    for l in lancamentos_hoje_rest:
        lancs_por_coop_hoje[l.cooperado_id].append(l)

    lancamentos_pendentes = []
    pendentes_chaves = set()

    for item in escalas_hoje_rest:
        coop_obj = item.get("coop")
        if not coop_obj:
            continue

        horario_txt = (item.get("horario") or "").strip()
        turno_txt = (item.get("turno") or "").strip()
        contrato_txt = (item.get("contrato") or "").strip()

        ini_min, fim_min = _extrair_inicio_fim_intervalo(horario_txt)
        referencia_fim = fim_min if fim_min is not None else ini_min

        if referencia_fim is None:
            continue

        if agora_min < referencia_fim:
            continue

        escalas_vencidas_do_coop = []
        for x in escalas_hoje_rest:
            x_coop = x.get("coop")
            if not x_coop or x_coop.id != coop_obj.id:
                continue

            x_ini, x_fim = _extrair_inicio_fim_intervalo(x.get("horario"))
            x_ref = x_fim if x_fim is not None else x_ini

            if x_ref is not None and agora_min >= x_ref:
                escalas_vencidas_do_coop.append(x)

        qtd_vencidas = len(escalas_vencidas_do_coop)
        qtd_lancadas = len(lancs_por_coop_hoje.get(coop_obj.id, []))

        if qtd_lancadas < qtd_vencidas:
            chave_pend = (
                coop_obj.id,
                turno_txt.lower(),
                horario_txt.lower(),
                contrato_txt.lower(),
            )

            if chave_pend not in pendentes_chaves:
                pendentes_chaves.add(chave_pend)
                lancamentos_pendentes.append({
                    "chave": chave_pend,
                    "cooperado_id": coop_obj.id,
                    "cooperado_nome": coop_obj.nome,
                    "turno": turno_txt,
                    "horario": horario_txt,
                    "contrato": contrato_txt,
                    "fim_min": referencia_fim,
                })

    lancamentos_pendentes.sort(
        key=lambda x: (x["fim_min"], (x["cooperado_nome"] or "").lower())
    )

    # -------------------- Lista de lançamentos (aba "lancamentos") --------------------
    lancamentos_periodo = []
    total_lanc_valor = 0.0
    total_lanc_entregas = 0

    if view == "lancamentos":
        q = (
            db.session.query(Lancamento, Cooperado)
            .join(Cooperado, Cooperado.id == Lancamento.cooperado_id)
            .filter(
                Lancamento.restaurante_id == rest.id,
                Lancamento.data >= di,
                Lancamento.data <= df,
            )
            .order_by(Lancamento.data.asc(), Lancamento.id.asc())
        )
        for lanc, coop in q.all():
            item = {
                "id": lanc.id,
                "data": lanc.data.strftime("%d/%m/%Y") if lanc.data else "",
                "hora_inicio": (
                    lanc.hora_inicio if isinstance(lanc.hora_inicio, str)
                    else (lanc.hora_inicio.strftime("%H:%M") if lanc.hora_inicio else "")
                ),
                "hora_fim": (
                    lanc.hora_fim if isinstance(lanc.hora_fim, str)
                    else (lanc.hora_fim.strftime("%H:%M") if lanc.hora_fim else "")
                ),
                "qtd_entregas": lanc.qtd_entregas or 0,
                "valor": float(lanc.valor or 0.0),
                "descricao": (lanc.descricao or ""),
                "cooperado_id": coop.id,
                "cooperado_nome": coop.nome,
                "contrato_nome": rest.nome,
            }
            lancamentos_periodo.append(item)

        total_lanc_valor = sum(x["valor"] for x in lancamentos_periodo)
        total_lanc_entregas = sum(x["qtd_entregas"] for x in lancamentos_periodo)

    # ---- URLs/flags para template
    from werkzeug.routing import BuildError
    try:
        url_lancar_producao = url_for("lancar_producao")
    except BuildError:
        url_lancar_producao = "/restaurante/lancar_producao"

    has_editar_lanc = ("editar_lancamento" in app.view_functions)

    # -------------------- Render --------------------
    return render_template(
        "restaurante_dashboard.html",
        rest=rest,
        cooperados=cooperados,
        filtro_inicio=di,
        filtro_fim=df,
        filtro_mes=(mes or ""),
        periodo_desc=periodo_desc,
        total_bruto=total_bruto,
        total_inss=total_inss,
        total_sest=total_sest,
        total_encargos=total_encargos,
        total_liquido=total_liquido,
        total_qtd=total_qtd,
        total_entregas=total_entregas,
        view=view,
        agenda=agenda,
        dias_list=dias_list,
        ref_data=ref,
        modo=modo,
        lancamentos_periodo=(lancamentos_periodo if view == "lancamentos" else []),
        total_lanc_valor=total_lanc_valor,
        total_lanc_entregas=total_lanc_entregas,
        url_lancar_producao=url_lancar_producao,
        has_editar_lanc=has_editar_lanc,
        escalados_hoje=escalados_hoje,
        cooperados_busca_manual=cooperados_busca_manual,
        lancamentos_pendentes=lancamentos_pendentes,
        hoje=hoje,
    )

# =========================
# Rotas de CRUD de lançamento
# =========================
@app.post("/restaurante/lancar_producao")
@role_required("restaurante")
def lancar_producao():
    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first()
    if not rest:
        abort(403)
    f = request.form

    # NOVO: captura a descrição do formulário (somente para o estabelecimento)
    desc_raw = (f.get("descricao") or "").strip()
    desc_val = desc_raw or None  # salva None se vazio

    # 1) cria o lançamento
    l = Lancamento(
        restaurante_id=rest.id,
        cooperado_id=f.get("cooperado_id", type=int),
        descricao=desc_val,                            # <<< NOVO: salvar descrição
        valor=f.get("valor", type=float),
        data=_parse_date(f.get("data")) or date.today(),
        hora_inicio=f.get("hora_inicio"),
        hora_fim=f.get("hora_fim"),
        qtd_entregas=f.get("qtd_entregas", type=int),
    )
    db.session.add(l)
    db.session.flush()  # garante l.id

    # 2) avaliação (opcional)
    g   = _clamp_star(f.get("av_geral"))
    p   = _clamp_star(f.get("av_pontualidade"))
    ed  = _clamp_star(f.get("av_educacao"))
    ef  = _clamp_star(f.get("av_eficiencia"))
    ap  = _clamp_star(f.get("av_apresentacao"))
    txt = (f.get("av_comentario") or "").strip()

    tem_avaliacao = any(x is not None for x in (g, p, ed, ef, ap)) or bool(txt)
    if tem_avaliacao:
        media = _media_ponderada(g, p, ed, ef, ap)
        senti = _analise_sentimento(txt)
        temas = _identifica_temas(txt)
        crise = _sinaliza_crise(g, txt)
        feed  = _gerar_feedback(p, ed, ef, ap, txt, senti)

        av = AvaliacaoCooperado(
            restaurante_id=rest.id,
            cooperado_id=l.cooperado_id,
            lancamento_id=l.id,
            estrelas_geral=g,
            estrelas_pontualidade=p,
            estrelas_educacao=ed,
            estrelas_eficiencia=ef,
            estrelas_apresentacao=ap,
            comentario=txt,
            media_ponderada=media,
            sentimento=senti,
            temas="; ".join(temas),
            alerta_crise=crise,
            feedback_motoboy=feed,
        )
        db.session.add(av)
        if crise:
            flash("⚠️ Avaliação crítica registrada (1★ + termo de risco). A gerência deve revisar.", "danger")

    db.session.commit()
    flash("Produção lançada" + (" + avaliação salva." if tem_avaliacao else "."), "success")
    return redirect(url_for("portal_restaurante", view="lancar"))

@app.route("/lancamentos/<int:id>/editar", methods=["GET", "POST"])
@role_required("restaurante")
def editar_lancamento(id):
    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first()
    l = Lancamento.query.get_or_404(id)
    if not rest or l.restaurante_id != rest.id:
        abort(403)

    if request.method == "POST":
        f = request.form
        l.valor = f.get("valor", type=float)
        l.data = _parse_date(f.get("data")) or l.data
        l.hora_inicio = f.get("hora_inicio")
        l.hora_fim = f.get("hora_fim")
        l.qtd_entregas = f.get("qtd_entregas", type=int)
        # NOVO: permitir atualizar descrição se o form de edição trouxer o campo
        if "descricao" in f:
            l.descricao = (f.get("descricao") or "").strip() or None
        db.session.commit()
        flash("Lançamento atualizado.", "success")
        return redirect(url_for("portal_restaurante", view="lancamentos",
                                data_inicio=(l.data and l.data.strftime("%Y-%m-%d"))))

    return render_template("editar_lancamento.html", lanc=l)

@app.get("/lancamentos/<int:id>/excluir")
@role_required("restaurante")
def excluir_lancamento(id):
    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first()
    l = Lancamento.query.get_or_404(id)
    if not rest or l.restaurante_id != rest.id:
        abort(403)

    db.session.execute(sa_delete(AvaliacaoCooperado).where(AvaliacaoCooperado.lancamento_id == id))
    db.session.execute(sa_delete(AvaliacaoRestaurante).where(AvaliacaoRestaurante.lancamento_id == id))
    db.session.delete(l)
    db.session.commit()
    flash("Lançamento excluído.", "success")
    return redirect(url_for("portal_restaurante", view="lancamentos"))

# =========================
# Compat: Restaurante Avisos
# =========================
@app.get("/api/rest/avisos/unread_count")
@role_required("restaurante")
def api_rest_avisos_unread_count():
    def _nocache(resp):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        return resp

    try:
        u_id = session.get("user_id")
        rest = Restaurante.query.filter_by(usuario_id=u_id).first()

        if not rest:
            return _nocache(
                jsonify(
                    ok=False,
                    unread=0,
                    count=0,
                    error="Restaurante não encontrado"
                )
            ), 404

        avisos_rest = get_avisos_for_restaurante(rest)

        lidos_ids_rest = {
            row[0]
            for row in db.session.query(AvisoLeitura.aviso_id)
            .filter_by(restaurante_id=rest.id)
            .all()
        }

        avisos_nao_lidos_count = sum(
            1 for a in avisos_rest if a.id not in lidos_ids_rest
        )

        return _nocache(
            jsonify(
                ok=True,
                unread=int(avisos_nao_lidos_count),
                count=int(avisos_nao_lidos_count)
            )
        ), 200

    except Exception as e:
        db.session.rollback()

        try:
            current_app.logger.exception(
                "Erro ao calcular /api/rest/avisos/unread_count"
            )
        except Exception:
            pass

        return _nocache(
            jsonify(
                ok=False,
                unread=0,
                count=0,
                error=str(e)
            )
        ), 500

# =========================
# Documentos (Admin + Público)
# =========================
@app.route("/admin/documentos")
@admin_perm_required("documentos", "ver")
def admin_documentos():
    documentos = Documento.query.order_by(Documento.enviado_em.desc()).all()
    return render_template("admin_documentos.html", documentos=documentos)


@app.post("/admin/documentos/upload")
@admin_perm_required("documentos", "criar")
def admin_upload_documento():
    f = request.form
    titulo = (f.get("titulo") or "").strip()
    categoria = (f.get("categoria") or "outro").strip()
    descricao = (f.get("descricao") or "").strip()
    arquivo = request.files.get("arquivo")

    if not titulo or not (arquivo and arquivo.filename):
        flash("Preencha o título e selecione o arquivo.", "warning")
        return redirect(url_for("admin_documentos"))

    nome_unico = salvar_documento_upload(arquivo)
    if not nome_unico:
        flash("Falha ao salvar o arquivo.", "danger")
        return redirect(url_for("admin_documentos"))

    d = Documento(
        titulo=titulo,
        categoria=categoria,
        descricao=descricao,
        arquivo_url=url_for("serve_documento", nome=nome_unico),
        arquivo_nome=nome_unico,
        enviado_em=datetime.utcnow(),
    )
    db.session.add(d)
    db.session.commit()

    flash("Documento enviado.", "success")
    return redirect(url_for("admin_documentos"))


@app.get("/admin/documentos/<int:doc_id>/delete")
@admin_perm_required("documentos", "excluir")
def admin_delete_documento(doc_id):
    d = Documento.query.get_or_404(doc_id)

    try:
        p = resolve_documento_path(d.arquivo_nome)
        if p and os.path.exists(p):
            os.remove(p)

        if d.arquivo_url and d.arquivo_url.startswith("/static/uploads/docs/"):
            legacy_path = os.path.join(BASE_DIR, d.arquivo_url.lstrip("/"))
            if os.path.exists(legacy_path):
                os.remove(legacy_path)
    except Exception:
        pass

    db.session.delete(d)
    db.session.commit()
    flash("Documento removido.", "success")
    return redirect(url_for("admin_documentos"))


@app.route("/documentos")
def documentos_publicos():
    uid = session.get("user_id")
    if not uid:
        return redirect(url_for("login"))

    documentos = Documento.query.order_by(Documento.enviado_em.desc()).all()
    return render_template("documentos_publicos.html", documentos=documentos)


@app.route("/documentos/<int:doc_id>/baixar")
def baixar_documento(doc_id):
    doc = Documento.query.get_or_404(doc_id)

    path = resolve_documento_path(doc.arquivo_nome)
    if not path or not os.path.exists(path):
        abort(404)

    return send_file(
        path,
        as_attachment=True,
        download_name=os.path.basename(doc.arquivo_nome),
    )


# =========================
# Inicialização automática do DB em servidores (Gunicorn/Render)
# =========================
try:
    with app.app_context():
        init_db()
except Exception as _e:
    try:
        app.logger.warning(f"Falha ao inicializar DB: {_e}")
    except Exception:
        pass


@app.errorhandler(413)
def too_large(e):
    flash("Arquivo excede o tamanho máximo permitido (32MB).", "danger")
    return redirect(url_for("admin_documentos"))


# ==== TABELAS (upload/abrir/baixar) =========================================
from pathlib import Path
import logging

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _tabelas_base_dir() -> Path:
    """
    Usa o diretório persistente já definido no topo do app:
    TABELAS_DIR = os.path.join(PERSIST_ROOT, "tabelas")
    """
    p = Path(TABELAS_DIR)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _norm_txt(s: str) -> str:
    s = unicodedata.normalize("NFD", (s or "").strip())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def _guess_mimetype_from_path(path: str) -> str:
    mt, _ = mimetypes.guess_type(path)
    return mt or "application/octet-stream"


def _enforce_restaurante_titulo(tabela, restaurante):
    """
    Regra: restaurante só acessa a tabela cujo TÍTULO == NOME/LOGIN do restaurante.
    """
    login_nome = (
        getattr(getattr(restaurante, "usuario_ref", None), "usuario", None)
        or getattr(restaurante, "usuario", None)
        or (restaurante.nome or "")
    )
    if _norm_txt(tabela.titulo) != _norm_txt(login_nome):
        abort(403)


def _serve_tabela_or_redirect(tabela, *, as_attachment: bool):
    """
    Resolve e serve o arquivo da Tabela:
    - http(s) => redirect
    - prioriza diretório persistente
    - aceita nome simples, relativo e absoluto
    - faz fallback para diretório legado static/uploads/tabelas
    """
    url = (tabela.arquivo_url or "").strip()
    if not url:
        abort(404)

    if url.startswith(("http://", "https://")):
        return redirect(url)

    tabelas_dir = _tabelas_base_dir()
    base_dir = Path(BASE_DIR)

    raw = url.lstrip("/")
    raw_no_q = raw.split("?", 1)[0].split("#", 1)[0]
    fname = (raw_no_q.split("/")[-1] if raw_no_q else "").strip()

    candidates = []

    # 1) persistente
    if fname:
        candidates.append(tabelas_dir / fname)

    # 2) legado static/uploads/tabelas
    if fname:
        candidates.append(Path(STATIC_TABLES) / fname)

    # 3) relativo ao projeto
    if raw_no_q:
        candidates.append(base_dir / raw_no_q)

    # 4) absoluto
    p = Path(url)
    if p.is_absolute():
        candidates.append(p)

    # 5) legados extras
    if fname:
        candidates.append(base_dir / "uploads" / "tabelas" / fname)
        candidates.append(base_dir / "static" / "uploads" / "tabelas" / fname)

    file_path = next((c for c in candidates if c.exists() and c.is_file()), None)

    if not file_path:
        try:
            log.warning(
                "Arquivo de Tabela não encontrado. id=%s titulo=%r arquivo_url=%r tents=%r",
                getattr(tabela, "id", None),
                getattr(tabela, "titulo", None),
                tabela.arquivo_url,
                [str(c) for c in candidates],
            )
        except Exception:
            pass
        abort(404)

    return send_file(
        str(file_path),
        as_attachment=as_attachment,
        download_name=(tabela.arquivo_nome or file_path.name),
        mimetype=_guess_mimetype_from_path(str(file_path)),
    )


# ---------------------------------------------------------------------------
# Admin: listar / upload / delete
# ---------------------------------------------------------------------------
@app.get("/admin/tabelas", endpoint="admin_tabelas")
@admin_perm_required("tabelas", "ver")
def admin_tabelas():
    tabelas = Tabela.query.order_by(Tabela.enviado_em.desc(), Tabela.id.desc()).all()
    return render_template("admin_tabelas.html", tabelas=tabelas)


@app.post("/admin/tabelas/upload", endpoint="admin_upload_tabela")
@admin_perm_required("tabelas", "criar")
def admin_upload_tabela():
    f = request.form
    titulo = (f.get("titulo") or "").strip()
    descricao = (f.get("descricao") or "").strip() or None

    arquivo = (
        request.files.get("arquivo")
        or request.files.get("file")
        or request.files.get("tabela")
    )

    if not titulo or not (arquivo and arquivo.filename):
        flash("Preencha o título e selecione o arquivo.", "warning")
        return redirect(url_for("admin_tabelas"))

    base_dir = _tabelas_base_dir()

    raw = secure_filename(arquivo.filename)
    stem, ext = os.path.splitext(raw)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "-", stem) or "arquivo"
    final_name = f"{safe_stem}_{ts}{ext or ''}"

    dest = base_dir / final_name
    arquivo.save(str(dest))

    t = Tabela(
        titulo=titulo,
        descricao=descricao,
        arquivo_url=final_name,
        arquivo_nome=arquivo.filename,
        enviado_em=datetime.utcnow(),
    )
    db.session.add(t)
    db.session.commit()

    flash("Tabela publicada.", "success")
    return redirect(url_for("admin_tabelas"))


@app.get("/admin/tabelas/<int:tab_id>/delete", endpoint="admin_delete_tabela")
@admin_perm_required("tabelas", "excluir")
def admin_delete_tabela(tab_id: int):
    t = Tabela.query.get_or_404(tab_id)

    try:
        url = (t.arquivo_url or "").strip()
        if url and not url.startswith(("http://", "https://")):
            fname = url.split("?", 1)[0].split("#", 1)[0].split("/")[-1]

            persistent_path = _tabelas_base_dir() / fname
            legacy_path = Path(STATIC_TABLES) / fname

            if persistent_path.exists():
                persistent_path.unlink()
            elif legacy_path.exists():
                legacy_path.unlink()
    except Exception:
        pass

    db.session.delete(t)
    db.session.commit()
    flash("Tabela excluída.", "success")
    return redirect(url_for("admin_tabelas"))


# ---------------------------------------------------------------------------
# Cooperado/Admin/Restaurante: listagem
# ---------------------------------------------------------------------------
@app.get("/tabelas", endpoint="tabelas_publicas")
def tabelas_publicas():
    if session.get("user_tipo") not in {"admin", "cooperado", "restaurante"}:
        return redirect(url_for("login"))

    tabs = Tabela.query.order_by(Tabela.enviado_em.desc(), Tabela.id.desc()).all()

    items = [{
        "id": t.id,
        "titulo": t.titulo,
        "descricao": getattr(t, "descricao", None),
        "enviado_em": t.enviado_em,
        "arquivo_nome": getattr(t, "arquivo_nome", None),
        "abrir_url": url_for("tabela_abrir", tab_id=t.id),
        "baixar_url": url_for("baixar_tabela", tab_id=t.id),
    } for t in tabs]

    back_href = url_for("portal_cooperado") if (
        session.get("user_tipo") == "cooperado" and "portal_cooperado" in current_app.view_functions
    ) else ""

    return render_template(
        "tabelas_publicas.html",
        tabelas=tabs,
        items=items,
        back_href=back_href,
    )


# ---------------------------------------------------------------------------
# Abrir / Baixar compartilhado
# ---------------------------------------------------------------------------
@app.get("/tabelas/<int:tab_id>/abrir", endpoint="tabela_abrir")
def tabela_abrir(tab_id: int):
    if session.get("user_tipo") not in {"admin", "cooperado", "restaurante"}:
        return redirect(url_for("login"))

    t = Tabela.query.get_or_404(tab_id)

    if session.get("user_tipo") == "restaurante":
        rest = Restaurante.query.filter_by(usuario_id=session.get("user_id")).first_or_404()
        _enforce_restaurante_titulo(t, rest)

    return _serve_tabela_or_redirect(t, as_attachment=False)


@app.get("/tabelas/<int:tab_id>/baixar", endpoint="baixar_tabela")
def baixar_tabela(tab_id: int):
    if session.get("user_tipo") not in {"admin", "cooperado", "restaurante"}:
        return redirect(url_for("login"))

    t = Tabela.query.get_or_404(tab_id)

    if session.get("user_tipo") == "restaurante":
        rest = Restaurante.query.filter_by(usuario_id=session.get("user_id")).first_or_404()
        _enforce_restaurante_titulo(t, rest)

    return _serve_tabela_or_redirect(t, as_attachment=True)


# ---------------------------------------------------------------------------
# Restaurante: vê/abre/baixa SOMENTE a própria tabela
# ---------------------------------------------------------------------------
@app.get("/rest/tabelas", endpoint="rest_tabelas")
def rest_tabelas():
    if session.get("user_tipo") != "restaurante":
        return redirect(url_for("login"))

    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first_or_404()

    login_nome = (
        getattr(getattr(rest, "usuario_ref", None), "usuario", None)
        or getattr(rest, "usuario", None)
        or (rest.nome or "")
    )
    alvo_norm = _norm_txt(login_nome)

    candidatos = Tabela.query.order_by(Tabela.enviado_em.desc()).all()
    tabela_exata = next((t for t in candidatos if _norm_txt(t.titulo) == alvo_norm), None)

    has_portal_restaurante = ("portal_restaurante" in current_app.view_functions)

    return render_template(
        "restaurantes_tabelas.html",
        restaurante=rest,
        login_nome=login_nome,
        tabela=tabela_exata,
        has_portal_restaurante=has_portal_restaurante,
        back_href=url_for("portal_restaurante") if has_portal_restaurante else url_for("rest_tabelas"),
        current_year=datetime.utcnow().year,
    )


@app.get("/rest/tabelas/<int:tabela_id>/abrir", endpoint="rest_tabela_abrir")
def rest_tabela_abrir(tabela_id: int):
    if session.get("user_tipo") != "restaurante":
        return redirect(url_for("login"))

    rest = Restaurante.query.filter_by(usuario_id=session.get("user_id")).first_or_404()
    t = Tabela.query.get_or_404(tabela_id)
    _enforce_restaurante_titulo(t, rest)

    return _serve_tabela_or_redirect(t, as_attachment=False)


@app.get("/rest/tabelas/<int:tabela_id>/download", endpoint="rest_tabela_download")
def rest_tabela_download(tabela_id: int):
    if session.get("user_tipo") != "restaurante":
        return redirect(url_for("login"))

    rest = Restaurante.query.filter_by(usuario_id=session.get("user_id")).first_or_404()
    t = Tabela.query.get_or_404(tabela_id)
    _enforce_restaurante_titulo(t, rest)

    return _serve_tabela_or_redirect(t, as_attachment=True)


# ---------------------------------------------------------------------------
# Diagnóstico rápido (admin)
# ---------------------------------------------------------------------------
@app.get("/admin/tabelas/scan", endpoint="admin_tabelas_scan")
@admin_required
def admin_tabelas_scan():
    base = _tabelas_base_dir()
    items = []

    for t in Tabela.query.order_by(Tabela.enviado_em.desc()).all():
        url = (t.arquivo_url or "").strip()
        fname = (url.split("?", 1)[0].split("#", 1)[0]).split("/")[-1] if url else ""

        persistent_resolved = str(base / fname) if fname else ""
        legacy_resolved = str(Path(STATIC_TABLES) / fname) if fname else ""

        exists_persistent = Path(persistent_resolved).exists() if persistent_resolved else False
        exists_legacy = Path(legacy_resolved).exists() if legacy_resolved else False

        items.append({
            "id": t.id,
            "titulo": t.titulo,
            "arquivo_nome": t.arquivo_nome,
            "arquivo_url": t.arquivo_url,
            "persistent_resolved": persistent_resolved,
            "legacy_resolved": legacy_resolved,
            "exists_persistent": bool(exists_persistent),
            "exists_legacy": bool(exists_legacy),
        })

    return jsonify({
        "tabelas_dir_persistente": str(base),
        "tabelas_dir_legado": str(STATIC_TABLES),
        "items": items,
    })


# ---------------------------------------------------------------------------
# Normalizador: deixa arquivo_url só com o NOME do arquivo
# ---------------------------------------------------------------------------
@app.post("/admin/tabelas/normalize-arquivo-url", endpoint="admin_tabelas_normalize_arquivo_url")
@admin_required
def admin_tabelas_normalize_arquivo_url():
    alterados = 0

    for t in Tabela.query.all():
        url = (t.arquivo_url or "").strip()
        if not url or url.startswith(("http://", "https://")):
            continue

        fname = url.split("?", 1)[0].split("#", 1)[0].split("/")[-1]
        if fname and fname != url:
            t.arquivo_url = fname
            alterados += 1

    if alterados:
        db.session.commit()

    return jsonify({"ok": True, "alterados": alterados})

# =========================
# AVISOS — Ações (cooperado)
# =========================

@app.post("/avisos/<int:aviso_id>/lido", endpoint="marcar_aviso_lido")
@role_required("cooperado")
def marcar_aviso_lido(aviso_id: int):
    u_id = session.get("user_id")
    coop = Cooperado.query.filter_by(usuario_id=u_id).first_or_404()

    aviso = Aviso.query.get_or_404(aviso_id)

    ja_lido = AvisoLeitura.query.filter_by(
        cooperado_id=coop.id, aviso_id=aviso.id
    ).first()

    if not ja_lido:
        db.session.add(AvisoLeitura(
            cooperado_id=coop.id,
            aviso_id=aviso.id,
            lido_em=datetime.utcnow(),
        ))
        db.session.commit()

    # volta para a lista; se quiser voltar ancorado: + f"#aviso-{aviso.id}"
    return redirect(url_for("portal_cooperado_avisos"))

@app.post("/avisos/marcar-todos", endpoint="marcar_todos_avisos_lidos")
@role_required("cooperado")
def marcar_todos_avisos_lidos():
    u_id = session.get("user_id")
    coop = Cooperado.query.filter_by(usuario_id=u_id).first_or_404()

    # todos avisos visíveis ao cooperado
    avisos = get_avisos_for_cooperado(coop)

    # ids já lidos
    lidos_ids = {
        a_id for (a_id,) in db.session.query(AvisoLeitura.aviso_id)
        .filter(AvisoLeitura.cooperado_id == coop.id).all()
    }

    # persiste só os que faltam
    now = datetime.utcnow()
    for a in avisos:
        if a.id not in lidos_ids:
            db.session.add(AvisoLeitura(
                cooperado_id=coop.id,
                aviso_id=a.id,
                lido_em=now,
            ))

    db.session.commit()
    return redirect(url_for("portal_cooperado_avisos"))

@app.get("/portal/restaurante/avisos")
@role_required("restaurante")
def portal_restaurante_avisos():
    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first_or_404()

    # avisos aplicáveis
    try:
        avisos_db = get_avisos_for_restaurante(rest)
    except NameError:
        # fallback: global + restaurante (associados ou broadcast)
        avisos_db = (Aviso.query
                     .filter(Aviso.ativo.is_(True))
                     .filter(or_(Aviso.tipo == "global", Aviso.tipo == "restaurante"))
                     .order_by(Aviso.fixado.desc(), Aviso.criado_em.desc())
                     .all())

    # ids já lidos
    lidos_ids = {
        a_id for (a_id,) in db.session.query(AvisoLeitura.aviso_id)
        .filter(AvisoLeitura.restaurante_id == rest.id).all()
    }

    def corpo_do_aviso(a: Aviso) -> str:
        for k in ("corpo_html","html","conteudo_html","mensagem_html","descricao_html","texto_html",
                  "corpo","conteudo","mensagem","descricao","texto","resumo","body","content"):
            v = getattr(a, k, None)
            if isinstance(v, str) and v.strip():
                return v
        return ""

    avisos = [{
        "id": a.id,
        "titulo": a.titulo or "Aviso",
        "criado_em": a.criado_em,
        "lido": (a.id in lidos_ids),
        "prioridade_alta": (str(a.prioridade or "").lower() == "alta"),
        "corpo_html": corpo_do_aviso(a),
    } for a in avisos_db]

    avisos_nao_lidos_count = sum(1 for x in avisos if not x["lido"])
    return render_template(
        "portal_restaurante_avisos.html",   # crie/clone seu template
        avisos=avisos,
        avisos_nao_lidos_count=avisos_nao_lidos_count,
        current_year=datetime.now().year,
    )

@app.post("/avisos-restaurante/marcar-todos", endpoint="marcar_todos_avisos_lidos_restaurante")
@role_required("restaurante")
def marcar_todos_avisos_lidos_restaurante():
    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first_or_404()

    try:
        avisos = get_avisos_for_restaurante(rest)
    except NameError:
        avisos = (Aviso.query
                  .filter(Aviso.ativo.is_(True))
                  .filter(or_(Aviso.tipo == "global", Aviso.tipo == "restaurante"))
                  .all())

    lidos_ids = {
        a_id for (a_id,) in db.session.query(AvisoLeitura.aviso_id)
        .filter(AvisoLeitura.restaurante_id == rest.id).all()
    }

    now = datetime.utcnow()
    for a in avisos:
        if a.id not in lidos_ids:
            db.session.add(AvisoLeitura(
                restaurante_id=rest.id, aviso_id=a.id, lido_em=now
            ))
    db.session.commit()
    return redirect(url_for("portal_restaurante_avisos"))

# =========================
# Avisos: contagem de não lidos (Cooperado/Restaurante)
# =========================
def _nocache_json(payload: dict, status: int = 200):
    resp = jsonify(payload)
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp, status


@app.get("/avisos/unread_count")
@app.get("/avisos/unread_count/")
@with_db_retry
def avisos_unread_count():
    """
    Retorna a quantidade de avisos não lidos para o usuário logado.
    Responde sempre em JSON e evita quebrar o painel quando houver sessão vazia,
    usuário sem vínculo, ou qualquer falha na busca.
    """
    user_id = session.get("user_id")
    user_tipo = (session.get("user_tipo") or "").strip().lower()

    if not user_id:
        return _nocache_json({"ok": False, "unread": 0, "count": 0, "error": "Sessão ausente"}, 401)

    try:
        count = 0

        if user_tipo == "cooperado":
            coop = Cooperado.query.filter_by(usuario_id=user_id).first()
            if not coop:
                return _nocache_json({"ok": False, "unread": 0, "count": 0, "error": "Cooperado não encontrado"}, 404)

            avisos = get_avisos_for_cooperado(coop) or []
            lidos_ids = {
                aviso_id
                for (aviso_id,) in db.session.query(AvisoLeitura.aviso_id)
                .filter(AvisoLeitura.cooperado_id == coop.id)
                .all()
            }
            count = sum(1 for a in avisos if a.id not in lidos_ids)

        elif user_tipo == "restaurante":
            rest = Restaurante.query.filter_by(usuario_id=user_id).first()
            if not rest:
                return _nocache_json({"ok": False, "unread": 0, "count": 0, "error": "Restaurante não encontrado"}, 404)

            avisos = get_avisos_for_restaurante(rest) or []
            lidos_ids = {
                aviso_id
                for (aviso_id,) in db.session.query(AvisoLeitura.aviso_id)
                .filter(AvisoLeitura.restaurante_id == rest.id)
                .all()
            }
            count = sum(1 for a in avisos if a.id not in lidos_ids)

        else:
            return _nocache_json({"ok": True, "unread": 0, "count": 0, "error": "Tipo de usuário sem avisos"}, 200)

        return _nocache_json({"ok": True, "unread": int(count), "count": int(count)}, 200)

    except Exception as e:
        db.session.rollback()
        try:
            current_app.logger.exception("Erro ao calcular /avisos/unread_count")
        except Exception:
            pass
        return _nocache_json({"ok": False, "unread": 0, "count": 0, "error": str(e)}, 500)

import click

@app.cli.command("init-db")
def init_db_command():
    """Roda init_db() manualmente."""
    click.echo("Rodando init_db() ...")
    with app.app_context():
        init_db()
    click.echo("init_db() concluído.")

# =========================
# Main
# =========================
if __name__ == "__main__":
    with app.app_context():
        init_db()
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
    
